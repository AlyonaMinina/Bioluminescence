#!/usr/bin/env python3
"""
biolum_analysis.py — Bioluminescence Image Analysis Tool
Run on your laptop to analyze images from experiment folders.

Usage:
    python biolum_analysis.py

Then open: http://localhost:5001
"""

from flask import Flask, Response, request, jsonify, send_file
from pathlib import Path
from datetime import datetime
import numpy as np
import json
import io
import os
import webbrowser
import threading
import sys

app = Flask(__name__)

# ── helpers ───────────────────────────────────────────────────────────────────

def detect_image_type(stem):
    s = stem.lower()
    if any(kw in s for kw in ('biolum', 'bio', 'dark', 'night', 'lum')):
        return 'biolum'
    if any(kw in s for kw in ('day', 'light', 'white', 'bright')):
        return 'day'
    return 'unknown'

def find_pairs(folder):
    """Find JPEG+NEF pairs in a folder, grouped by stem."""
    folder = Path(folder)
    if not folder.exists():
        return []
    stems = {}
    for f in sorted(folder.iterdir()):
        if f.suffix.lower() in ('.jpg', '.jpeg', '.nef', '.raw'):
            stem = f.stem
            if stem not in stems:
                stems[stem] = {'stem': stem, 'jpg': None, 'nef': None,
                               'type': detect_image_type(stem)}
            if f.suffix.lower() in ('.jpg', '.jpeg'):
                stems[stem]['jpg'] = str(f)
            else:
                stems[stem]['nef'] = str(f)
    return [v for v in stems.values() if v['jpg'] or v['nef']]

def get_experiments(base_folder):
    """Return experiment subfolders."""
    base = Path(base_folder)
    if not base.exists():
        return []
    return [str(f) for f in sorted(base.iterdir(), reverse=True) if f.is_dir()]

def measure_nef(nef_path, rois, jpeg_size):
    """
    Measure integrated density from NEF for each ROI.
    rois: list of {x, y, w, h} in JPEG pixel coordinates
    jpeg_size: (width, height) of the JPEG used for ROI drawing
    Returns list of measurements per ROI.
    """
    import rawpy

    with rawpy.imread(nef_path) as raw:
        # postprocess to get RGB array (full resolution)
        rgb = raw.postprocess(
            output_bps=16,
            no_auto_bright=True,
            use_camera_wb=True
        )

    nef_h, nef_w = rgb.shape[:2]
    jpeg_w, jpeg_h = jpeg_size

    scale_x = nef_w / jpeg_w
    scale_y = nef_h / jpeg_h

    results = []
    for i, roi in enumerate(rois):
        # scale ROI coordinates to NEF resolution
        x1 = max(0, int(roi['x'] * scale_x))
        y1 = max(0, int(roi['y'] * scale_y))
        x2 = min(nef_w, int((roi['x'] + roi['w']) * scale_x))
        y2 = min(nef_h, int((roi['y'] + roi['h']) * scale_y))

        region = rgb[y1:y2, x1:x2]
        rh, rw = region.shape[:2]

        shape = roi.get('shape', 'rect')
        if shape == 'circle' and rh > 0 and rw > 0:
            cy_c, cx_c = rh / 2.0, rw / 2.0
            ry_c, rx_c = rh / 2.0, rw / 2.0
            Y, X = np.ogrid[:rh, :rw]
            mask = ((X - cx_c) / (rx_c + 1e-9))**2 + ((Y - cy_c) / (ry_c + 1e-9))**2 <= 1.0
        else:
            mask = np.ones((rh, rw), dtype=bool)

        area = int(np.sum(mask))

        r_ch = region[:, :, 0].astype(np.float64)
        g_ch = region[:, :, 1].astype(np.float64)
        b_ch = region[:, :, 2].astype(np.float64)

        results.append({
            'roi_number': i + 1,
            'roi_x_jpeg': roi['x'],
            'roi_y_jpeg': roi['y'],
            'roi_w_jpeg': roi['w'],
            'roi_h_jpeg': roi['h'],
            'roi_x_nef': x1,
            'roi_y_nef': y1,
            'roi_w_nef': x2 - x1,
            'roi_h_nef': y2 - y1,
            'area_px': area,
            'mean_R': float(np.mean(r_ch[mask])),
            'mean_G': float(np.mean(g_ch[mask])),
            'mean_B': float(np.mean(b_ch[mask])),
            'intden_R': float(np.sum(r_ch[mask])),
            'intden_G': float(np.sum(g_ch[mask])),
            'intden_B': float(np.sum(b_ch[mask])),
        })

    return results

def save_results(folder, stem, rois, measurements, jpeg_size, snapshot_png="", save_dir=""):
    """Save ROIs as JSON, snapshot PNG, and measurements as Excel."""
    if save_dir and Path(save_dir).exists():
        analysis_dir = Path(save_dir)
    else:
        analysis_dir = Path(folder) / f"Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}"
        analysis_dir.mkdir(parents=True, exist_ok=True)

    # Save ROIs
    roi_file = analysis_dir / f"{stem}_rois.json"
    with open(roi_file, 'w') as f:
        json.dump({'stem': stem, 'jpeg_size': jpeg_size, 'rois': rois}, f, indent=2)

    # Save snapshot PNG
    if snapshot_png and ',' in snapshot_png:
        import base64
        img_bytes = base64.b64decode(snapshot_png.split(',', 1)[1])
        with open(analysis_dir / f"{stem}_snapshot.png", 'wb') as f:
            f.write(img_bytes)

    if not measurements:
        return str(analysis_dir)

    # Excel
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Measurements"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_border = Border(bottom=Side(style='medium', color="2E74B5"))

    # non-null nb value means background was measured for that ROI
    has_nb = any(m.get('mean_B_nb') is not None for m in measurements)

    headers = ['Sample', 'ROI', 'Area (px)', 'Mean B', 'Mean G', 'Mean R']
    if has_nb:
        headers += ['Mean B-Bckg', 'Mean G-Bckg', 'Mean R-Bckg']

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = header_border

    fills = {
        'B':  PatternFill(start_color="E8EEFF", end_color="E8EEFF", fill_type="solid"),
        'G':  PatternFill(start_color="E8FFE8", end_color="E8FFE8", fill_type="solid"),
        'R':  PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"),
        'Bn': PatternFill(start_color="D0DDFF", end_color="D0DDFF", fill_type="solid"),
        'Gn': PatternFill(start_color="D0FFD8", end_color="D0FFD8", fill_type="solid"),
        'Rn': PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid"),
    }
    data_font = Font(name="Calibri")
    bckg_font = Font(name="Calibri", italic=True, color="8B6914")
    row_border = Border(bottom=Side(style='thin', color="D0D0D0"))

    def _r(v, d=1):
        return round(v, d) if isinstance(v, (int, float)) and v is not None else ''

    # col index (1-based) of first channel column
    ch_base = 4  # D = Mean B

    for m in measurements:
        is_bckg = m.get('roi_type') == 'bckg'
        sample = m.get('sample_name') or stem
        row = [sample, m.get('roi_label', ''), m.get('area_px', 0),
               _r(m.get('mean_B')), _r(m.get('mean_G')), _r(m.get('mean_R'))]
        if has_nb:
            row += [_r(m.get('mean_B_nb')), _r(m.get('mean_G_nb')), _r(m.get('mean_R_nb'))]
        ws.append(row)
        rn = ws.max_row
        for cell in ws[rn]:
            cell.font = bckg_font if is_bckg else data_font
            cell.border = row_border
        col_map = [(ch_base,'B'),(ch_base+1,'G'),(ch_base+2,'R')]
        if has_nb:
            col_map += [(ch_base+3,'Bn'),(ch_base+4,'Gn'),(ch_base+5,'Rn')]
        for col, ch in col_map:
            ws.cell(rn, col).fill = fills[ch]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    col_widths = [('A',28),('B',10),('C',10),('D',10),('E',10),('F',10)]
    if has_nb:
        col_widths += [('G',12),('H',12),('I',12)]
    for col, w in col_widths:
        ws.column_dimensions[col].width = w

    xlsx_file = analysis_dir / "session_measurements.xlsx"
    wb.save(xlsx_file)

    return str(analysis_dir)

# ── routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return HTML_PAGE

@app.route("/browse", methods=["POST"])
def browse():
    data = request.get_json()
    folder = data.get("folder", "")
    pairs = find_pairs(folder)
    analysis_dir = ""
    try:
        p = Path(folder)
        if p.exists():
            analysis_dir = str(p / f"Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}")
            Path(analysis_dir).mkdir(parents=True, exist_ok=True)
    except Exception:
        analysis_dir = ""
    return jsonify({"pairs": pairs, "folder": folder, "analysis_dir": analysis_dir})

@app.route("/ls", methods=["POST"])
def ls():
    """List directories and image files in a path for the folder navigator."""
    data = request.get_json()
    path = data.get("path", "")

    # default to home / drives on Windows
    if not path:
        if sys.platform == "win32":
            import string
            drives = [f"{d}:\\" for d in string.ascii_uppercase
                      if Path(f"{d}:\\").exists()]
            return jsonify({"path": "", "parent": None,
                            "dirs": [{"name": d, "path": d} for d in drives],
                            "has_images": False})
        else:
            path = str(Path.home())

    p = Path(path)
    if not p.exists() or not p.is_dir():
        return jsonify({"error": "Not a directory"}), 400

    try:
        dirs = []
        for item in sorted(p.iterdir()):
            try:
                if item.is_dir() and not item.name.startswith('.'):
                    dirs.append({"name": item.name, "path": str(item)})
            except PermissionError:
                pass

        # check if folder has images
        has_images = any(
            f.suffix.lower() in ('.jpg','.jpeg','.nef','.raw')
            for f in p.iterdir()
            if f.is_file()
        )

        parent = str(p.parent) if p.parent != p else None
        return jsonify({
            "path": str(p),
            "parent": parent,
            "dirs": dirs,
            "has_images": has_images
        })
    except PermissionError:
        return jsonify({"error": "Permission denied"}), 403

@app.route("/experiments", methods=["POST"])
def experiments():
    data = request.get_json()
    base = data.get("folder", "")
    exps = get_experiments(base)
    return jsonify({"experiments": exps})

@app.route("/image")
def serve_image():
    path = request.args.get("path", "")
    try:
        return send_file(path, mimetype="image/jpeg")
    except Exception:
        return "", 404

@app.route("/image_size")
def image_size():
    path = request.args.get("path", "")
    try:
        from PIL import Image
        with Image.open(path) as img:
            return jsonify({"width": img.width, "height": img.height})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/measure", methods=["POST"])
def measure():
    data = request.get_json()
    nef_path = data.get("nef_path", "")
    rois = data.get("rois", [])
    jpeg_size = data.get("jpeg_size", [1, 1])

    if not nef_path or not Path(nef_path).exists():
        return jsonify({"error": "NEF file not found"}), 400
    if not rois:
        return jsonify({"error": "No ROIs defined"}), 400

    try:
        results = measure_nef(nef_path, rois, jpeg_size)
        return jsonify({"measurements": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/save", methods=["POST"])
def save():
    data = request.get_json()
    folder = data.get("folder", "")
    save_dir = data.get("save_dir") or ""
    stem = data.get("stem", "analysis")
    rois = data.get("rois", [])
    measurements = data.get("measurements", [])
    session_measurements = data.get("session_measurements") or measurements
    jpeg_size = data.get("jpeg_size", [1, 1])
    snapshot_png = data.get("snapshot_png", "")

    try:
        out_dir = save_results(folder, stem, rois, session_measurements, jpeg_size,
                               snapshot_png, save_dir=save_dir)
        return jsonify({"ok": True, "dir": out_dir})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/save_pdf", methods=["POST"])
def save_pdf():
    import base64
    data = request.get_json()
    save_dir = data.get("dir", "")
    pdf_data = data.get("pdf_data", "")
    filename = data.get("filename", "biolum_summary.pdf")
    if not save_dir or not pdf_data:
        return jsonify({"ok": False, "error": "Missing dir or pdf_data"})
    try:
        p = Path(save_dir)
        p.mkdir(parents=True, exist_ok=True)
        raw = pdf_data.split(',', 1)[-1]
        pdf_bytes = base64.b64decode(raw)
        out_path = p / filename
        with open(out_path, 'wb') as f:
            f.write(pdf_bytes)
        return jsonify({"ok": True, "path": str(out_path)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/load_rois", methods=["POST"])
def load_rois():
    data = request.get_json()
    path = data.get("path", "")
    try:
        with open(path, 'r') as f:
            roi_data = json.load(f)
        return jsonify({"ok": True, "data": roi_data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/list_roi_files", methods=["POST"])
def list_roi_files():
    data = request.get_json()
    folder = data.get("folder", "")
    files = []
    for p in Path(folder).rglob("*_rois.json"):
        files.append({"name": p.name, "path": str(p)})
    return jsonify({"files": files})

@app.route("/download")
def download():
    path = request.args.get("path", "")
    try:
        return send_file(path, as_attachment=True)
    except Exception:
        return "", 404

@app.route("/pick_folder")
def pick_folder():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes('-topmost', True)
    folder = filedialog.askdirectory(title="Select Experiment Folder")
    root.destroy()
    return jsonify({"path": folder or ""})

# ── HTML ──────────────────────────────────────────────────────────────────────

HTML_PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Biolum Analysis</title>
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@300;400;500&family=IBM+Plex+Sans:wght@300;400;600&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    :root {
      --bg:      #06090d;
      --panel:   #0b0f15;
      --panel2:  #0f1520;
      --border:  #162030;
      --border2: #1e2d42;
      --accent:  #4af0c4;
      --amber:   #f0c44a;
      --red:     #f04a6a;
      --text:    #c0d8f0;
      --muted:   #385060;
      --mono:    'IBM Plex Mono', monospace;
      --sans:    'IBM Plex Sans', sans-serif;
    }

    html, body {
      height: 100%;
      background: var(--bg);
      color: var(--text);
      font-family: var(--sans);
      font-size: 13px;
      overflow: hidden;
    }

    /* ── layout ── */
    .app {
      display: grid;
      grid-template-rows: 48px 1fr;
      height: 100vh;
    }

    header {
      display: flex;
      align-items: center;
      gap: 16px;
      padding: 0 20px;
      border-bottom: 1px solid var(--border);
      background: var(--panel);
    }

    .brand {
      font-family: var(--mono);
      font-size: 13px;
      letter-spacing: 0.15em;
      color: var(--accent);
      text-transform: uppercase;
      flex-shrink: 0;
    }

    .header-path {
      font-family: var(--mono);
      font-size: 11px;
      color: var(--muted);
      flex: 1;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    .body {
      display: grid;
      grid-template-columns: 280px 1fr;
      overflow: hidden;
    }

    /* tabs-container fills the app grid's second row */
    .tabs-container { height: 100%; }

    /* ── sidebar ── */
    .sidebar {
      border-right: 1px solid var(--border);
      background: var(--panel);
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }

    .sidebar-section {
      border-bottom: 1px solid var(--border);
      flex-shrink: 0;
    }

    .sidebar-section-header {
      padding: 8px 14px;
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.18em;
      color: var(--muted);
      text-transform: uppercase;
      background: rgba(255,255,255,0.015);
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    .pair-list {
      flex: 1;
      overflow-y: auto;
      padding: 6px;
    }

    .dir-item {
      padding: 5px 10px;
      cursor: pointer;
      color: var(--text);
      display: flex;
      align-items: center;
      gap: 6px;
      border-bottom: 1px solid var(--border);
      transition: background 0.1s;
    }
    .dir-item:hover { background: rgba(74,240,196,0.05); }
    .dir-item.has-images { color: var(--accent); }
    .dir-item.parent { color: var(--amber); }
    .dir-item .dir-icon { font-size: 12px; flex-shrink: 0; }

    .pair-item {
      padding: 8px 10px;
      border-radius: 3px;
      cursor: pointer;
      font-family: var(--mono);
      font-size: 11px;
      color: var(--text);
      border: 1px solid transparent;
      transition: all 0.15s;
      margin-bottom: 3px;
      word-break: break-word;
      line-height: 1.5;
    }

    .pair-item:hover { background: rgba(74,240,196,0.05); border-color: var(--border2); }
    .pair-item.active-day    { background: rgba(240,196,74,0.08); border-color: var(--amber); color: var(--amber); }
    .pair-item.active-biolum { background: rgba(74,240,196,0.08); border-color: var(--accent); color: var(--accent); }

    .pair-badges {
      display: flex;
      gap: 4px;
      margin-top: 4px;
    }

    .badge {
      font-size: 9px;
      padding: 1px 5px;
      border-radius: 2px;
      letter-spacing: 0.08em;
      font-family: var(--mono);
    }
    .badge-jpg { background: rgba(74,240,196,0.15); color: var(--accent); }
    .badge-nef { background: rgba(240,196,74,0.15); color: var(--amber); }
    .badge-none { background: rgba(240,74,106,0.15); color: var(--red); }
    .badge-day { background: rgba(240,196,74,0.2); color: var(--amber); }
    .badge-biolum { background: rgba(74,240,196,0.2); color: var(--accent); }
    .badge-unknown { background: rgba(192,216,240,0.1); color: var(--muted); }

    /* ── folder input ── */
    .folder-input-wrap {
      padding: 10px;
      display: flex;
      flex-direction: column;
      gap: 6px;
    }

    input[type="text"] {
      width: 100%;
      background: var(--bg);
      border: 1px solid var(--border2);
      border-radius: 3px;
      color: var(--text);
      font-family: var(--mono);
      font-size: 11px;
      padding: 6px 10px;
      outline: none;
    }
    input[type="text"]:focus { border-color: var(--accent); }

    .btn {
      font-family: var(--mono);
      font-size: 11px;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      padding: 7px 12px;
      border-radius: 3px;
      border: 1px solid;
      cursor: pointer;
      transition: all 0.15s;
      width: 100%;
    }
    .btn-primary { background: #000; border-color: var(--accent); color: var(--accent); font-weight: 600; }
    .btn-primary:hover { background: rgba(74,240,196,0.08); }
    .btn-primary:disabled { opacity: 0.3; cursor: not-allowed; }
    .btn-ghost { background: transparent; border-color: var(--border2); color: var(--text); }
    .btn-ghost:hover { border-color: var(--accent); color: var(--accent); }
    .btn-amber { background: rgba(240,196,74,0.1); border-color: var(--amber); color: var(--amber); }
    .btn-amber:hover { background: rgba(240,196,74,0.2); }
    .btn-red { background: transparent; border-color: var(--red); color: var(--red); }
    .btn-red:hover { background: rgba(240,74,106,0.1); }
    .btn-sm { padding: 4px 10px; font-size: 10px; width: auto; }

    /* ── main panel ── */
    .main {
      display: grid;
      grid-template-rows: 1fr 220px;
      overflow: hidden;
    }

    /* ── canvas area ── */
    .canvas-area {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 2px;
      background: var(--border);
      overflow: hidden;
      position: relative;
    }

    .canvas-wrap {
      position: relative;
      background: #000;
      overflow: hidden;
      display: flex;
      align-items: center;
      justify-content: center;
    }

    .canvas-label {
      position: absolute;
      top: 10px;
      left: 50%;
      transform: translateX(-50%);
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.15em;
      color: #fff;
      background: rgba(6,9,13,0.85);
      padding: 3px 10px;
      border-radius: 2px;
      text-transform: uppercase;
      font-weight: 600;
      pointer-events: none;
      z-index: 10;
    }

    .canvas-clear-btn {
      position: absolute;
      top: 10px;
      right: 10px;
      background: rgba(240,74,106,0.15);
      border: 1px solid var(--red);
      color: var(--red);
      font-family: var(--mono);
      font-size: 10px;
      padding: 2px 8px;
      border-radius: 2px;
      cursor: pointer;
      z-index: 10;
      transition: background 0.15s;
    }
    .canvas-clear-btn:hover { background: rgba(240,74,106,0.35); }

    .canvas-sample {
      position: absolute;
      bottom: 10px;
      left: 50%;
      transform: translateX(-50%);
      font-family: var(--mono);
      font-size: 11px;
      color: var(--accent);
      background: rgba(6,9,13,0.85);
      padding: 3px 12px;
      border-radius: 2px;
      pointer-events: none;
      z-index: 10;
      white-space: nowrap;
      max-width: 90%;
      overflow: hidden;
      text-overflow: ellipsis;
    }

    canvas {
      max-width: 100%;
      max-height: 100%;
      cursor: crosshair;
      display: block;
    }

    .no-image {
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      gap: 8px;
      color: var(--muted);
      font-family: var(--mono);
      font-size: 11px;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      width: 100%;
      height: 100%;
    }

    /* ── bottom panel ── */
    .bottom-panel {
      border-top: 1px solid var(--border);
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }

    .roi-toolbar {
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 5px 12px;
      border-bottom: 1px solid var(--border);
      background: var(--panel);
      flex-shrink: 0;
      flex-wrap: wrap;
    }

    .roi-toolbar-label {
      font-family: var(--mono);
      font-size: 9px;
      letter-spacing: 0.18em;
      color: var(--muted);
      text-transform: uppercase;
      flex-shrink: 0;
    }

    .roi-mode-row {
      display: flex;
      gap: 4px;
      align-items: center;
    }

    .mode-btn {
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      padding: 4px 8px;
      border-radius: 3px;
      border: 1px solid var(--text);
      cursor: pointer;
      color: var(--text);
      background: transparent;
      transition: all 0.15s;
      text-align: center;
      white-space: nowrap;
    }
    .mode-btn.active { border-color: var(--accent); color: var(--accent); background: rgba(74,240,196,0.07); }
    .mode-btn:hover:not(.active) { background: rgba(255,255,255,0.04); }

    .checkbox-row {
      display: flex;
      align-items: center;
      gap: 6px;
      font-family: var(--mono);
      font-size: 11px;
      color: var(--text);
      cursor: pointer;
    }
    .checkbox-row input { accent-color: var(--accent); }

    /* ── results table ── */
    .results-area {
      overflow-y: auto;
      background: var(--panel2);
      padding: 8px 12px;
    }

    .results-header {
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.15em;
      color: var(--muted);
      text-transform: uppercase;
      margin-bottom: 8px;
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    table {
      width: 100%;
      border-collapse: collapse;
      font-family: var(--mono);
      font-size: 11px;
    }

    th {
      text-align: left;
      padding: 4px 8px;
      font-size: 10px;
      letter-spacing: 0.1em;
      color: var(--muted);
      text-transform: uppercase;
      border-bottom: 1px solid var(--border);
      white-space: nowrap;
    }

    td {
      padding: 4px 8px;
      border-bottom: 1px solid var(--border);
      color: #ddeeff;
      white-space: nowrap;
    }

    td.ch-r { color: #ff9999; }
    td.ch-g { color: #99ffbb; }
    td.ch-b { color: #99bbff; }
    td.ch-r-nb { color: #ffcccc; font-style: italic; }
    td.ch-g-nb { color: #ccffdd; font-style: italic; }
    td.ch-b-nb { color: #ccdeff; font-style: italic; }
    td.bckg-row { color: var(--amber); }

    tr:hover td { background: rgba(255,255,255,0.02); }

    .status-bar {
      font-family: var(--mono);
      font-size: 11px;
      color: var(--muted);
      padding: 4px 0;
    }
    .status-ok  { color: var(--accent); }
    .status-err { color: var(--red); }
    .status-wrn { color: var(--amber); }

    /* ── top-level view tabs ── */
    .top-tab-btn {
      font-family: var(--mono);
      font-size: 11px;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      padding: 0 18px;
      height: 100%;
      border: none;
      border-bottom: 2px solid transparent;
      background: transparent;
      color: var(--muted);
      cursor: pointer;
      transition: all 0.15s;
      flex-shrink: 0;
    }
    .top-tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); }
    .top-tab-btn:hover:not(.active) { color: var(--text); }

    .tabs-container { overflow: hidden; display: grid; }
    .top-tab-pane { display: none; overflow: hidden; }
    #top-tab-measure.active {
      display: grid;
      grid-template-columns: 280px 1fr;
    }
    #top-tab-summary.active {
      display: flex;
      flex-direction: column;
      background: var(--panel2);
      padding: 16px;
      gap: 12px;
    }
  </style>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js"></script>
</head>
<body>
<div class="app">

  <header>
    <div class="brand">◉ Biolum Analysis</div>
    <button class="top-tab-btn active" id="top-tab-btn-measure" onclick="switchTopTab('measure')">Measurement</button>
    <button class="top-tab-btn" id="top-tab-btn-summary" onclick="switchTopTab('summary')">Analysis Summary</button>
    <div class="header-path" id="header-path">No folder loaded</div>
    <div style="display:flex;gap:8px;flex-shrink:0;">
      <button class="btn btn-ghost btn-sm" onclick="measureAll()" id="btn-measure">▶ Measure ROIs</button>
      <button class="btn btn-amber btn-sm" onclick="saveResults()" id="btn-save">↓ Save Results</button>
      <button class="btn btn-ghost btn-sm" onclick="showLoadRoiDialog()">↑ Load ROIs</button>
    </div>
  </header>

  <div class="tabs-container">
  <div class="top-tab-pane active" id="top-tab-measure">

    <!-- sidebar -->
    <div class="sidebar">
      <div class="sidebar-section">
        <div class="sidebar-section-header" id="experiment-label" style="font-size:11px;color:var(--accent);">
          No experiment loaded
        </div>
        <div style="padding:6px 8px;border-bottom:1px solid var(--border);display:flex;flex-direction:column;gap:4px;">
          <button class="btn btn-primary btn-sm" style="width:100%;" onclick="pickFolder()">Browse...</button>
          <div style="display:flex;gap:4px;">
            <input type="text" id="folder-input" placeholder="Or paste path..."
                   style="flex:1;font-size:10px;padding:4px 8px;"
                   onkeydown="if(event.key==='Enter') loadFolder(this.value)">
            <button class="btn btn-ghost btn-sm" onclick="loadFolder(document.getElementById('folder-input').value)" style="flex-shrink:0;">Go</button>
          </div>
        </div>
      </div>
      <div class="sidebar-section" style="padding:8px 10px;">
        <div class="status-bar" id="status">Ready. Load an experiment folder to begin.</div>
      </div>
      <div class="sidebar-section-header" style="padding:8px 14px;font-size:11px;color:var(--text);letter-spacing:0.1em;">
        Images within experiment
        <span style="font-size:10px;color:var(--accent);" id="pair-count"></span>
      </div>
      <div class="pair-list" id="pair-list">
        <div style="padding:20px;text-align:center;font-family:var(--mono);font-size:11px;color:var(--muted);">
          No folder loaded
        </div>
      </div>
    </div>

    <!-- main -->
    <div class="main">

      <!-- canvases -->
      <div class="canvas-area">
        <div class="canvas-wrap" id="day-wrap">
          <div class="canvas-label">DAY</div>
          <div class="no-image" id="day-placeholder">
            <span style="font-size:24px;">☀</span>
            <span>Select a day image</span>
          </div>
          <canvas id="day-canvas" style="display:none"></canvas>
          <button class="canvas-clear-btn" id="day-clear-btn" onclick="clearPanel('day')" style="display:none;" title="Clear day image">✕ Clear</button>
          <div class="canvas-sample" id="day-sample-label" style="display:none;"></div>
        </div>
        <div class="canvas-wrap" id="biolum-wrap">
          <div class="canvas-label">BIOLUM</div>
          <div class="no-image" id="biolum-placeholder">
            <span style="font-size:24px;">◉</span>
            <span>Select a biolum image</span>
          </div>
          <canvas id="biolum-canvas" style="display:none"></canvas>
          <button class="canvas-clear-btn" id="biolum-clear-btn" onclick="clearPanel('biolum')" style="display:none;" title="Clear biolum image">✕ Clear</button>
          <div class="canvas-sample" id="biolum-sample-label" style="display:none;"></div>
        </div>
      </div>

      <!-- bottom panel -->
      <div class="bottom-panel">

        <!-- compact ROI toolbar -->
        <div class="roi-toolbar">
          <span class="roi-toolbar-label">ROI</span>
          <div class="roi-mode-row">
            <div class="mode-btn active" id="mode-draw" onclick="setMode('draw')">✚ Draw</div>
            <div class="mode-btn" id="mode-move" onclick="setMode('move')">✥ Move</div>
            <div class="mode-btn" id="mode-resize" onclick="setMode('resize')">⤢ Resize</div>
            <div class="mode-btn" id="mode-pan" onclick="setMode('pan')">✋ Pan</div>
          </div>
          <select id="roi-shape" onchange="roiShape=this.value"
                  style="font-family:var(--mono);font-size:10px;background:var(--bg);color:var(--text);border:1px solid var(--border2);border-radius:3px;padding:3px 6px;cursor:pointer;">
            <option value="rect">■ Rect</option>
            <option value="circle">● Circle</option>
          </select>
          <div class="mode-btn" id="type-bckg" onclick="toggleBckgType()" title="Toggle background ROI type">Bckg</div>
          <label class="checkbox-row" style="flex-shrink:0;">
            <input type="checkbox" id="fix-size" onchange="toggleFixSize()">
            Lock size
          </label>
          <label class="checkbox-row" style="flex-shrink:0;">
            <input type="checkbox" id="rename-mode" onchange="toggleRenameMode()">
            Rename ROIs
          </label>
          <button class="btn btn-ghost btn-sm" onclick="resetView()">⊡ Fit</button>
          <button class="btn btn-red btn-sm" onclick="deleteSelected()">✕ Del ROI</button>
          <button class="btn btn-ghost btn-sm" onclick="clearAllRois()">Clear all</button>
        </div>

        <!-- measurements -->
        <div class="results-area" id="tab-meas">
          <div class="results-header">
            <span>Measurements</span>
            <div style="display:flex;gap:6px;" id="download-links"></div>
          </div>
          <div id="results-table-wrap">
            <div style="font-family:var(--mono);font-size:11px;color:var(--muted);">
              Draw ROIs then click ▶ Measure ROIs
            </div>
          </div>
        </div>

      </div>
    </div>
  </div>

  <!-- Analysis Summary — full-size pane -->
  <div class="top-tab-pane" id="top-tab-summary">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-shrink:0;">
      <span style="font-family:var(--mono);font-size:11px;color:var(--muted);letter-spacing:0.15em;text-transform:uppercase;">Mean IntDen per sample — background subtracted</span>
      <button class="btn btn-sm" onclick="exportSummaryPDF()" style="margin-right:6px;">Export PDF</button>
      <button class="btn btn-red btn-sm" onclick="clearAnalysisData()">Clear data</button>
    </div>
    <div id="summary-empty" style="font-family:var(--mono);font-size:13px;color:var(--muted);text-align:center;padding:40px;">
      Measure samples in the Measurement tab to build the summary
    </div>
    <div style="display:flex;gap:16px;flex:1;min-height:0;" id="summary-plots">
      <canvas id="plot-b" style="flex:1;min-width:0;"></canvas>
      <canvas id="plot-g" style="flex:1;min-width:0;"></canvas>
      <canvas id="plot-r" style="flex:1;min-width:0;"></canvas>
    </div>
  </div>

  </div><!-- end .tabs-container -->
</div>

<!-- Load ROIs dialog -->
<div id="roi-dialog" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.7);z-index:1000;align-items:center;justify-content:center;">
  <div style="background:var(--panel);border:1px solid var(--border2);border-radius:4px;padding:20px;width:500px;max-height:60vh;display:flex;flex-direction:column;gap:12px;">
    <div style="font-family:var(--mono);font-size:12px;letter-spacing:0.1em;color:var(--accent);text-transform:uppercase;">Load Saved ROIs</div>
    <div id="roi-file-list" style="overflow-y:auto;flex:1;display:flex;flex-direction:column;gap:4px;"></div>
    <button class="btn btn-ghost" onclick="document.getElementById('roi-dialog').style.display='none'">Cancel</button>
  </div>
</div>

<script>
  // fix Windows backslashes for URL passing
  var _BS = String.fromCharCode(92);
  function fixPath(p) { return p ? p.split(_BS).join('/') : ''; }

  // ── state ──
  let dayPair = null;
  let biolumPair = null;
  let currentFolder = null;
  let jpegSize = {w: 1, h: 1};
  let rois = [];
  let selectedRoi = -1;
  let mode = 'draw';
  let roiShape = 'rect';
  let roiType = 'roi';
  let fixedSize = null;
  let isDrawing = false;
  let drawStart = {x:0, y:0};
  let dragStart = {x:0, y:0};
  let dragRoiStart = null;
  let measurements = [];
  let analysisData = [];
  let analysisDataUid = 0;
  let renameMode = false;
  let lastSaveDir = null;
  let sessionMeasurements = [];
  const SUMMARY_CHANNELS = [
    {id:'plot-b', key:'mean_B_nb', color:'#88bbff', colorExport:'#1030c0', label:'Mean B-Bckg'},
    {id:'plot-g', key:'mean_G_nb', color:'#55e888', colorExport:'#0e7a2e', label:'Mean G-Bckg'},
    {id:'plot-r', key:'mean_R_nb', color:'#ff9999', colorExport:'#b01010', label:'Mean R-Bckg'},
  ];
  let view = {scale: 1, dx: 0, dy: 0};
  let panStart = null;
  let dragRoiRef = null;

  // canvas refs
  const dayCanvas = document.getElementById('day-canvas');
  const dayCtx = dayCanvas.getContext('2d');
  const biolumCanvas = document.getElementById('biolum-canvas');
  const biolumCtx = biolumCanvas.getContext('2d');
  let dayImg = null;
  let biolumImg = null;

  // ROI colours
  const ROI_COLORS = ['#4af0c4','#f0c44a','#f04a6a','#a04af0','#4a80f0','#f0804a'];

  // ── status ──
  function setStatus(msg, type='') {
    const el = document.getElementById('status');
    el.textContent = msg;
    el.className = 'status-bar' + (type ? ' status-' + type : '');
  }

  async function pickFolder() {
    setStatus('Opening folder picker...', '');
    try {
      const r = await fetch('/pick_folder');
      const data = await r.json();
      if (data.path) {
        document.getElementById('folder-input').value = data.path;
        await loadFolder(data.path);
      } else {
        setStatus('No folder selected.', '');
      }
    } catch(e) {
      setStatus('Picker error: ' + e.message, 'err');
    }
  }

  // ── folder loading ──
  async function loadFolder(folder) {
    folder = (folder || document.getElementById('folder-input').value).trim();
    if (!folder) return;
    setStatus('Loading...', '');
    try {
      const r = await fetch('/browse', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({folder})
      });
      const data = await r.json();
      currentFolder = folder;
      lastSaveDir = data.analysis_dir || null;
      sessionMeasurements = [];
      analysisData = [];
      analysisDataUid = 0;
      document.getElementById('header-path').textContent = folder;
      const name = fixPath(folder).split('/').filter(Boolean).pop() || folder;
      document.getElementById('experiment-label').textContent = name;
      document.getElementById('folder-input').value = folder;
      dayPair = null; biolumPair = null; dayImg = null; biolumImg = null;
      setSampleLabel('day', null); setSampleLabel('biolum', null);
      renderPairList(data.pairs);
      const dirMsg = lastSaveDir ? (' → ' + fixPath(lastSaveDir).split('/').filter(Boolean).pop()) : '';
      setStatus('Loaded ' + data.pairs.length + ' images' + dirMsg, 'ok');
    } catch(e) {
      setStatus('Load error: ' + e.message, 'err');
    }
  }

  function renderPairList(pairs) {
    const list = document.getElementById('pair-list');
    document.getElementById('pair-count').textContent = pairs.length + ' pairs';
    if (!pairs.length) {
      list.innerHTML = '<div style="padding:20px;text-align:center;font-family:var(--mono);font-size:11px;color:var(--muted);">No image pairs found</div>';
      return;
    }
    list.innerHTML = pairs.map((p, i) => `
      <div class="pair-item" onclick="loadPair(${i})" id="pair-${i}" data-pair='${JSON.stringify(p)}'>
        <div style="display:flex;align-items:center;justify-content:space-between;gap:4px;">
          <span style="overflow:hidden;text-overflow:ellipsis;">${p.stem}</span>
          <span class="badge badge-${p.type}" style="flex-shrink:0;">${p.type.toUpperCase()}</span>
        </div>
      </div>
    `).join('');
  }

  function setSampleLabel(side, stem) {
    const el = document.getElementById(side + '-sample-label');
    if (stem) {
      el.textContent = sampleName(stem);
      el.style.display = '';
    } else {
      el.textContent = '';
      el.style.display = 'none';
    }
  }

  async function loadPair(idx) {
    const el = document.getElementById('pair-' + idx);
    const pair = JSON.parse(el.dataset.pair);
    const isDay = pair.type !== 'biolum';
    const newSample = sampleName(pair.stem);

    // Enforce same-sample pairing
    if (isDay && biolumPair) {
      const existing = sampleName(biolumPair.stem);
      if (newSample.toLowerCase() !== existing.toLowerCase()) {
        setStatus('Sample mismatch: "' + newSample + '" vs loaded biolum "' + existing + '". Clear the biolum image first.', 'err');
        return;
      }
    } else if (!isDay && dayPair) {
      const existing = sampleName(dayPair.stem);
      if (newSample.toLowerCase() !== existing.toLowerCase()) {
        setStatus('Sample mismatch: "' + newSample + '" vs loaded day "' + existing + '". Clear the day image first.', 'err');
        return;
      }
    }

    // Update active highlights
    document.querySelectorAll('.pair-item').forEach(e => e.classList.remove('active', 'active-day', 'active-biolum'));
    el.classList.add('active', isDay ? 'active-day' : 'active-biolum');

    if (isDay) {
      dayPair = pair;
      rois = [];
      measurements = [];
      selectedRoi = -1;
      document.getElementById('results-table-wrap').innerHTML =
        '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click ▶ Measure ROIs</div>';
      document.getElementById('download-links').innerHTML = '';
    } else {
      biolumPair = pair;
    }

    if (pair.jpg) {
      if (isDay) {
        await loadImageToCanvas(pair.jpg, dayCanvas, dayCtx, 'day-placeholder');
        const sr = await fetch('/image_size?path=' + encodeURIComponent(pair.jpg));
        const sd = await sr.json();
        jpegSize = {w: sd.width, h: sd.height};
        dayImg = new Image();
        dayImg.src = '/image?path=' + encodeURIComponent(pair.jpg);
        await new Promise(r => { dayImg.onload = r; });
      } else {
        await loadImageToCanvas(pair.jpg, biolumCanvas, biolumCtx, 'biolum-placeholder');
        biolumImg = new Image();
        biolumImg.src = '/image?path=' + encodeURIComponent(pair.jpg);
        await new Promise(r => { biolumImg.onload = r; });
      }
    }

    setSampleLabel(isDay ? 'day' : 'biolum', pair.stem);
    document.getElementById((isDay ? 'day' : 'biolum') + '-clear-btn').style.display = '';
    drawAll();

    if (isDay && !biolumPair) {
      setStatus('"' + newSample + '" day loaded — select the matching biolum image.', 'wrn');
    } else if (!isDay && !dayPair) {
      setStatus('"' + newSample + '" biolum loaded — select the matching day image.', 'wrn');
    } else {
      setStatus('Set complete: ' + newSample, 'ok');
    }
  }

  async function loadImageToCanvas(path, canvas, ctx, placeholderId) {
    const img = new Image();
    img.src = '/image?path=' + encodeURIComponent(path);
    await new Promise(r => { img.onload = r; img.onerror = r; });
    const wrap = canvas.parentElement;
    const maxW = wrap.clientWidth;
    const maxH = wrap.clientHeight;
    const scale = Math.min(maxW / img.naturalWidth, maxH / img.naturalHeight, 1);
    canvas.width  = Math.round(img.naturalWidth  * scale);
    canvas.height = Math.round(img.naturalHeight * scale);
    ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
    canvas.style.display = 'block';
    document.getElementById(placeholderId).style.display = 'none';
  }

  // ── ROI drawing ──
  function setMode(m) {
    mode = m;
    ['draw','move','resize','pan'].forEach(mm => {
      document.getElementById('mode-' + mm).classList.toggle('active', mm === m);
    });
    const cur = m === 'draw' ? 'crosshair' : m === 'move' ? 'move' : m === 'pan' ? 'grab' : 'nw-resize';
    dayCanvas.style.cursor = cur;
    biolumCanvas.style.cursor = cur;
  }

  function resetView() {
    view = {scale: 1, dx: 0, dy: 0};
    drawAll();
  }

  function applyViewTransform(ctx, canvas) {
    const cx = canvas.width / 2, cy = canvas.height / 2;
    ctx.translate(cx + view.dx, cy + view.dy);
    ctx.scale(view.scale, view.scale);
    ctx.translate(-cx, -cy);
  }

  function dayToBiolum(pos) {
    if (!biolumCanvas.width || !dayCanvas.width) return pos;
    return {x: pos.x * biolumCanvas.width / dayCanvas.width,
            y: pos.y * biolumCanvas.height / dayCanvas.height};
  }

  function biolumToDay(pos) {
    if (!dayCanvas.width || !biolumCanvas.width) return pos;
    return {x: pos.x * dayCanvas.width / biolumCanvas.width,
            y: pos.y * dayCanvas.height / biolumCanvas.height};
  }

  function drawPreview(startDay, endDay) {
    [{ctx: dayCtx, canvas: dayCanvas, s: startDay, e: endDay},
     {ctx: biolumCtx, canvas: biolumCanvas,
      s: dayToBiolum(startDay), e: dayToBiolum(endDay)}
    ].forEach(({ctx, canvas, s, e}) => {
      if (!canvas.width) return;
      ctx.save();
      applyViewTransform(ctx, canvas);
      ctx.strokeStyle = 'rgba(74,240,196,0.8)';
      ctx.lineWidth = 2;
      ctx.setLineDash([4,4]);
      const px = Math.min(s.x,e.x), py = Math.min(s.y,e.y);
      const pw = Math.abs(e.x-s.x), ph = Math.abs(e.y-s.y);
      if (roiShape === 'circle') {
        ctx.beginPath();
        ctx.ellipse(px+pw/2, py+ph/2, pw/2, ph/2, 0, 0, Math.PI*2);
        ctx.stroke();
      } else {
        ctx.strokeRect(s.x, s.y, e.x-s.x, e.y-s.y);
      }
      ctx.setLineDash([]);
      ctx.restore();
    });
  }

  function constrainSquare(start, end) {
    const dx = end.x - start.x, dy = end.y - start.y;
    const size = Math.min(Math.abs(dx), Math.abs(dy));
    return {x: start.x + Math.sign(dx) * size, y: start.y + Math.sign(dy) * size};
  }

  function commitRoi(endPosDay) {
    let x = Math.min(drawStart.x, endPosDay.x);
    let y = Math.min(drawStart.y, endPosDay.y);
    let w = Math.abs(endPosDay.x - drawStart.x);
    let h = Math.abs(endPosDay.y - drawStart.y);
    if (fixedSize) { w = fixedSize.w; h = fixedSize.h; }
    if (w > 5 && h > 5) {
      rois.push({x, y, w, h, shape: roiShape, type: roiType});
      sortRois();
      selectedRoi = rois.length - 1;
    }
    isDrawing = false;
    drawAll();
  }

  function sampleName(stem) {
    let m = stem.match(/^(.+?)_biolum/i);
    if (m) return m[1].trim();
    m = stem.match(/^(.+?)_day/i);
    if (m) return m[1].trim();
    m = stem.match(/^(.+?)_light/i);
    if (m) return m[1].trim();
    return stem.trim();
  }

  function toggleFixSize() {
    const checked = document.getElementById('fix-size').checked;
    if (checked && rois.length > 0) {
      const last = rois[rois.length - 1];
      fixedSize = {w: last.w, h: last.h};
    } else {
      fixedSize = null;
    }
  }

  function canvasPos(canvas, e) {
    const rect = canvas.getBoundingClientRect();
    const sx = (e.clientX - rect.left) * (canvas.width / rect.width);
    const sy = (e.clientY - rect.top)  * (canvas.height / rect.height);
    const cx = canvas.width / 2, cy = canvas.height / 2;
    return {
      x: (sx - cx - view.dx) / view.scale + cx,
      y: (sy - cy - view.dy) / view.scale + cy,
      sx, sy
    };
  }

  function onWheel(canvas, e) {
    e.preventDefault();
    const rect = canvas.getBoundingClientRect();
    const sx = (e.clientX - rect.left) * (canvas.width / rect.width);
    const sy = (e.clientY - rect.top)  * (canvas.height / rect.height);
    const cx = canvas.width / 2, cy = canvas.height / 2;
    const factor = e.deltaY < 0 ? 1.12 : 1 / 1.12;
    const ns = Math.max(0.1, Math.min(20, view.scale * factor));
    const ratio = ns / view.scale;
    view.dx = (1 - ratio) * (sx - cx) + ratio * view.dx;
    view.dy = (1 - ratio) * (sy - cy) + ratio * view.dy;
    view.scale = ns;
    drawAll();
  }
  dayCanvas.addEventListener('wheel', e => onWheel(dayCanvas, e), {passive: false});
  biolumCanvas.addEventListener('wheel', e => onWheel(biolumCanvas, e), {passive: false});

  function hitTest(x, y) {
    // returns index of ROI hit, -1 if none
    // check in reverse so top-most ROI is selected first
    for (let i = rois.length - 1; i >= 0; i--) {
      const r = rois[i];
      if (x >= r.x && x <= r.x + r.w && y >= r.y && y <= r.y + r.h) return i;
    }
    return -1;
  }

  function resizeHandle(roi, x, y) {
    // returns true if near bottom-right corner
    const dx = x - (roi.x + roi.w);
    const dy = y - (roi.y + roi.h);
    return Math.abs(dx) < 10 && Math.abs(dy) < 10;
  }

  function roiMouseDown(pos, rawPos) {
    if (mode === 'pan') {
      panStart = {sx: rawPos.sx, sy: rawPos.sy, dx: view.dx, dy: view.dy};
      dayCanvas.style.cursor = 'grabbing';
      biolumCanvas.style.cursor = 'grabbing';
      return;
    }
    if (!dayPair && !biolumPair) return;
    if (mode === 'draw') {
      const preHit = hitTest(pos.x, pos.y);
      if (preHit >= 0) { selectedRoi = preHit; drawAll(); return; }
      isDrawing = true; drawStart = pos; selectedRoi = -1;
    } else if (mode === 'move') {
      const hit = hitTest(pos.x, pos.y);
      if (hit >= 0) {
        selectedRoi = hit; dragStart = pos;
        dragRoiRef = rois[hit]; dragRoiStart = {...rois[hit]};
      }
    } else if (mode === 'resize') {
      const hit = hitTest(pos.x, pos.y);
      if (hit >= 0) {
        selectedRoi = hit; dragStart = pos;
        dragRoiRef = rois[hit]; dragRoiStart = {...rois[hit]};
      }
    }
    drawAll();
  }

  function roiMouseMove(pos, rawPos, buttons, shiftKey) {
    if (mode === 'pan' && panStart && buttons === 1) {
      view.dx = panStart.dx + rawPos.sx - panStart.sx;
      view.dy = panStart.dy + rawPos.sy - panStart.sy;
      drawAll(); return;
    }
    if (!isDrawing && mode !== 'move' && mode !== 'resize') return;
    if (isDrawing && mode === 'draw') {
      const ep = shiftKey ? constrainSquare(drawStart, pos) : pos;
      drawAll(); drawPreview(drawStart, ep);
    } else if (mode === 'move' && dragRoiRef && buttons === 1) {
      dragRoiRef.x = dragRoiStart.x + pos.x - dragStart.x;
      dragRoiRef.y = dragRoiStart.y + pos.y - dragStart.y;
      drawAll();
    } else if (mode === 'resize' && dragRoiRef && buttons === 1) {
      dragRoiRef.w = Math.max(10, dragRoiStart.w + pos.x - dragStart.x);
      dragRoiRef.h = Math.max(10, dragRoiStart.h + pos.y - dragStart.y);
      drawAll();
    }
  }

  function roiMouseUp(pos, shiftKey) {
    if (mode === 'pan') {
      panStart = null;
      const cur = 'grab';
      dayCanvas.style.cursor = cur; biolumCanvas.style.cursor = cur; return;
    }
    if (isDrawing && mode === 'draw') {
      commitRoi(shiftKey ? constrainSquare(drawStart, pos) : pos);
    } else if ((mode === 'move' || mode === 'resize') && dragRoiRef) {
      sortRois();
      selectedRoi = rois.indexOf(dragRoiRef);
      dragRoiRef = null;
      drawAll();
    }
  }

  // day canvas events
  dayCanvas.addEventListener('mousedown', e => {
    const p = canvasPos(dayCanvas, e);
    roiMouseDown(p, p);
  });
  dayCanvas.addEventListener('mousemove', e => {
    const p = canvasPos(dayCanvas, e);
    roiMouseMove(p, p, e.buttons, e.shiftKey);
  });
  dayCanvas.addEventListener('mouseup', e => {
    roiMouseUp(canvasPos(dayCanvas, e), e.shiftKey);
  });
  dayCanvas.addEventListener('dblclick', e => {
    const hit = hitTest(canvasPos(dayCanvas, e).x, canvasPos(dayCanvas, e).y);
    if (hit >= 0) { selectedRoi = hit; drawAll(); }
  });

  // biolum canvas events — coords converted to day space
  biolumCanvas.addEventListener('mousedown', e => {
    const raw = canvasPos(biolumCanvas, e);
    roiMouseDown(biolumToDay(raw), raw);
  });
  biolumCanvas.addEventListener('mousemove', e => {
    const raw = canvasPos(biolumCanvas, e);
    roiMouseMove(biolumToDay(raw), raw, e.buttons, e.shiftKey);
  });
  biolumCanvas.addEventListener('mouseup', e => {
    roiMouseUp(biolumToDay(canvasPos(biolumCanvas, e)), e.shiftKey);
  });
  biolumCanvas.addEventListener('dblclick', e => {
    const pos = biolumToDay(canvasPos(biolumCanvas, e));
    const hit = hitTest(pos.x, pos.y);
    if (hit >= 0) { selectedRoi = hit; drawAll(); }
  });

  function sortRois() {
    // sort left to right, top to bottom (by row then column)
    const rowThreshold = 20;
    rois.sort((a, b) => {
      if (Math.abs(a.y - b.y) > rowThreshold) return a.y - b.y;
      return a.x - b.x;
    });
  }

  function drawOneCanvas(ctx, canvas, img, isBiolum) {
    ctx.setTransform(1, 0, 0, 1, 0, 0);
    ctx.clearRect(0, 0, canvas.width, canvas.height);
    if (!img) return;
    ctx.save();
    applyViewTransform(ctx, canvas);
    ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
    if (isBiolum && dayCanvas.width > 0) {
      drawRoisScaled(ctx, canvas.width / dayCanvas.width, canvas.height / dayCanvas.height);
    } else {
      drawRois(ctx, canvas);
    }
    ctx.restore();
  }

  function drawAll() {
    drawOneCanvas(dayCtx, dayCanvas, dayImg, false);
    drawOneCanvas(biolumCtx, biolumCanvas, biolumImg, true);
  }

  function getRoiLabel(roi, idx) {
    if (roi.type === 'bckg') {
      const n = rois.slice(0, idx+1).filter(r => r.type === 'bckg').length;
      return 'B' + n;
    }
    const n = rois.slice(0, idx+1).filter(r => r.type !== 'bckg').length;
    return String(n);
  }

  function drawSingleRoi(ctx, x, y, w, h, shape, color, isSelected, label) {
    ctx.strokeStyle = color;
    ctx.lineWidth = isSelected ? 2.5 : 1.5;
    ctx.fillStyle = color + '22';
    if (shape === 'circle') {
      ctx.beginPath();
      ctx.ellipse(x+w/2, y+h/2, w/2, h/2, 0, 0, Math.PI*2);
      ctx.stroke(); ctx.fill();
    } else {
      ctx.strokeRect(x, y, w, h);
      ctx.fillRect(x, y, w, h);
      ctx.fillStyle = color;
      ctx.fillRect(x+w-6, y+h-6, 6, 6);
    }
    ctx.fillStyle = color;
    ctx.font = 'bold 12px IBM Plex Mono';
    ctx.fillText(label, x+4, y+14);
  }

  function drawRois(ctx, canvas) {
    rois.forEach((roi, i) => {
      drawSingleRoi(ctx, roi.x, roi.y, roi.w, roi.h,
        roi.shape||'rect', ROI_COLORS[i % ROI_COLORS.length], roi === dragRoiRef || i === selectedRoi, getRoiLabel(roi,i));
    });
  }

  function drawRoisScaled(ctx, sx, sy) {
    rois.forEach((roi, i) => {
      drawSingleRoi(ctx, roi.x*sx, roi.y*sy, roi.w*sx, roi.h*sy,
        roi.shape||'rect', ROI_COLORS[i % ROI_COLORS.length], roi === dragRoiRef || i === selectedRoi, getRoiLabel(roi,i));
    });
  }

  function deleteSelected() {
    if (selectedRoi >= 0 && selectedRoi < rois.length) {
      rois.splice(selectedRoi, 1);
      selectedRoi = -1;
      dragRoiRef = null;
      drawAll();
    }
  }

  function clearAllRois() {
    rois = [];
    selectedRoi = -1;
    measurements = [];
    drawAll();
  }

  // ── measurement ──
  async function measureAll() {
    if (!dayPair && !biolumPair) { setStatus('Load images first', 'err'); return; }
    if (!rois.length) { setStatus('Draw at least one ROI first', 'wrn'); return; }
    if (!biolumPair || !biolumPair.nef) { setStatus('No biolum NEF loaded — select a biolum image with a NEF file', 'err'); return; }

    sortRois();
    setStatus('Reading NEF and measuring... (this may take a few seconds)', '');
    document.getElementById('btn-measure').disabled = true;

    // convert canvas ROIs to JPEG pixel coords
    const scaleX = jpegSize.w / dayCanvas.width;
    const scaleY = jpegSize.h / dayCanvas.height;
    const scaledRois = rois.map(r => ({
      x: Math.round(r.x * scaleX),
      y: Math.round(r.y * scaleY),
      w: Math.round(r.w * scaleX),
      h: Math.round(r.h * scaleY),
      shape: r.shape || 'rect',
      type: r.type || 'roi',
    }));

    const r = await fetch('/measure', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({
        nef_path: biolumPair.nef,
        rois: scaledRois,
        jpeg_size: [jpegSize.w, jpegSize.h]
      })
    });
    const data = await r.json();
    document.getElementById('btn-measure').disabled = false;

    if (data.error) { setStatus('Error: ' + data.error, 'err'); return; }
    const sample = biolumPair ? sampleName(biolumPair.stem) : (dayPair ? sampleName(dayPair.stem) : '');

    // compute background mean if any background ROIs exist
    const bckgMeas = data.measurements.filter((m, i) => scaledRois[i] && scaledRois[i].type === 'bckg');
    let bckgMean = null;
    if (bckgMeas.length > 0) {
      const avg = key => bckgMeas.reduce((s, m) => s + m[key], 0) / bckgMeas.length;
      bckgMean = {
        mean_R: avg('mean_R'), mean_G: avg('mean_G'), mean_B: avg('mean_B'),
        intden_R: avg('intden_R'), intden_G: avg('intden_G'), intden_B: avg('intden_B'),
      };
    }

    let roiCount = 0, bckgCount = 0;
    measurements = data.measurements.map((m, i) => {
      const roiObj = scaledRois[i] || {};
      const isB = roiObj.type === 'bckg';
      if (isB) bckgCount++; else roiCount++;
      const roi_label = isB ? ('Bckg ' + bckgCount) : String(roiCount);
      const nb = bckgMean ? {
        mean_R_nb:   isB ? null : m.mean_R   - bckgMean.mean_R,
        mean_G_nb:   isB ? null : m.mean_G   - bckgMean.mean_G,
        mean_B_nb:   isB ? null : m.mean_B   - bckgMean.mean_B,
        intden_R_nb: isB ? null : m.intden_R - bckgMean.intden_R,
        intden_G_nb: isB ? null : m.intden_G - bckgMean.intden_G,
        intden_B_nb: isB ? null : m.intden_B - bckgMean.intden_B,
      } : {};
      return {...m, sample_name: sample, roi_type: roiObj.type||'roi', roi_label, ...nb};
    });

    // accumulate session measurements (all ROIs including bckg)
    sessionMeasurements.push(...measurements);

    // accumulate for Analysis Summary (non-background ROIs only)
    measurements.filter(m => m.roi_type !== 'bckg').forEach(m => {
      const uid = analysisDataUid++;
      m._uid = uid;
      analysisData.push({
        uid,
        sample: m.sample_name,
        roiLabel: m.roi_label,
        mean_R_nb: m.mean_R_nb !== undefined && m.mean_R_nb !== null ? m.mean_R_nb : m.mean_R,
        mean_G_nb: m.mean_G_nb !== undefined && m.mean_G_nb !== null ? m.mean_G_nb : m.mean_G,
        mean_B_nb: m.mean_B_nb !== undefined && m.mean_B_nb !== null ? m.mean_B_nb : m.mean_B,
      });
    });

    renderTable(measurements);
    setStatus(`Measured ${measurements.length} ROIs from NEF (16-bit)`, 'ok');
  }

  function renderTable(data) {
    if (!data.length) return;
    const wrap = document.getElementById('results-table-wrap');
    const hasBckg = data.some(m => m.mean_R_nb !== undefined && m.mean_R_nb !== null);
    const valueCols = [
      {label:'Area (px)', key:'area_px'},
      {label:'Mean B', key:'mean_B', cls:'ch-b'},
      {label:'Mean G', key:'mean_G', cls:'ch-g'},
      {label:'Mean R', key:'mean_R', cls:'ch-r'},
      ...(hasBckg ? [
        {label:'Mean B-Bckg', key:'mean_B_nb', cls:'ch-b-nb'},
        {label:'Mean G-Bckg', key:'mean_G_nb', cls:'ch-g-nb'},
        {label:'Mean R-Bckg', key:'mean_R_nb', cls:'ch-r-nb'},
      ] : []),
    ];
    const fmtVal = v => {
      if (v === null || v === undefined) return '—';
      if (typeof v === 'number') return v.toLocaleString(undefined, {maximumFractionDigits: 1});
      return v;
    };
    const roiHeader = renameMode ? 'ROI Name' : 'ROI';
    const sampleHeader = renameMode ? '' : '<th>Sample</th>';
    wrap.innerHTML = `
      <table>
        <thead><tr>${sampleHeader}<th>${roiHeader}</th>${valueCols.map(c => `<th>${c.label}</th>`).join('')}</tr></thead>
        <tbody>
          ${data.map((row, ri) => {
            const isB = row.roi_type === 'bckg';
            const rowCls = isB ? ' class="bckg-row"' : '';
            const sampleCell = renameMode ? '' : `<td>${fmtVal(row.sample_name)}</td>`;
            const roiCell = (renameMode && !isB)
              ? `<td><input type="text" value="${row.roi_label}" style="font-family:var(--mono);font-size:11px;background:transparent;border:1px solid var(--border2);color:var(--text);border-radius:2px;padding:1px 4px;width:80px;" onchange="renameMeasurement(${ri}, this.value)"></td>`
              : `<td>${fmtVal(row.roi_label)}</td>`;
            return `<tr${rowCls}>${sampleCell}${roiCell}${valueCols.map(c => {
              const cls = isB && (c.cls||'').includes('-nb') ? '' : (c.cls||'');
              return `<td class="${cls}">${fmtVal(row[c.key])}</td>`;
            }).join('')}</tr>`;
          }).join('')}
        </tbody>
      </table>`;
  }

  // ── rename ROIs ──
  function toggleRenameMode() {
    renameMode = document.getElementById('rename-mode').checked;
    if (measurements.length) renderTable(measurements);
    renderSummary();
  }

  function renameMeasurement(idx, newLabel) {
    const m = measurements[idx];
    if (!m || m.roi_type === 'bckg') return;
    m.roi_label = newLabel;
    if (m._uid !== undefined) {
      const ad = analysisData.find(d => d.uid === m._uid);
      if (ad) ad.roiLabel = newLabel;
    }
    renderSummary();
  }

  // ── save ──
  async function saveResults() {
    if ((!dayPair && !biolumPair) || !rois.length) { setStatus('Nothing to save', 'wrn'); return; }

    const scaleX = jpegSize.w / dayCanvas.width;
    const scaleY = jpegSize.h / dayCanvas.height;
    const scaledRois = rois.map(r => ({
      ...r,
      x: Math.round(r.x * scaleX), y: Math.round(r.y * scaleY),
      w: Math.round(r.w * scaleX), h: Math.round(r.h * scaleY),
    }));

    const saveStem = biolumPair ? biolumPair.stem : dayPair.stem;
    const snapshot_png = captureSnapshotDataURL() || '';
    setStatus('Saving...', '');
    const r = await fetch('/save', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({
        folder: currentFolder,
        save_dir: lastSaveDir,
        stem: saveStem,
        rois: scaledRois,
        measurements,
        session_measurements: sessionMeasurements,
        jpeg_size: [jpegSize.w, jpegSize.h],
        snapshot_png,
      })
    });
    const data = await r.json();
    if (data.ok) {
      lastSaveDir = data.dir;
      setStatus('Saved to: ' + data.dir, 'ok');
      const links = document.getElementById('download-links');
      links.innerHTML = `
        <a href="/download?path=${encodeURIComponent(fixPath(data.dir) + '/session_measurements.xlsx')}"
           style="font-family:var(--mono);font-size:10px;color:var(--amber);text-decoration:none;border:1px solid var(--border2);padding:2px 8px;border-radius:2px;">
           ↓ Excel</a>`;
    } else {
      setStatus('Save failed: ' + data.error, 'err');
    }
  }

  // ── load ROIs ──
  async function showLoadRoiDialog() {
    if (!currentFolder) { setStatus('Load a folder first', 'wrn'); return; }
    const r = await fetch('/list_roi_files', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({folder: currentFolder})
    });
    const data = await r.json();
    const list = document.getElementById('roi-file-list');
    if (!data.files.length) {
      list.innerHTML = '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">No saved ROI files found in this folder.</div>';
    } else {
      list.innerHTML = data.files.map(f => `
        <div data-path="${f.path}" onclick="loadRoiFile(this.dataset.path)"
             style="padding:8px 10px;border:1px solid var(--border2);border-radius:3px;cursor:pointer;font-family:var(--mono);font-size:11px;color:var(--text);">
          ${f.name}
        </div>`).join('');
    }
    document.getElementById('roi-dialog').style.display = 'flex';
  }

  function clearPanel(side) {
    if (side === 'day') {
      dayPair = null; dayImg = null;
      dayCanvas.style.display = 'none';
      document.getElementById('day-placeholder').style.display = '';
      document.getElementById('day-clear-btn').style.display = 'none';
      setSampleLabel('day', null);
      document.querySelectorAll('.pair-item.active-day').forEach(e => e.classList.remove('active', 'active-day'));
      rois = []; measurements = []; selectedRoi = -1;
      document.getElementById('results-table-wrap').innerHTML =
        '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click ▶ Measure ROIs</div>';
      document.getElementById('download-links').innerHTML = '';
    } else {
      biolumPair = null; biolumImg = null;
      biolumCanvas.style.display = 'none';
      document.getElementById('biolum-placeholder').style.display = '';
      document.getElementById('biolum-clear-btn').style.display = 'none';
      setSampleLabel('biolum', null);
      document.querySelectorAll('.pair-item.active-biolum').forEach(e => e.classList.remove('active', 'active-biolum'));
    }
    drawAll();
    setStatus('Panel cleared. Select a new image.', '');
  }

  // ── background type toggle ──
  function toggleBckgType() {
    roiType = roiType === 'bckg' ? 'roi' : 'bckg';
    document.getElementById('type-bckg').classList.toggle('active', roiType === 'bckg');
  }

  // ── top-level tab switching ──
  function switchTopTab(tab) {
    ['measure','summary'].forEach(t => {
      document.getElementById('top-tab-' + t).classList.toggle('active', t === tab);
      document.getElementById('top-tab-btn-' + t).classList.toggle('active', t === tab);
    });
    if (tab === 'summary') renderSummary();
  }

  // ── analysis summary / box plots ──
  function clearAnalysisData() {
    analysisData = [];
    renderSummary();
  }

  // ── Tukey HSD + CLD ──

  // Studentized range critical values q(α=0.05, k, df) — Tukey-Kramer table
  const _TDF = [4,5,6,7,8,9,10,12,15,20,24,30,40,60,120,Infinity];
  const _TQCRIT = {
    2:[3.93,3.64,3.46,3.34,3.26,3.20,3.15,3.08,3.01,2.95,2.92,2.89,2.86,2.83,2.80,2.77],
    3:[5.04,4.60,4.34,4.16,4.04,3.95,3.88,3.77,3.67,3.58,3.53,3.49,3.44,3.40,3.36,3.31],
    4:[5.76,5.22,4.90,4.68,4.53,4.41,4.33,4.20,4.08,3.96,3.90,3.85,3.79,3.74,3.68,3.63],
    5:[6.29,5.67,5.30,5.06,4.89,4.76,4.65,4.51,4.37,4.23,4.17,4.10,4.04,3.98,3.92,3.86],
    6:[6.71,6.03,5.63,5.36,5.17,5.02,4.91,4.75,4.59,4.45,4.37,4.30,4.23,4.16,4.10,4.03],
    7:[7.05,6.33,5.90,5.61,5.40,5.24,5.12,4.95,4.78,4.62,4.54,4.46,4.39,4.31,4.24,4.17],
    8:[7.35,6.58,6.12,5.82,5.60,5.43,5.30,5.12,4.94,4.77,4.68,4.60,4.52,4.44,4.36,4.29],
  };

  function tukeyQCrit(k, df) {
    const row = _TQCRIT[Math.min(Math.max(k, 2), 8)];
    if (df <= _TDF[0]) return row[0];
    if (!isFinite(df)) return row[row.length - 1];
    for (let i = 0; i < _TDF.length - 1; i++) {
      if (df >= _TDF[i] && df <= _TDF[i+1]) {
        const t = (df - _TDF[i]) / (_TDF[i+1] - _TDF[i]);
        return row[i] + t * (row[i+1] - row[i]);
      }
    }
    return row[row.length - 1];
  }

  function tukeyHSD(groupData) {
    // groupData: [{name, vals}]
    const k = groupData.length;
    if (k < 2) return null;
    const grps = groupData.map(g => {
      const n = g.vals.length, mean = g.vals.reduce((a,b)=>a+b,0)/n;
      const ss = g.vals.reduce((a,b)=>a+(b-mean)**2, 0);
      return {name:g.name, n, mean, ss};
    });
    const N = grps.reduce((a,g)=>a+g.n, 0);
    const df = N - k;
    if (df < 1) return null;
    const MSE = grps.reduce((a,g)=>a+g.ss,0) / df;
    if (MSE <= 0) return null;
    const qCrit = tukeyQCrit(k, df);
    const sig = Array.from({length:k}, ()=>Array(k).fill(false));
    for (let i = 0; i < k; i++)
      for (let j = i+1; j < k; j++) {
        const q = Math.abs(grps[i].mean - grps[j].mean) /
                  (Math.sqrt(MSE) * Math.sqrt(0.5*(1/grps[i].n + 1/grps[j].n)));
        sig[i][j] = sig[j][i] = q > qCrit;
      }
    return {sig};
  }

  function computeCLD(k, sig, sortedIdx) {
    // Piepho (2004) absorption algorithm
    if (k === 0) return [];
    if (k === 1) return ['a'];
    const sets = Array.from({length:k}, ()=>new Set([0]));
    let next = 1, changed = true;
    while (changed) {
      changed = false;
      for (let pi = 0; pi < k; pi++) {
        const i = sortedIdx[pi];
        for (let pj = pi+1; pj < k; pj++) {
          const j = sortedIdx[pj];
          if (!sig[i][j]) continue;
          const common = [...sets[i]].filter(l=>sets[j].has(l));
          if (!common.length) continue;
          changed = true;
          for (const l of common) {
            const lp = next++;
            const Li = new Set(), Lj = new Set();
            for (let m = 0; m < k; m++) {
              if (!sets[m].has(l)) continue;
              if (!sig[i][m]) Li.add(m);
              if (!sig[j][m]) Lj.add(m);
            }
            for (let m = 0; m < k; m++) {
              if (!sets[m].has(l)) continue;
              if (Li.has(m) && !Lj.has(m)) { sets[m].delete(l); sets[m].add(lp); }
              else if (Li.has(m) &&  Lj.has(m)) { sets[m].add(lp); }
              // Lj only or neither: keep l unchanged
            }
          }
        }
      }
    }
    // Build ordered letter map: first letter of highest-mean group → 'a'
    const order = [];
    for (const si of sortedIdx)
      for (const l of [...sets[si]].sort((a,b)=>a-b))
        if (!order.includes(l)) order.push(l);
    const alpha = 'abcdefghijklmnop';
    const lmap = Object.fromEntries(order.map((l,i)=>[l, alpha[i]]));
    return Array.from({length:k}, (_,i) =>
      [...sets[i]].sort((a,b)=>order.indexOf(a)-order.indexOf(b)).map(l=>lmap[l]).join('')
    );
  }

  function computeStats(values) {
    if (!values.length) return null;
    const s = [...values].sort((a,b)=>a-b);
    const n = s.length;
    const q = p => { const i=p*(n-1); const lo=Math.floor(i),hi=Math.ceil(i); return s[lo]+(s[hi]-s[lo])*(i-lo); };
    return {min:s[0], q1:q(0.25), median:q(0.5), q3:q(0.75), max:s[n-1], values:s};
  }

  function summaryGroupKey(d) { return renameMode ? (d.roiLabel || d.sample) : d.sample; }

  function summaryYRange() {
    const allVals = SUMMARY_CHANNELS.flatMap(ch =>
      analysisData.map(d => d[ch.key]).filter(v => v != null && isFinite(v))
    );
    if (!allVals.length) return {yMin: 0, yMax: 1};
    const gMin = Math.min(0, ...allVals), gMax = Math.max(0, ...allVals);
    const pad = (gMax - gMin) * 0.12 || Math.abs(gMax) * 0.12 || 1;
    return {yMin: gMin - pad, yMax: gMax + pad};
  }

  function renderSummary() {
    const empty = document.getElementById('summary-empty');
    const plots = document.getElementById('summary-plots');
    if (!analysisData.length) {
      empty.style.display = ''; plots.style.display = 'none'; return;
    }
    empty.style.display = 'none'; plots.style.display = 'flex';
    const groups = [...new Set(analysisData.map(summaryGroupKey))];
    const {yMin, yMax} = summaryYRange();
    const plotsDiv = document.getElementById('summary-plots');
    const gap = 16;
    const n = SUMMARY_CHANNELS.length;
    const canvasW = Math.floor((plotsDiv.clientWidth  - gap * (n - 1)) / n) || 400;
    const canvasH = plotsDiv.clientHeight || 500;
    SUMMARY_CHANNELS.forEach(ch => {
      const canvas = document.getElementById(ch.id);
      canvas.width  = canvasW;
      canvas.height = canvasH;
      canvas.style.width  = canvasW + 'px';
      canvas.style.height = canvasH + 'px';
      drawBoxPlot(canvas, groups, ch.key, ch.color, ch.label, yMin, yMax, 'dark');
    });
  }

  function fmtVal(v) {
    const a = Math.abs(v);
    if (a === 0) return '0';
    if (a >= 1e9) return (v/1e9).toFixed(2) + ' G';
    if (a >= 1e6) return (v/1e6).toFixed(2) + ' M';
    if (a >= 1e3) return (v/1e3).toFixed(1) + ' k';
    if (a >= 1)   return v.toFixed(1);
    return v.toPrecision(3);
  }

  function drawBoxPlot(canvas, samples, key, color, title, yMinOverride = null, yMaxOverride = null, theme = 'dark') {
    const ctx = canvas.getContext('2d');
    const W = canvas.width, H = canvas.height;
    ctx.clearRect(0, 0, W, H);

    const isLight = theme === 'light';
    const bgCol    = isLight ? '#ffffff' : '#0b0f15';
    const gridCol  = isLight ? '#cccccc' : '#1e2d42';
    const axisCol  = isLight ? '#666666' : '#2e4060';
    const tickCol  = isLight ? '#333333' : '#7090a8';
    const xLblCol  = isLight ? '#111111' : '#c8dff0';
    const medCol   = isLight ? '#000000' : '#ffffff';

    ctx.fillStyle = bgCol; ctx.fillRect(0, 0, W, H);

    // determine Y range
    const allVals = analysisData.map(d => d[key]).filter(v => v != null && isFinite(v));
    if (!allVals.length) return;
    let yMin, yMax;
    if (yMinOverride !== null && yMaxOverride !== null) {
      yMin = yMinOverride; yMax = yMaxOverride;
    } else {
      yMin = Math.min(...allVals); yMax = Math.max(...allVals);
      const pad = (yMax - yMin) * 0.12 || Math.abs(yMax) * 0.12 || 1;
      yMin -= pad; yMax += pad;
    }

    ctx.font = '13px IBM Plex Mono';
    let maxLabelW = 0;
    for (let t = 0; t <= 5; t++) {
      const v = yMin + (yMax - yMin) * t / 5;
      maxLabelW = Math.max(maxLabelW, ctx.measureText(fmtVal(v)).width);
    }
    const mg = {top: 44, right: 20, bottom: 80, left: Math.ceil(maxLabelW) + 20};
    const pw = W - mg.left - mg.right;
    const ph = H - mg.top  - mg.bottom;
    const toY = v => mg.top + ph - (v - yMin) / (yMax - yMin) * ph;

    // title
    ctx.fillStyle = color;
    ctx.font = 'bold 15px IBM Plex Mono';
    ctx.textAlign = 'center';
    ctx.fillText(title, W / 2, 26);

    // grid lines + y-axis ticks
    const ticks = 5;
    for (let t = 0; t <= ticks; t++) {
      const v = yMin + (yMax - yMin) * t / ticks;
      const y = toY(v);
      ctx.strokeStyle = gridCol; ctx.lineWidth = 0.5;
      ctx.beginPath(); ctx.moveTo(mg.left, y); ctx.lineTo(mg.left + pw, y); ctx.stroke();
      ctx.fillStyle = tickCol;
      ctx.font = '13px IBM Plex Mono';
      ctx.textAlign = 'right';
      ctx.fillText(fmtVal(v), mg.left - 8, y + 5);
    }

    // axes
    ctx.strokeStyle = axisCol; ctx.lineWidth = 1.5;
    ctx.beginPath();
    ctx.moveTo(mg.left, mg.top); ctx.lineTo(mg.left, mg.top + ph);
    ctx.lineTo(mg.left + pw, mg.top + ph); ctx.stroke();

    const slotW = pw / samples.length;
    const boxW = Math.min(slotW * 0.5, 70);

    const sampleData = samples.map(s => ({
      sample: s,
      vals: analysisData.filter(d => summaryGroupKey(d) === s).map(d => d[key]).filter(v => v != null && isFinite(v)),
    }));
    const stats = sampleData.map(sd => ({...sd, stats: computeStats(sd.vals)}));

    stats.forEach(({sample, vals, stats: st}, si) => {
      const cx = mg.left + (si + 0.5) * slotW;

      // x-axis label
      ctx.fillStyle = xLblCol;
      ctx.font = '13px IBM Plex Mono';
      ctx.textAlign = 'center';
      const maxChars = Math.floor(slotW / 8);
      const lbl = sample.length > maxChars ? sample.slice(0, maxChars - 1) + '…' : sample;
      ctx.fillText(lbl, cx, mg.top + ph + 22);

      if (!vals.length) return;

      // jittered data points
      const ptAlpha = isLight ? 'aa' : 'bb';
      vals.forEach(v => {
        const jitter = (Math.random() - 0.5) * boxW * 0.55;
        ctx.beginPath();
        ctx.arc(cx + jitter, toY(v), 4, 0, Math.PI * 2);
        ctx.fillStyle = color + ptAlpha; ctx.fill();
      });

      if (!st || vals.length < 2) return;

      const capW = boxW * 0.35;

      // whiskers + caps
      const wAlpha = isLight ? 'dd' : 'cc';
      ctx.strokeStyle = color + wAlpha; ctx.lineWidth = 1.5;
      ctx.beginPath();
      ctx.moveTo(cx, toY(st.min)); ctx.lineTo(cx, toY(st.q1));
      ctx.moveTo(cx, toY(st.q3)); ctx.lineTo(cx, toY(st.max));
      ctx.moveTo(cx - capW, toY(st.min)); ctx.lineTo(cx + capW, toY(st.min));
      ctx.moveTo(cx - capW, toY(st.max)); ctx.lineTo(cx + capW, toY(st.max));
      ctx.stroke();

      // IQR box
      const qy = toY(st.q3), qh = toY(st.q1) - toY(st.q3);
      const boxAlpha = isLight ? '28' : '30';
      ctx.fillStyle = color + boxAlpha; ctx.strokeStyle = color; ctx.lineWidth = 2;
      ctx.fillRect(cx - boxW/2, qy, boxW, qh);
      ctx.strokeRect(cx - boxW/2, qy, boxW, qh);

      // median line
      const my = toY(st.median);
      ctx.strokeStyle = medCol; ctx.lineWidth = 2.5;
      ctx.beginPath(); ctx.moveTo(cx - boxW/2, my); ctx.lineTo(cx + boxW/2, my); ctx.stroke();
    });

    // ── Tukey HSD CLD letters ──
    if (stats.length >= 2) {
      const tukey = tukeyHSD(stats.map(s => ({name: s.sample, vals: s.vals})));
      if (tukey) {
        const sortedIdx = [...Array(stats.length).keys()]
          .sort((a, b) => (stats[b].stats ? stats[b].stats.median : 0) -
                          (stats[a].stats ? stats[a].stats.median : 0));
        const cld = computeCLD(stats.length, tukey.sig, sortedIdx);
        ctx.font = 'bold 13px IBM Plex Mono';
        ctx.textAlign = 'center';
        ctx.fillStyle = isLight ? '#222222' : '#dddddd';
        stats.forEach(({vals, stats: st}, si) => {
          if (!vals.length || !cld[si]) return;
          const cx = mg.left + (si + 0.5) * slotW;
          const topY = st ? toY(st.max) : Math.min(...vals.map(v => toY(v)));
          ctx.fillText(cld[si], cx, topY - 6);
        });
      }
    }

    // ── legend ──
    ctx.font = '9px IBM Plex Mono';
    ctx.textAlign = 'left';
    ctx.fillStyle = isLight ? '#888888' : '#506070';
    ctx.fillText('Tukey HSD α=0.05 — groups sharing a letter are not significantly different', mg.left, H - 6);
  }

  // ── PDF export ──
  async function exportSummaryPDF() {
    if (!analysisData.length) { setStatus('No data to export', 'wrn'); return; }
    if (!window.jspdf) { setStatus('PDF library not loaded', 'wrn'); return; }
    const groups = [...new Set(analysisData.map(summaryGroupKey))];
    const {yMin, yMax} = summaryYRange();

    // Canvas sized to match the per-plot slot on landscape A4 (3 plots, 8mm margins)
    // A4 landscape: 297x210mm → slot ≈ 88mm wide × 194mm tall → ratio 0.454
    const CW = 720, CH = 1590;
    const dataURLs = SUMMARY_CHANNELS.map(ch => {
      const c = document.createElement('canvas');
      c.width = CW; c.height = CH;
      drawBoxPlot(c, groups, ch.key, ch.colorExport, ch.label, yMin, yMax, 'light');
      return c.toDataURL('image/png');
    });

    const {jsPDF} = window.jspdf;
    const pdf = new jsPDF({orientation: 'landscape', unit: 'mm', format: 'a4'});
    const PW = pdf.internal.pageSize.getWidth();
    const PH = pdf.internal.pageSize.getHeight();
    const margin = 8;
    const n = SUMMARY_CHANNELS.length;
    const plotW = (PW - margin * (n + 1)) / n;
    const plotH = plotW * CH / CW;           // maintain canvas aspect ratio — no distortion
    const yOff = (PH - plotH) / 2;          // centre vertically on page
    dataURLs.forEach((url, i) => {
      pdf.addImage(url, 'PNG', margin + i * (plotW + margin), yOff, plotW, plotH);
    });

    const pdfDataUrl = pdf.output('datauristring');
    if (lastSaveDir) {
      const r = await fetch('/save_pdf', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({dir: lastSaveDir, pdf_data: pdfDataUrl, filename: 'biolum_summary.pdf'})
      });
      const d = await r.json();
      if (d.ok) setStatus('PDF saved: ' + d.path, 'ok');
      else setStatus('PDF save failed: ' + d.error, 'err');
    } else {
      pdf.save('biolum_summary.pdf');
      setStatus('PDF downloaded — save measurements first to write to folder', 'wrn');
    }
  }

  // ── snapshot ──
  function captureSnapshotDataURL() {
    if (!dayImg && !biolumImg) return null;
    const dW = dayCanvas.width, dH = dayCanvas.height;
    const bW = biolumCanvas.width, bH = biolumCanvas.height;
    const gap = 4;
    const W = (dW||0) + (bW||0) + (dW && bW ? gap : 0);
    const H = Math.max(dH||0, bH||0);
    const off = document.createElement('canvas');
    off.width = W||800; off.height = H||600;
    const ctx = off.getContext('2d');
    ctx.fillStyle = '#06090d'; ctx.fillRect(0, 0, off.width, off.height);

    if (dayImg && dW) {
      ctx.drawImage(dayImg, 0, 0, dW, dH);
      ctx.save(); drawRoisOnCtx(ctx, 1, 1); ctx.restore();
    }
    const bx = (dW||0) + (dW && bW ? gap : 0);
    if (biolumImg && bW) {
      ctx.drawImage(biolumImg, bx, 0, bW, bH);
      ctx.save(); ctx.translate(bx, 0);
      drawRoisOnCtx(ctx, dW ? bW/dW : 1, dH ? bH/dH : 1);
      ctx.restore();
    }

    const addLabel = (text, x, y) => {
      ctx.font = '10px IBM Plex Mono';
      const tw = ctx.measureText(text).width;
      ctx.fillStyle = 'rgba(6,9,13,0.75)'; ctx.fillRect(x, y, tw+12, 18);
      ctx.fillStyle = '#4af0c4'; ctx.fillText(text, x+6, y+13);
    };
    if (dayImg && dW)    addLabel('DAY',    6,    6);
    if (biolumImg && bW) addLabel('BIOLUM', bx+6, 6);

    return off.toDataURL('image/png');
  }

  function saveSnapshot() {
    const dataURL = captureSnapshotDataURL();
    if (!dataURL) { setStatus('Load images first', 'wrn'); return; }
    const stem = biolumPair ? biolumPair.stem : (dayPair ? dayPair.stem : 'snapshot');
    const a = document.createElement('a');
    a.href = dataURL; a.download = stem + '_snapshot.png';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
    setStatus('Snapshot downloaded: ' + stem + '_snapshot.png', 'ok');
  }

  function drawRoisOnCtx(ctx, sx, sy) {
    sortRois();
    rois.forEach((roi, i) => {
      const color = ROI_COLORS[i % ROI_COLORS.length];
      const label = getRoiLabel(roi, i);
      const x = roi.x*sx, y = roi.y*sy, w = roi.w*sx, h = roi.h*sy;
      ctx.strokeStyle = color; ctx.lineWidth = 1.5;
      ctx.fillStyle = color + '22';
      if ((roi.shape||'rect') === 'circle') {
        ctx.beginPath(); ctx.ellipse(x+w/2, y+h/2, w/2, h/2, 0, 0, Math.PI*2);
        ctx.stroke(); ctx.fill();
      } else {
        ctx.strokeRect(x, y, w, h); ctx.fillRect(x, y, w, h);
      }
      ctx.fillStyle = color; ctx.font = 'bold 11px IBM Plex Mono';
      ctx.fillText(label, x+4, y+13);
    });
  }

  async function loadRoiFile(path) {
    document.getElementById('roi-dialog').style.display = 'none';
    const r = await fetch('/load_rois', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({path})
    });
    const data = await r.json();
    if (!data.ok) { setStatus('Failed to load ROIs: ' + data.error, 'err'); return; }

    const saved = data.data;
    const scaleX = dayCanvas.width / saved.jpeg_size[0];
    const scaleY = dayCanvas.height / saved.jpeg_size[1];

    rois = saved.rois.map(r => ({
      ...r,
      x: r.x * scaleX, y: r.y * scaleY,
      w: r.w * scaleX, h: r.h * scaleY,
    }));
    selectedRoi = -1;
    drawAll();
    setStatus(`Loaded ${rois.length} ROIs from ${saved.stem}`, 'ok');
  }
</script>
</body>
</html>"""

# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print()
    print("  Biolum Analysis Tool")
    print("  ─────────────────────────────────────────")
    print("  Opening browser at http://localhost:5001")
    print("  Press Ctrl+C to quit")
    print("  ─────────────────────────────────────────")
    print()

    def open_browser():
        import time
        time.sleep(1)
        webbrowser.open("http://localhost:5001")

    threading.Thread(target=open_browser, daemon=True).start()

    try:
        app.run(host="127.0.0.1", port=5001, threaded=True)
    except KeyboardInterrupt:
        print("\nBye!")
        sys.exit(0)
