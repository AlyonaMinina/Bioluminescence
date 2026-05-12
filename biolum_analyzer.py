#!/usr/bin/env python3
"""
Bioluminescence Image Analysis Tool
Run on your laptop to analyze images from experiment folders.

Usage:
    python biolum_analyzer.py

Then open: http://localhost:5001
"""

from flask import Flask, Response, request, jsonify, send_file
from pathlib import Path
from datetime import datetime
import numpy as np
import json
import io
import os
import re
import webbrowser
import threading
import sys
import uuid
import time

APP_DIR = Path(__file__).resolve().parent
APP_ICON = Path(os.environ.get("BIOLUM_ICON", APP_DIR / "biolum_icon.ico")).expanduser()

app = Flask(__name__)
MEASURE_JOBS = {}
MEASURE_JOBS_LOCK = threading.Lock()

# ── helpers ───────────────────────────────────────────────────────────────────

def detect_image_type(stem):
    s = stem.lower()
    if re.search(r'(^|_)tl\d+(_|$)', s):
        return 'biolum'
    if any(kw in s for kw in ('biolum', 'bio', 'dark', 'night', 'lum')):
        return 'biolum'
    if any(kw in s for kw in ('day', 'light', 'white', 'bright')):
        return 'day'
    return 'unknown'

def parse_timelapse_stem(stem):
    """Return stack metadata for names like sample_tl0001_1sec_193206."""
    m = re.match(r'^(?P<sample>.+?)_tl(?P<frame>\d+)_(?P<exposure>[^_]+sec)(?:_|$)', stem, re.I)
    if not m:
        return None
    return {
        "sample": m.group("sample").strip(),
        "frame": int(m.group("frame")),
        "exposure": m.group("exposure"),
    }

def parse_frame_timestamp(stem):
    """Return HHMMSS timestamp from the final filename token, if present."""
    m = re.search(r'_(\d{6})$', stem)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%H%M%S")
    except ValueError:
        return None

def find_pairs(folder):
    """Find JPEG+NEF pairs in a folder, grouped by stem."""
    folder = Path(folder)
    if not folder.exists():
        return []
    stems = {}
    for f in sorted(folder.iterdir()):
        if f.suffix.lower() in ('.jpg', '.jpeg', '.nef', '.raw'):
            stem = f.stem
            if stem not in stems:
                stems[stem] = {'stem': stem, 'jpg': None, 'nef': None,
                               'type': detect_image_type(stem)}
            if f.suffix.lower() in ('.jpg', '.jpeg'):
                stems[stem]['jpg'] = str(f)
            else:
                stems[stem]['nef'] = str(f)

    stacks = {}
    singles = []
    for item in stems.values():
        if not (item['jpg'] or item['nef']):
            continue
        tl = parse_timelapse_stem(item['stem'])
        if not tl:
            singles.append(item)
            continue

        key = (tl["sample"].lower(), tl["exposure"].lower())
        if key not in stacks:
            stacks[key] = {
                "stem": f"{tl['sample']}_{tl['exposure']}_timelapse",
                "sample": tl["sample"],
                "exposure": tl["exposure"],
                "type": "biolum",
                "is_stack": True,
                "frames": [],
                "jpg": None,
                "nef": None,
            }
        frame = {**item, "frame": tl["frame"], "sample": tl["sample"], "exposure": tl["exposure"]}
        stacks[key]["frames"].append(frame)

    for stack in stacks.values():
        stack["frames"].sort(key=lambda f: f["frame"])
        first = stack["frames"][0]
        stack["jpg"] = first.get("jpg")
        stack["nef"] = first.get("nef")

    return singles + list(stacks.values())

def get_experiments(base_folder):
    """Return experiment subfolders."""
    base = Path(base_folder)
    if not base.exists():
        return []
    return [str(f) for f in sorted(base.iterdir(), reverse=True) if f.is_dir()]

def measure_nef(nef_path, rois, jpeg_size, progress_callback=None):
    """
    Measure integrated density from NEF for each ROI.
    rois: list of {x, y, w, h} in JPEG pixel coordinates
    jpeg_size: (width, height) of the JPEG used for ROI drawing
    Returns list of measurements per ROI.
    """
    import rawpy

    with rawpy.imread(nef_path) as raw:
        # postprocess to get RGB array (full resolution)
        rgb = raw.postprocess(
            output_bps=16,
            no_auto_bright=True,
            use_camera_wb=False
        )

    nef_h, nef_w = rgb.shape[:2]
    jpeg_w, jpeg_h = jpeg_size

    scale_x = nef_w / jpeg_w
    scale_y = nef_h / jpeg_h

    results = []
    for i, roi in enumerate(rois):
        if progress_callback:
            progress_callback(i, len(rois), roi, "measuring")

        # scale ROI coordinates to NEF resolution
        x1 = max(0, int(roi['x'] * scale_x))
        y1 = max(0, int(roi['y'] * scale_y))
        x2 = min(nef_w, int((roi['x'] + roi['w']) * scale_x))
        y2 = min(nef_h, int((roi['y'] + roi['h']) * scale_y))

        region = rgb[y1:y2, x1:x2]
        rh, rw = region.shape[:2]

        shape = roi.get('shape', 'rect')
        if shape == 'circle' and rh > 0 and rw > 0:
            cy_c, cx_c = rh / 2.0, rw / 2.0
            ry_c, rx_c = rh / 2.0, rw / 2.0
            Y, X = np.ogrid[:rh, :rw]
            mask = ((X - cx_c) / (rx_c + 1e-9))**2 + ((Y - cy_c) / (ry_c + 1e-9))**2 <= 1.0
        else:
            mask = np.ones((rh, rw), dtype=bool)

        area = int(np.sum(mask))

        r_ch = region[:, :, 0].astype(np.float64)
        g_ch = region[:, :, 1].astype(np.float64)
        b_ch = region[:, :, 2].astype(np.float64)

        results.append({
            'roi_number': i + 1,
            'roi_x_jpeg': roi['x'],
            'roi_y_jpeg': roi['y'],
            'roi_w_jpeg': roi['w'],
            'roi_h_jpeg': roi['h'],
            'roi_x_nef': x1,
            'roi_y_nef': y1,
            'roi_w_nef': x2 - x1,
            'roi_h_nef': y2 - y1,
            'area_px': area,
            'mean_R': float(np.mean(r_ch[mask])),
            'mean_G': float(np.mean(g_ch[mask])),
            'mean_B': float(np.mean(b_ch[mask])),
            'intden_R': float(np.sum(r_ch[mask])),
            'intden_G': float(np.sum(g_ch[mask])),
            'intden_B': float(np.sum(b_ch[mask])),
        })
        if progress_callback:
            progress_callback(i + 1, len(rois), roi, "finished")

    return results

def _job_update(job_id, **updates):
    with MEASURE_JOBS_LOCK:
        job = MEASURE_JOBS.setdefault(job_id, {})
        job.update(updates)

def _job_status(job_id, status):
    with MEASURE_JOBS_LOCK:
        job = MEASURE_JOBS.get(job_id, {})
        return job.get("status") == status

def _job_snapshot(job_id):
    with MEASURE_JOBS_LOCK:
        job = MEASURE_JOBS.get(job_id)
        return dict(job) if job else None

def _roi_display_name(roi, index):
    roi_number = roi.get("roi_number") or index + 1
    if roi.get("type") == "bckg":
        return f"background ROI {roi_number}"
    return f"ROI {roi_number}"

def _measure_job_worker(job_id, data):
    mode = data.get("mode", "single")
    rois = data.get("rois", [])
    jpeg_size = data.get("jpeg_size", [1, 1])
    all_measurements = []

    try:
        if mode == "stack":
            frames = data.get("frames", [])
            valid_frames = [f for f in frames if f.get("nef") and Path(f.get("nef")).exists()]
            if not valid_frames:
                raise ValueError("No NEF files found in stack")
            valid_frames = sorted(valid_frames, key=lambda f: int(f.get("frame", 0)))
            total = len(valid_frames) * len(rois)
            _job_update(job_id, status="running", done=0, total=total,
                        message=f"Preparing {len(valid_frames)} frames and {len(rois)} ROIs...")

            base_ts = parse_frame_timestamp(valid_frames[0].get("stem", ""))
            prev_ts = base_ts
            day_offset = 0

            for frame_idx, frame in enumerate(valid_frames):
                stem = frame.get("stem", "")
                ts = parse_frame_timestamp(stem)
                if frame_idx == 0 or not base_ts or not ts:
                    elapsed_min = 0.0
                else:
                    if prev_ts and ts < prev_ts:
                        day_offset += 1
                    elapsed_sec = (ts - base_ts).total_seconds() + day_offset * 24 * 3600
                    elapsed_min = elapsed_sec / 60.0
                if ts:
                    prev_ts = ts

                frame_label = f"frame {frame_idx + 1} of {len(valid_frames)}"
                frame_base = frame_idx * len(rois)

                def progress(done_in_frame, total_in_frame, roi, phase):
                    roi_pos = done_in_frame - 1 if phase == "finished" else done_in_frame
                    roi_idx = min(max(roi_pos, 0), max(total_in_frame - 1, 0))
                    roi_name = _roi_display_name(roi, roi_idx)
                    done = frame_base + done_in_frame
                    if phase == "finished":
                        done = frame_base + done_in_frame
                        message = f"Finished {roi_name} in {frame_label}"
                    else:
                        message = f"Measuring {roi_name} in {frame_label}"
                    _job_update(job_id, done=done, total=total, message=message)

                frame_results = measure_nef(frame["nef"], rois, jpeg_size, progress_callback=progress)
                for m in frame_results:
                    m["frame_number"] = int(frame.get("frame", frame_idx + 1))
                    m["frame_index"] = frame_idx
                    m["elapsed_min"] = float(elapsed_min)
                    m["frame_stem"] = stem
                    m["nef_path"] = frame.get("nef", "")
                all_measurements.extend(frame_results)
                _job_update(job_id, done=frame_base + len(rois), total=total,
                            message=f"Finished {frame_label}")
        else:
            nef_path = data.get("nef_path", "")
            if not nef_path or not Path(nef_path).exists():
                raise ValueError("NEF file not found")
            total = len(rois)
            _job_update(job_id, status="running", done=0, total=total,
                        message=f"Preparing single image with {total} ROIs...")

            def progress(done, total_rois, roi, phase):
                roi_pos = done - 1 if phase == "finished" else done
                roi_idx = min(max(roi_pos, 0), max(total_rois - 1, 0))
                roi_name = _roi_display_name(roi, roi_idx)
                message = f"Finished {roi_name}" if phase == "finished" else f"Measuring {roi_name} of {total_rois}"
                _job_update(job_id, done=done, total=total_rois, message=message)
                time.sleep(0.035)

            all_measurements = measure_nef(nef_path, rois, jpeg_size, progress_callback=progress)

        final_total = max(1, len(rois)) if mode != "stack" else max(1, len(valid_frames) * len(rois))
        _job_update(job_id, status="finishing", done=final_total, total=final_total,
                    message="Measurement complete")
        time.sleep(0.35)
        if _job_status(job_id, "finishing"):
            _job_update(job_id, status="done", done=final_total, total=final_total,
                        message="Measurement complete", measurements=all_measurements)
    except Exception as e:
        _job_update(job_id, status="error", error=str(e), message=f"Error: {e}")

def save_results(folder, stem, rois, measurements, jpeg_size, snapshot_png="", save_dir=""):
    """Save ROIs as JSON, snapshot PNG, and measurements as Excel."""
    if save_dir:
        analysis_dir = Path(save_dir)
        analysis_dir.mkdir(parents=True, exist_ok=True)
    else:
        analysis_dir = Path(folder) / f"Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}"
        analysis_dir.mkdir(parents=True, exist_ok=True)

    # Save ROIs
    roi_file = analysis_dir / f"{stem}_rois.json"
    with open(roi_file, 'w') as f:
        json.dump({'stem': stem, 'jpeg_size': jpeg_size, 'rois': rois}, f, indent=2)

    # Save snapshot PNG
    if snapshot_png and ',' in snapshot_png:
        import base64
        img_bytes = base64.b64decode(snapshot_png.split(',', 1)[1])
        with open(analysis_dir / f"{stem}_snapshot.png", 'wb') as f:
            f.write(img_bytes)

    if not measurements:
        return str(analysis_dir)

    # Excel
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    fills = {
        'B':  PatternFill(start_color="E8EEFF", end_color="E8EEFF", fill_type="solid"),
        'G':  PatternFill(start_color="E8FFE8", end_color="E8FFE8", fill_type="solid"),
        'R':  PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"),
        'Bn': PatternFill(start_color="D0DDFF", end_color="D0DDFF", fill_type="solid"),
        'Gn': PatternFill(start_color="D0FFD8", end_color="D0FFD8", fill_type="solid"),
        'Rn': PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid"),
    }
    data_font = Font(name="Calibri")
    bckg_font = Font(name="Calibri", italic=True, color="8B6914")
    row_border = Border(bottom=Side(style='thin', color="D0D0D0"))

    def _r(v, d=1):
        return round(v, d) if isinstance(v, (int, float)) and v is not None else ''

    xlsx_file = analysis_dir / "session_measurements.xlsx"

    if xlsx_file.exists():
        wb = load_workbook(xlsx_file)
        ws = wb.active
        existing_headers = [cell.value for cell in ws[1]]
        has_frame = 'Frame' in existing_headers
        has_elapsed = 'Elapsed (min)' in existing_headers
        has_nb = 'Mean B-Bckg' in existing_headers
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Measurements"

        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri")
        header_border = Border(bottom=Side(style='medium', color="2E74B5"))

        has_nb = any(m.get('mean_B_nb') is not None for m in measurements)
        has_frame = any(m.get('frame_number') is not None for m in measurements)
        has_elapsed = any(m.get('elapsed_min') is not None for m in measurements)

        headers = ['Sample', 'ROI']
        if has_frame:
            headers += ['Frame']
        if has_elapsed:
            headers += ['Elapsed (min)']
        headers += ['Area (px)', 'Raw Mean B', 'Raw Mean G', 'Raw Mean R']
        if has_nb:
            headers += ['Mean B-Bckg', 'Mean G-Bckg', 'Mean R-Bckg']

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = header_border

        ws.freeze_panes = "A2"
        for col in ws.columns:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = 28 if letter == 'A' else 12

    # col index (1-based) of first channel column
    ch_base = 4 + int(has_frame) + int(has_elapsed)

    for m in measurements:
        is_bckg = m.get('roi_type') == 'bckg'
        sample = m.get('sample_name') or stem
        row = [sample, m.get('roi_label', '')]
        if has_frame:
            row.append(m.get('frame_number', ''))
        if has_elapsed:
            row.append(_r(m.get('elapsed_min'), 2))
        row += [m.get('area_px', 0), _r(m.get('mean_B')), _r(m.get('mean_G')), _r(m.get('mean_R'))]
        if has_nb:
            row += [_r(m.get('mean_B_nb')), _r(m.get('mean_G_nb')), _r(m.get('mean_R_nb'))]
        ws.append(row)
        rn = ws.max_row
        for cell in ws[rn]:
            cell.font = bckg_font if is_bckg else data_font
            cell.border = row_border
        col_map = [(ch_base,'B'),(ch_base+1,'G'),(ch_base+2,'R')]
        if has_nb:
            col_map += [(ch_base+3,'Bn'),(ch_base+4,'Gn'),(ch_base+5,'Rn')]
        for col, ch in col_map:
            ws.cell(rn, col).fill = fills[ch]

    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_file)

    return str(analysis_dir)

# ── routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return HTML_PAGE

@app.route("/app_icon.ico")
@app.route("/favicon.ico")
def app_icon():
    if APP_ICON.exists():
        return send_file(APP_ICON, mimetype="image/vnd.microsoft.icon")
    return "", 404

@app.route("/browse", methods=["POST"])
def browse():
    data = request.get_json()
    folder = data.get("folder", "")
    pairs = find_pairs(folder)
    analysis_dir = ""
    try:
        p = Path(folder)
        if p.exists():
            analysis_dir = str(p / f"Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}")
    except Exception:
        analysis_dir = ""
    return jsonify({"pairs": pairs, "folder": folder, "analysis_dir": analysis_dir})

@app.route("/ls", methods=["POST"])
def ls():
    """List directories and image files in a path for the folder navigator."""
    data = request.get_json()
    path = data.get("path", "")

    # default to home / drives on Windows
    if not path:
        if sys.platform == "win32":
            import string
            drives = [f"{d}:\\" for d in string.ascii_uppercase
                      if Path(f"{d}:\\").exists()]
            return jsonify({"path": "", "parent": None,
                            "dirs": [{"name": d, "path": d} for d in drives],
                            "has_images": False})
        else:
            path = str(Path.home())

    p = Path(path)
    if not p.exists() or not p.is_dir():
        return jsonify({"error": "Not a directory"}), 400

    try:
        dirs = []
        for item in sorted(p.iterdir()):
            try:
                if item.is_dir() and not item.name.startswith('.'):
                    dirs.append({"name": item.name, "path": str(item)})
            except PermissionError:
                pass

        # check if folder has images
        has_images = any(
            f.suffix.lower() in ('.jpg','.jpeg','.nef','.raw')
            for f in p.iterdir()
            if f.is_file()
        )

        parent = str(p.parent) if p.parent != p else None
        return jsonify({
            "path": str(p),
            "parent": parent,
            "dirs": dirs,
            "has_images": has_images
        })
    except PermissionError:
        return jsonify({"error": "Permission denied"}), 403

@app.route("/experiments", methods=["POST"])
def experiments():
    data = request.get_json()
    base = data.get("folder", "")
    exps = get_experiments(base)
    return jsonify({"experiments": exps})

@app.route("/image")
def serve_image():
    path = request.args.get("path", "")
    try:
        return send_file(path, mimetype="image/jpeg")
    except Exception:
        return "", 404

@app.route("/image_size")
def image_size():
    path = request.args.get("path", "")
    try:
        from PIL import Image
        with Image.open(path) as img:
            return jsonify({"width": img.width, "height": img.height})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/measure", methods=["POST"])
def measure():
    data = request.get_json()
    nef_path = data.get("nef_path", "")
    rois = data.get("rois", [])
    jpeg_size = data.get("jpeg_size", [1, 1])

    if not nef_path or not Path(nef_path).exists():
        return jsonify({"error": "NEF file not found"}), 400
    if not rois:
        return jsonify({"error": "No ROIs defined"}), 400

    try:
        results = measure_nef(nef_path, rois, jpeg_size)
        return jsonify({"measurements": results})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/measure_stack", methods=["POST"])
def measure_stack():
    data = request.get_json()
    frames = data.get("frames", [])
    rois = data.get("rois", [])
    jpeg_size = data.get("jpeg_size", [1, 1])

    if not frames:
        return jsonify({"error": "No stack frames provided"}), 400
    if not rois:
        return jsonify({"error": "No ROIs defined"}), 400

    valid_frames = [f for f in frames if f.get("nef") and Path(f.get("nef")).exists()]
    if not valid_frames:
        return jsonify({"error": "No NEF files found in stack"}), 400

    valid_frames = sorted(valid_frames, key=lambda f: int(f.get("frame", 0)))
    base_ts = parse_frame_timestamp(valid_frames[0].get("stem", ""))
    prev_ts = base_ts
    day_offset = 0

    all_measurements = []
    try:
        for idx, frame in enumerate(valid_frames):
            stem = frame.get("stem", "")
            ts = parse_frame_timestamp(stem)
            if idx == 0 or not base_ts or not ts:
                elapsed_min = 0.0
            else:
                if prev_ts and ts < prev_ts:
                    day_offset += 1
                elapsed_sec = (ts - base_ts).total_seconds() + day_offset * 24 * 3600
                elapsed_min = elapsed_sec / 60.0
            if ts:
                prev_ts = ts

            frame_results = measure_nef(frame["nef"], rois, jpeg_size)
            for m in frame_results:
                m["frame_number"] = int(frame.get("frame", idx + 1))
                m["frame_index"] = idx
                m["elapsed_min"] = float(elapsed_min)
                m["frame_stem"] = stem
                m["nef_path"] = frame.get("nef", "")
            all_measurements.extend(frame_results)
        return jsonify({"measurements": all_measurements})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/measure_job", methods=["POST"])
def measure_job():
    data = request.get_json()
    mode = data.get("mode", "single")
    rois = data.get("rois", [])

    if not rois:
        return jsonify({"error": "No ROIs defined"}), 400
    if mode == "stack":
        frames = data.get("frames", [])
        if not frames:
            return jsonify({"error": "No stack frames provided"}), 400
    else:
        nef_path = data.get("nef_path", "")
        if not nef_path or not Path(nef_path).exists():
            return jsonify({"error": "NEF file not found"}), 400

    job_id = uuid.uuid4().hex
    _job_update(job_id, status="queued", done=0, total=max(1, len(rois)),
                message="Queued measurement...", measurements=[])
    worker = threading.Thread(target=_measure_job_worker, args=(job_id, data), daemon=True)
    worker.start()
    return jsonify({"job_id": job_id})

@app.route("/measure_job/<job_id>")
def measure_job_status(job_id):
    job = _job_snapshot(job_id)
    if not job:
        return jsonify({"error": "Measurement job not found"}), 404
    return jsonify(job)

@app.route("/save", methods=["POST"])
def save():
    data = request.get_json()
    folder = data.get("folder", "")
    save_dir = data.get("save_dir") or ""
    stem = data.get("stem", "analysis")
    rois = data.get("rois", [])
    measurements = data.get("measurements", [])
    session_measurements = data.get("session_measurements") or measurements
    jpeg_size = data.get("jpeg_size", [1, 1])
    snapshot_png = data.get("snapshot_png", "")

    try:
        out_dir = save_results(folder, stem, rois, session_measurements, jpeg_size,
                               snapshot_png, save_dir=save_dir)
        return jsonify({"ok": True, "dir": out_dir})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/save_pdf", methods=["POST"])
def save_pdf():
    import base64
    data = request.get_json()
    save_dir = data.get("dir", "")
    pdf_data = data.get("pdf_data", "")
    filename = data.get("filename", "biolum_summary.pdf")
    if not save_dir or not pdf_data:
        return jsonify({"ok": False, "error": "Missing dir or pdf_data"})
    try:
        p = Path(save_dir)
        p.mkdir(parents=True, exist_ok=True)
        raw = pdf_data.split(',', 1)[-1]
        pdf_bytes = base64.b64decode(raw)
        out_path = p / filename
        with open(out_path, 'wb') as f:
            f.write(pdf_bytes)
        return jsonify({"ok": True, "path": str(out_path)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/load_rois", methods=["POST"])
def load_rois():
    data = request.get_json()
    path = data.get("path", "")
    try:
        with open(path, 'r') as f:
            roi_data = json.load(f)
        return jsonify({"ok": True, "data": roi_data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/autosave_rois", methods=["POST"])
def autosave_rois():
    data = request.get_json()
    folder = data.get("folder", "")
    save_dir = data.get("save_dir") or ""
    stem = data.get("stem", "analysis")
    rois = data.get("rois", [])
    jpeg_size = data.get("jpeg_size", [1, 1])
    try:
        if save_dir:
            out_dir = Path(save_dir)
        else:
            out_dir = Path(folder) / f"Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{stem}_rois.json"
        with open(out_path, 'w') as f:
            json.dump({
                'stem': stem,
                'jpeg_size': jpeg_size,
                'rois': rois,
                'autosaved': True,
                'saved_at': datetime.now().isoformat(timespec='seconds'),
            }, f, indent=2)
        return jsonify({"ok": True, "path": str(out_path), "dir": str(out_dir)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/list_roi_files", methods=["POST"])
def list_roi_files():
    data = request.get_json()
    folder = data.get("folder", "")
    files = []
    try:
        root = Path(folder)
        if not root.exists():
            return jsonify({"files": [], "folder": folder, "error": "Folder not found"})
        for p in sorted(root.rglob("*_rois.json"), reverse=True):
            files.append({"name": p.name, "path": str(p), "folder": str(p.parent)})
        return jsonify({"files": files, "folder": str(root)})
    except Exception as e:
        return jsonify({"files": [], "folder": folder, "error": str(e)})

@app.route("/download")
def download():
    path = request.args.get("path", "")
    try:
        return send_file(path, as_attachment=True)
    except Exception:
        return "", 404

@app.route("/pick_folder")
def pick_folder():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    if APP_ICON.exists():
        try:
            root.iconbitmap(str(APP_ICON))
        except tk.TclError:
            pass
    root.title("Biolum Analyzer")
    root.withdraw()
    root.wm_attributes('-topmost', True)
    folder = filedialog.askdirectory(title="Select Experiment Folder")
    root.destroy()
    return jsonify({"path": folder or ""})

# ── HTML ──────────────────────────────────────────────────────────────────────

HTML_PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Biolum Analyzer</title>
  <link rel="icon" href="/favicon.ico" type="image/x-icon">
  <link rel="shortcut icon" href="/favicon.ico" type="image/x-icon">
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@300;400;500&family=IBM+Plex+Sans:wght@300;400;600&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    :root {
      --bg:      #06090d;
      --panel:   #0b0f15;
      --panel2:  #0f1520;
      --border:  #162030;
      --border2: #1e2d42;
      --accent:  #4af0c4;
      --amber:   #f0c44a;
      --red:     #f04a6a;
      --text:    #c0d8f0;
      --muted:   #385060;
      --mono:    'IBM Plex Mono', monospace;
      --sans:    'IBM Plex Sans', sans-serif;
    }

    html, body {
      height: 100%;
      background: var(--bg);
      color: var(--text);
      font-family: var(--sans);
      font-size: 13px;
      overflow: hidden;
    }

    /* ── layout ── */
    .app {
      display: grid;
      grid-template-rows: 48px 1fr;
      height: 100vh;
    }

    header {
      display: flex;
      align-items: center;
      gap: 16px;
      padding: 0 20px;
      border-bottom: 1px solid var(--border);
      background: var(--panel);
    }

    .brand {
      display: flex;
      align-items: center;
      gap: 8px;
      font-family: var(--mono);
      font-size: 13px;
      letter-spacing: 0.15em;
      color: var(--accent);
      text-transform: uppercase;
      flex-shrink: 0;
    }
    .brand-icon {
      width: 18px;
      height: 18px;
      object-fit: contain;
      flex-shrink: 0;
    }
    .brand-fallback {
      display: none;
      font-size: 13px;
      line-height: 1;
    }

    .header-path {
      font-family: var(--mono);
      font-size: 11px;
      color: var(--muted);
      flex: 1;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    .body {
      display: grid;
      grid-template-columns: 280px 1fr;
      overflow: hidden;
    }

    /* tabs-container fills the app grid's second row */
    .tabs-container { height: 100%; }

    /* ── sidebar ── */
    .sidebar {
      border-right: 1px solid var(--border);
      background: var(--panel);
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }

    .sidebar-section {
      border-bottom: 1px solid var(--border);
      flex-shrink: 0;
    }

    .sidebar-section-header {
      padding: 8px 14px;
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.18em;
      color: var(--muted);
      text-transform: uppercase;
      background: rgba(255,255,255,0.015);
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    .pair-list {
      flex: 1;
      overflow-y: auto;
      padding: 6px;
    }

    .dir-item {
      padding: 5px 10px;
      cursor: pointer;
      color: var(--text);
      display: flex;
      align-items: center;
      gap: 6px;
      border-bottom: 1px solid var(--border);
      transition: background 0.1s;
    }
    .dir-item:hover { background: rgba(74,240,196,0.05); }
    .dir-item.has-images { color: var(--accent); }
    .dir-item.parent { color: var(--amber); }
    .dir-item .dir-icon { font-size: 12px; flex-shrink: 0; }

    .pair-item {
      padding: 8px 10px;
      border-radius: 3px;
      cursor: pointer;
      font-family: var(--mono);
      font-size: 11px;
      color: var(--text);
      border: 1px solid transparent;
      transition: all 0.15s;
      margin-bottom: 3px;
      word-break: break-word;
      line-height: 1.5;
    }

    .pair-item:hover { background: rgba(74,240,196,0.05); border-color: var(--border2); }
    .pair-item.active-day    { background: rgba(240,196,74,0.08); border-color: var(--amber); color: var(--amber); }
    .pair-item.active-biolum { background: rgba(74,240,196,0.08); border-color: var(--accent); color: var(--accent); }

    .pair-badges {
      display: flex;
      gap: 4px;
      margin-top: 4px;
    }

    .badge {
      font-size: 9px;
      padding: 1px 5px;
      border-radius: 2px;
      letter-spacing: 0.08em;
      font-family: var(--mono);
    }
    .badge-jpg { background: rgba(74,240,196,0.15); color: var(--accent); }
    .badge-nef { background: rgba(240,196,74,0.15); color: var(--amber); }
    .badge-none { background: rgba(240,74,106,0.15); color: var(--red); }
    .badge-day { background: rgba(240,196,74,0.2); color: var(--amber); }
    .badge-biolum { background: rgba(74,240,196,0.2); color: var(--accent); }
    .badge-unknown { background: rgba(192,216,240,0.1); color: var(--muted); }

    /* ── folder input ── */
    .folder-input-wrap {
      padding: 10px;
      display: flex;
      flex-direction: column;
      gap: 6px;
    }

    input[type="text"] {
      width: 100%;
      background: var(--bg);
      border: 1px solid var(--border2);
      border-radius: 3px;
      color: var(--text);
      font-family: var(--mono);
      font-size: 11px;
      padding: 6px 10px;
      outline: none;
    }
    input[type="text"]:focus { border-color: var(--accent); }

    .btn {
      font-family: var(--mono);
      font-size: 11px;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      padding: 7px 12px;
      border-radius: 3px;
      border: 1px solid;
      cursor: pointer;
      transition: all 0.15s;
      width: 100%;
    }
    .btn-primary { background: #000; border-color: var(--accent); color: var(--accent); font-weight: 600; }
    .btn-primary:hover { background: rgba(74,240,196,0.08); }
    .btn-primary:disabled { opacity: 0.3; cursor: not-allowed; }
    .btn-ghost { background: transparent; border-color: var(--border2); color: var(--text); }
    .btn-ghost:hover { border-color: var(--accent); color: var(--accent); }
    .btn-amber { background: rgba(240,196,74,0.1); border-color: var(--amber); color: var(--amber); }
    .btn-amber:hover { background: rgba(240,196,74,0.2); }
    .btn-red { background: transparent; border-color: var(--red); color: var(--red); }
    .btn-red:hover { background: rgba(240,74,106,0.1); }
    .btn-sm { padding: 4px 10px; font-size: 10px; width: auto; }

    /* ── main panel ── */
    .main {
      display: grid;
      grid-template-rows: 1fr 220px;
      overflow: hidden;
    }

    /* ── canvas area ── */
    .canvas-area {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 2px;
      background: var(--border);
      overflow: hidden;
      position: relative;
    }

    .canvas-wrap {
      position: relative;
      background: #000;
      overflow: hidden;
      display: flex;
      align-items: center;
      justify-content: center;
    }

    .canvas-label {
      position: absolute;
      top: 10px;
      left: 50%;
      transform: translateX(-50%);
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.15em;
      color: #fff;
      background: rgba(6,9,13,0.85);
      padding: 3px 10px;
      border-radius: 2px;
      text-transform: uppercase;
      font-weight: 600;
      pointer-events: none;
      z-index: 10;
    }

    .canvas-clear-btn {
      position: absolute;
      top: 10px;
      right: 10px;
      background: rgba(240,74,106,0.15);
      border: 1px solid var(--red);
      color: var(--red);
      font-family: var(--mono);
      font-size: 10px;
      padding: 2px 8px;
      border-radius: 2px;
      cursor: pointer;
      z-index: 10;
      transition: background 0.15s;
    }
    .canvas-clear-btn:hover { background: rgba(240,74,106,0.35); }

    .canvas-sample {
      position: absolute;
      bottom: 10px;
      left: 50%;
      transform: translateX(-50%);
      font-family: var(--mono);
      font-size: 11px;
      color: var(--accent);
      background: rgba(6,9,13,0.85);
      padding: 3px 12px;
      border-radius: 2px;
      pointer-events: none;
      z-index: 10;
      white-space: nowrap;
      max-width: 90%;
      overflow: hidden;
      text-overflow: ellipsis;
    }

    canvas {
      max-width: 100%;
      max-height: 100%;
      cursor: crosshair;
      display: block;
    }

    .no-image {
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      gap: 8px;
      color: var(--muted);
      font-family: var(--mono);
      font-size: 11px;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      width: 100%;
      height: 100%;
    }

    /* ── bottom panel ── */
    .bottom-panel {
      border-top: 1px solid var(--border);
      display: flex;
      flex-direction: column;
      overflow: hidden;
    }

    .roi-toolbar {
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 5px 12px;
      border-bottom: 1px solid var(--border);
      background: var(--panel);
      flex-shrink: 0;
      flex-wrap: wrap;
    }

    .roi-toolbar-label {
      font-family: var(--mono);
      font-size: 9px;
      letter-spacing: 0.18em;
      color: var(--muted);
      text-transform: uppercase;
      flex-shrink: 0;
    }

    .roi-mode-row {
      display: flex;
      gap: 4px;
      align-items: center;
    }

    .mode-btn {
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      padding: 4px 8px;
      border-radius: 3px;
      border: 1px solid var(--text);
      cursor: pointer;
      color: var(--text);
      background: transparent;
      transition: all 0.15s;
      text-align: center;
      white-space: nowrap;
    }
    .mode-btn.active { border-color: var(--accent); color: var(--accent); background: rgba(74,240,196,0.07); }
    .mode-btn:hover:not(.active) { background: rgba(255,255,255,0.04); }

    .checkbox-row {
      display: flex;
      align-items: center;
      gap: 6px;
      font-family: var(--mono);
      font-size: 11px;
      color: var(--text);
      cursor: pointer;
    }
    .checkbox-row input { accent-color: var(--accent); }

    /* ── results table ── */
    .results-area {
      overflow-y: auto;
      background: var(--panel2);
      padding: 8px 12px;
    }

    .results-header {
      font-family: var(--mono);
      font-size: 10px;
      letter-spacing: 0.15em;
      color: var(--muted);
      text-transform: uppercase;
      margin-bottom: 8px;
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    table {
      width: 100%;
      border-collapse: collapse;
      font-family: var(--mono);
      font-size: 11px;
    }

    th {
      text-align: left;
      padding: 4px 8px;
      font-size: 10px;
      letter-spacing: 0.1em;
      color: var(--muted);
      text-transform: uppercase;
      border-bottom: 1px solid var(--border);
      white-space: nowrap;
    }

    td {
      padding: 4px 8px;
      border-bottom: 1px solid var(--border);
      color: #ddeeff;
      white-space: nowrap;
    }

    td.ch-r { color: #ff9999; }
    td.ch-g { color: #99ffbb; }
    td.ch-b { color: #99bbff; }
    td.ch-r-nb { color: #ffcccc; font-style: italic; }
    td.ch-g-nb { color: #ccffdd; font-style: italic; }
    td.ch-b-nb { color: #ccdeff; font-style: italic; }
    td.bckg-row { color: var(--amber); }

    tr:hover td { background: rgba(255,255,255,0.02); }

    .status-bar {
      font-family: var(--mono);
      font-size: 11px;
      color: var(--muted);
      padding: 4px 0;
    }
    .status-ok  { color: var(--accent); }
    .status-err { color: var(--red); }
    .status-wrn { color: var(--amber); }

    /* ── top-level view tabs ── */
    .top-tab-btn {
      font-family: var(--mono);
      font-size: 11px;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      padding: 0 18px;
      height: 100%;
      border: none;
      border-bottom: 2px solid transparent;
      background: transparent;
      color: var(--muted);
      cursor: pointer;
      transition: all 0.15s;
      flex-shrink: 0;
    }
    .top-tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); }
    .top-tab-btn:hover:not(.active) { color: var(--text); }

    .tabs-container { overflow: hidden; display: grid; }
    .top-tab-pane { display: none; overflow: hidden; }
    #top-tab-measure.active {
      display: grid;
      grid-template-columns: 280px 1fr;
    }
    #top-tab-summary.active {
      display: flex;
      flex-direction: column;
      background: var(--panel2);
      padding: 16px;
      gap: 12px;
    }
  </style>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js"></script>
</head>
<body>
<div class="app">

  <header>
    <div class="brand">
      <img class="brand-icon" src="/app_icon.ico" alt="" onerror="this.style.display='none';this.nextElementSibling.style.display='inline';">
      <span class="brand-fallback">◉</span>
      <span>Biolum Analyzer</span>
    </div>
    <button class="top-tab-btn active" id="top-tab-btn-measure" onclick="switchTopTab('measure')">Measurement</button>
    <button class="top-tab-btn" id="top-tab-btn-summary" onclick="switchTopTab('summary')">Analysis Summary</button>
    <div class="header-path" id="header-path">No folder loaded</div>
    <div style="display:flex;gap:8px;flex-shrink:0;">
      <button class="btn btn-ghost btn-sm" onclick="measureAll()" id="btn-measure">▶ Measure ROIs</button>
      <button class="btn btn-amber btn-sm" onclick="saveResults()" id="btn-save">↓ Save Results</button>
      <button class="btn btn-ghost btn-sm" onclick="showLoadRoiDialog()">↑ Load ROIs</button>
    </div>
  </header>

  <div class="tabs-container">
  <div class="top-tab-pane active" id="top-tab-measure">

    <!-- sidebar -->
    <div class="sidebar">
      <div class="sidebar-section">
        <div class="sidebar-section-header" id="experiment-label" style="font-size:11px;color:var(--accent);">
          No experiment loaded
        </div>
        <div style="padding:6px 8px;border-bottom:1px solid var(--border);display:flex;flex-direction:column;gap:4px;">
          <button class="btn btn-primary btn-sm" style="width:100%;" onclick="pickFolder()">Browse...</button>
          <div style="display:flex;gap:4px;">
            <input type="text" id="folder-input" placeholder="Or paste path..."
                   style="flex:1;font-size:10px;padding:4px 8px;"
                   onkeydown="if(event.key==='Enter') loadFolder(this.value)">
            <button class="btn btn-ghost btn-sm" onclick="loadFolder(document.getElementById('folder-input').value)" style="flex-shrink:0;">Go</button>
          </div>
        </div>
      </div>
      <div class="sidebar-section" style="padding:8px 10px;">
        <div class="status-bar" id="status">Ready. Load an experiment folder to begin.</div>
      </div>
      <div class="sidebar-section-header" style="padding:8px 14px;font-size:11px;color:var(--text);letter-spacing:0.1em;">
        Images within experiment
        <span style="font-size:10px;color:var(--accent);" id="pair-count"></span>
      </div>
      <div class="pair-list" id="pair-list">
        <div style="padding:20px;text-align:center;font-family:var(--mono);font-size:11px;color:var(--muted);">
          No folder loaded
        </div>
      </div>
    </div>

    <!-- main -->
    <div class="main">

      <!-- canvases -->
      <div class="canvas-area">
        <div class="canvas-wrap" id="day-wrap">
          <div class="canvas-label">DAY</div>
          <div class="no-image" id="day-placeholder">
            <span style="font-size:24px;">☀</span>
            <span>Select a day image</span>
          </div>
          <canvas id="day-canvas" style="display:none"></canvas>
          <button class="canvas-clear-btn" id="day-clear-btn" onclick="clearPanel('day')" style="display:none;" title="Clear day image">✕ Clear</button>
          <div class="canvas-sample" id="day-sample-label" style="display:none;"></div>
        </div>
        <div class="canvas-wrap" id="biolum-wrap">
          <div class="canvas-label">BIOLUM</div>
          <div class="no-image" id="biolum-placeholder">
            <span style="font-size:24px;">◉</span>
            <span>Select a biolum image</span>
          </div>
          <canvas id="biolum-canvas" style="display:none"></canvas>
          <button class="canvas-clear-btn" id="biolum-clear-btn" onclick="clearPanel('biolum')" style="display:none;" title="Clear biolum image">✕ Clear</button>
          <div id="tl-frame-controls" style="display:none;position:absolute;left:10px;right:10px;bottom:34px;align-items:center;gap:8px;background:rgba(6,9,13,0.88);border:1px solid var(--border2);border-radius:3px;padding:5px 8px;font-family:var(--mono);font-size:10px;color:var(--text);">
            <button class="btn btn-ghost btn-sm" style="width:auto;padding:2px 7px;" onclick="stepBiolumStackFrame(-1)">‹</button>
            <input id="tl-frame-slider" type="range" min="0" max="0" value="0" step="1" style="flex:1;" oninput="loadBiolumStackFrame(parseInt(this.value))">
            <button class="btn btn-ghost btn-sm" style="width:auto;padding:2px 7px;" onclick="stepBiolumStackFrame(1)">›</button>
            <span id="tl-frame-label" style="min-width:76px;text-align:right;color:var(--accent);">Frame 1/1</span>
          </div>
          <div class="canvas-sample" id="biolum-sample-label" style="display:none;"></div>
        </div>
      </div>

      <!-- bottom panel -->
      <div class="bottom-panel">

        <!-- compact ROI toolbar -->
        <div class="roi-toolbar">
          <span class="roi-toolbar-label">ROI</span>
          <div class="roi-mode-row">
            <div class="mode-btn active" id="mode-draw" onclick="setMode('draw')">✚ Draw</div>
            <div class="mode-btn" id="mode-move" onclick="setMode('move')">✥ Move</div>
            <div class="mode-btn" id="mode-resize" onclick="setMode('resize')">⤢ Resize</div>
            <div class="mode-btn" id="mode-pan" onclick="setMode('pan')">✋ Pan</div>
          </div>
          <select id="roi-shape" onchange="roiShape=this.value"
                  style="font-family:var(--mono);font-size:10px;background:var(--bg);color:var(--text);border:1px solid var(--border2);border-radius:3px;padding:3px 6px;cursor:pointer;">
            <option value="rect">■ Rect</option>
            <option value="circle">● Circle</option>
          </select>
          <div class="mode-btn" id="type-bckg" onclick="toggleBckgType()" title="Toggle background ROI type">Bckg</div>
          <label class="checkbox-row" style="flex-shrink:0;">
            <input type="checkbox" id="fix-size" onchange="toggleFixSize()">
            Lock size
          </label>
          <label class="checkbox-row" style="flex-shrink:0;">
            <input type="checkbox" id="rename-mode" onchange="toggleRenameMode()">
            Rename ROIs
          </label>
          <button class="btn btn-ghost btn-sm" onclick="resetView()">⊡ Fit</button>
          <button class="btn btn-red btn-sm" onclick="deleteSelected()">✕ Del ROI</button>
          <button class="btn btn-ghost btn-sm" onclick="clearAllRois()">Clear all</button>
        </div>

        <!-- measurements -->
        <div class="results-area" id="tab-meas">
          <div class="results-header">
            <span>Measurements</span>
            <div style="display:flex;gap:6px;" id="download-links"></div>
          </div>
          <div id="results-table-wrap">
            <div style="font-family:var(--mono);font-size:11px;color:var(--muted);">
              Draw ROIs then click ▶ Measure ROIs
            </div>
          </div>
        </div>

      </div>
    </div>
  </div>

  <!-- Analysis Summary — full-size pane -->
  <div class="top-tab-pane" id="top-tab-summary">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-shrink:0;">
      <span style="font-family:var(--mono);font-size:11px;color:var(--muted);letter-spacing:0.15em;text-transform:uppercase;">Mean IntDen per sample</span>
      <div style="display:flex;align-items:center;gap:8px;">
        <div id="ratio-controls" style="display:flex;align-items:center;gap:5px;font-family:var(--mono);font-size:10px;color:var(--muted);">
          <label style="display:flex;align-items:center;gap:5px;color:var(--text);margin-right:8px;">
            <input id="summary-use-bckg" type="checkbox" onchange="renderSummary()">
            <span>Bckg-subtracted</span>
          </label>
          <label style="display:flex;align-items:center;gap:5px;color:var(--text);margin-right:8px;">
            <input id="summary-show-ratio" type="checkbox" onchange="renderSummary()">
            <span>Show ratio</span>
          </label>
          <span>Ratio</span>
          <select id="ratio-num" onchange="renderSummary()" style="font-family:var(--mono);font-size:10px;background:var(--bg);color:var(--text);border:1px solid var(--border2);border-radius:3px;padding:3px 5px;">
            <option value="G" selected>Green</option>
            <option value="R">Red</option>
            <option value="B">Blue</option>
          </select>
          <span>/</span>
          <select id="ratio-den" onchange="renderSummary()" style="font-family:var(--mono);font-size:10px;background:var(--bg);color:var(--text);border:1px solid var(--border2);border-radius:3px;padding:3px 5px;">
            <option value="R" selected>Red</option>
            <option value="G">Green</option>
            <option value="B">Blue</option>
          </select>
        </div>
        <div id="summary-order-controls" style="display:none;align-items:center;gap:4px;font-family:var(--mono);font-size:10px;color:var(--muted);">
          <span>Order</span>
          <select id="summary-order-sample" style="font-family:var(--mono);font-size:10px;background:var(--bg);color:var(--text);border:1px solid var(--border2);border-radius:3px;padding:3px 5px;max-width:150px;"></select>
          <button class="btn btn-ghost btn-sm" onclick="moveSummarySample(-1)" title="Move selected sample left">←</button>
          <button class="btn btn-ghost btn-sm" onclick="moveSummarySample(1)" title="Move selected sample right">→</button>
        </div>
        <button class="btn btn-sm" onclick="exportSummaryPDF()" style="margin-right:6px;">Export PDF</button>
        <button class="btn btn-red btn-sm" onclick="clearAnalysisData()">Clear data</button>
      </div>
    </div>
    <div id="summary-empty" style="font-family:var(--mono);font-size:13px;color:var(--muted);text-align:center;padding:40px;">
      Measure samples in the Measurement tab to build the summary
    </div>
    <div style="display:flex;gap:16px;flex:1;min-height:220px;" id="summary-plots">
      <canvas id="plot-b" style="flex:1;min-width:0;"></canvas>
      <canvas id="plot-g" style="flex:1;min-width:0;"></canvas>
      <canvas id="plot-r" style="flex:1;min-width:0;"></canvas>
    </div>
    <div style="display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:16px;flex:0.8;min-height:220px;" id="summary-ratio-plot">
      <canvas id="plot-ratio" style="grid-column:2;min-width:0;"></canvas>
    </div>
  </div>

  </div><!-- end .tabs-container -->
</div>

<!-- Load ROIs dialog -->
<div id="roi-dialog" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.7);z-index:1000;align-items:center;justify-content:center;">
  <div style="background:var(--panel);border:1px solid var(--border2);border-radius:4px;padding:20px;width:500px;max-height:60vh;display:flex;flex-direction:column;gap:12px;">
    <div style="font-family:var(--mono);font-size:12px;letter-spacing:0.1em;color:var(--accent);text-transform:uppercase;">Load Saved ROIs</div>
    <div style="display:flex;gap:6px;align-items:center;">
      <input id="roi-search-folder" type="text" placeholder="ROI search folder"
             style="flex:1;background:var(--bg);border:1px solid var(--border2);border-radius:3px;color:var(--text);font-family:var(--mono);font-size:11px;padding:6px 8px;">
      <button class="btn btn-ghost btn-sm" style="width:auto;" onclick="pickRoiSearchFolder()">Browse</button>
      <button class="btn btn-sm" style="width:auto;" onclick="refreshRoiFileList()">Search</button>
    </div>
    <div id="roi-file-list" style="overflow-y:auto;flex:1;display:flex;flex-direction:column;gap:4px;"></div>
    <button class="btn btn-ghost" onclick="document.getElementById('roi-dialog').style.display='none'">Cancel</button>
  </div>
</div>

<script>
  // fix Windows backslashes for URL passing
  var _BS = String.fromCharCode(92);
  function fixPath(p) { return p ? p.split(_BS).join('/') : ''; }
  function escHtml(v) {
    return String(v ?? '').replace(/[&<>"']/g, ch => ({
      '&':'&amp;', '<':'&lt;', '>':'&gt;', '"':'&quot;', "'":'&#39;'
    }[ch]));
  }

  // ── state ──
  let dayPair = null;
  let biolumPair = null;
  let currentFolder = null;
  let jpegSize = {w: 1, h: 1};
  let rois = [];
  let selectedRoi = -1;
  let mode = 'draw';
  let roiShape = 'rect';
  let roiType = 'roi';
  let fixedSize = null;
  let isDrawing = false;
  let drawStart = {x:0, y:0};
  let dragStart = {x:0, y:0};
  let dragRoiStart = null;
  let measurements = [];
  let analysisData = [];
  let analysisDataUid = 0;
  let summarySampleOrder = [];
  let renameMode = false;
  let lastSaveDir = null;
  let sessionMeasurements = [];
  let roiAutosaveTimer = null;
  let unsavedResults = false;
  let savePromptShown = false;
  const SUMMARY_CHANNELS = [
    {id:'plot-b', channel:'B', rawKey:'mean_B', nbKey:'mean_B_nb', color:'#88bbff', colorExport:'#1030c0', rawLabel:'Mean B', nbLabel:'Mean B-Bckg'},
    {id:'plot-g', channel:'G', rawKey:'mean_G', nbKey:'mean_G_nb', color:'#55e888', colorExport:'#0e7a2e', rawLabel:'Mean G', nbLabel:'Mean G-Bckg'},
    {id:'plot-r', channel:'R', rawKey:'mean_R', nbKey:'mean_R_nb', color:'#ff9999', colorExport:'#b01010', rawLabel:'Mean R', nbLabel:'Mean R-Bckg'},
  ];
  const RATIO_CHANNELS = {
    B: {rawKey:'mean_B', nbKey:'mean_B_nb', label:'Blue', short:'B'},
    G: {rawKey:'mean_G', nbKey:'mean_G_nb', label:'Green', short:'G'},
    R: {rawKey:'mean_R', nbKey:'mean_R_nb', label:'Red', short:'R'},
  };
  let view = {scale: 1, dx: 0, dy: 0};
  let panStart = null;
  let dragRoiRef = null;

  window.addEventListener('beforeunload', e => {
    if (!unsavedResults) return;
    e.preventDefault();
    e.returnValue = '';
  });

  // canvas refs
  const dayCanvas = document.getElementById('day-canvas');
  const dayCtx = dayCanvas.getContext('2d');
  const biolumCanvas = document.getElementById('biolum-canvas');
  const biolumCtx = biolumCanvas.getContext('2d');
  let dayImg = null;
  let biolumImg = null;
  let biolumFrameIndex = 0;

  // ROI colours
  const ROI_COLORS = ['#4af0c4','#f0c44a','#f04a6a','#a04af0','#4a80f0','#f0804a'];

  // ── status ──
  function setStatus(msg, type='') {
    const el = document.getElementById('status');
    el.textContent = msg;
    el.className = 'status-bar' + (type ? ' status-' + type : '');
  }

  async function pickFolder() {
    setStatus('Opening folder picker...', '');
    try {
      const r = await fetch('/pick_folder');
      const data = await r.json();
      if (data.path) {
        document.getElementById('folder-input').value = data.path;
        await loadFolder(data.path);
      } else {
        setStatus('No folder selected.', '');
      }
    } catch(e) {
      setStatus('Picker error: ' + e.message, 'err');
    }
  }

  // ── folder loading ──
  async function loadFolder(folder) {
    folder = (folder || document.getElementById('folder-input').value).trim();
    if (!folder) return;
    setStatus('Loading...', '');
    try {
      const r = await fetch('/browse', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({folder})
      });
      const data = await r.json();
      currentFolder = folder;
      lastSaveDir = data.analysis_dir || null;
      sessionMeasurements = [];
      analysisData = [];
      analysisDataUid = 0;
      summarySampleOrder = [];
      document.getElementById('header-path').textContent = folder;
      const name = fixPath(folder).split('/').filter(Boolean).pop() || folder;
      document.getElementById('experiment-label').textContent = name;
      document.getElementById('folder-input').value = folder;
      dayPair = null; biolumPair = null; dayImg = null; biolumImg = null;
      document.getElementById('tl-frame-controls').style.display = 'none';
      setSampleLabel('day', null); setSampleLabel('biolum', null);
      renderPairList(data.pairs);
      const dirMsg = lastSaveDir ? (' → ' + fixPath(lastSaveDir).split('/').filter(Boolean).pop()) : '';
      setStatus('Loaded ' + data.pairs.length + ' image items' + dirMsg, 'ok');
    } catch(e) {
      setStatus('Load error: ' + e.message, 'err');
    }
  }

  function renderPairList(pairs) {
    const list = document.getElementById('pair-list');
    document.getElementById('pair-count').textContent = pairs.length + ' items';
    if (!pairs.length) {
      list.innerHTML = '<div style="padding:20px;text-align:center;font-family:var(--mono);font-size:11px;color:var(--muted);">No image pairs found</div>';
      return;
    }
    list.innerHTML = pairs.map((p, i) => `
      <div class="pair-item" onclick="loadPair(${i})" id="pair-${i}" data-pair="${encodeURIComponent(JSON.stringify(p))}">
        <div style="display:flex;align-items:center;justify-content:space-between;gap:4px;">
          <span style="overflow:hidden;text-overflow:ellipsis;">${escHtml(p.stem)}</span>
          <span class="badge badge-${p.type}" style="flex-shrink:0;">${p.is_stack ? 'TL' : p.type.toUpperCase()}</span>
        </div>
        ${p.is_stack ? `<div style="font-size:10px;color:var(--muted);margin-top:4px;">${p.frames.length} frames · ${p.exposure}</div>` : ''}
      </div>
    `).join('');
  }

  function setSampleLabel(side, stem) {
    const el = document.getElementById(side + '-sample-label');
    if (stem) {
      el.textContent = sampleName(stem);
      el.style.display = '';
    } else {
      el.textContent = '';
      el.style.display = 'none';
    }
  }

  function clearCurrentMeasurementView() {
    measurements = [];
    selectedRoi = -1;
    document.getElementById('results-table-wrap').innerHTML =
      '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click â–¶ Measure ROIs</div>';
    document.getElementById('download-links').innerHTML = '';
    unsavedResults = false;
    savePromptShown = false;
  }

  async function loadPair(idx) {
    const el = document.getElementById('pair-' + idx);
    const pair = JSON.parse(decodeURIComponent(el.dataset.pair));
    const isDay = pair.type !== 'biolum';
    const newSample = pair.sample || sampleName(pair.stem);

    // Enforce same-sample pairing
    if (isDay && biolumPair) {
      const existing = biolumPair.sample || sampleName(biolumPair.stem);
      if (newSample.toLowerCase() !== existing.toLowerCase()) {
        setStatus('Sample mismatch: "' + newSample + '" vs loaded biolum "' + existing + '". Clear the biolum image first.', 'err');
        return;
      }
    } else if (!isDay && dayPair) {
      const existing = dayPair.sample || sampleName(dayPair.stem);
      if (newSample.toLowerCase() !== existing.toLowerCase()) {
        setStatus('Sample mismatch: "' + newSample + '" vs loaded day "' + existing + '". Clear the day image first.', 'err');
        return;
      }
    }

    // Update active highlights
    document.querySelectorAll('.pair-item').forEach(e => e.classList.remove('active', 'active-day', 'active-biolum'));
    el.classList.add('active', isDay ? 'active-day' : 'active-biolum');

    if (isDay) {
      dayPair = pair;
      rois = [];
      measurements = [];
      selectedRoi = -1;
      document.getElementById('results-table-wrap').innerHTML =
        '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click ▶ Measure ROIs</div>';
      document.getElementById('download-links').innerHTML = '';
    } else {
      biolumPair = pair;
      clearCurrentMeasurementView();
    }

    if (pair.jpg) {
      if (isDay) {
        await loadImageToCanvas(pair.jpg, dayCanvas, dayCtx, 'day-placeholder');
        const sr = await fetch('/image_size?path=' + encodeURIComponent(pair.jpg));
        const sd = await sr.json();
        jpegSize = {w: sd.width, h: sd.height};
        dayImg = new Image();
        dayImg.src = '/image?path=' + encodeURIComponent(pair.jpg);
        await new Promise(r => { dayImg.onload = r; });
      } else {
        if (pair.is_stack) {
          await loadBiolumStackFrame(0, pair);
        } else {
          document.getElementById('tl-frame-controls').style.display = 'none';
          await loadImageToCanvas(pair.jpg, biolumCanvas, biolumCtx, 'biolum-placeholder');
          if (!dayPair) {
            const sr = await fetch('/image_size?path=' + encodeURIComponent(pair.jpg));
            const sd = await sr.json();
            jpegSize = {w: sd.width, h: sd.height};
          }
          biolumImg = new Image();
          biolumImg.src = '/image?path=' + encodeURIComponent(pair.jpg);
          await new Promise(r => { biolumImg.onload = r; });
        }
      }
    }

    setSampleLabel(isDay ? 'day' : 'biolum', pair.sample || pair.stem);
    document.getElementById((isDay ? 'day' : 'biolum') + '-clear-btn').style.display = '';
    drawAll();

    if (isDay && !biolumPair) {
      setStatus('"' + newSample + '" day loaded — select the matching biolum image.', 'wrn');
    } else if (!isDay && !dayPair) {
      setStatus('"' + newSample + '" biolum loaded — select the matching day image.', 'wrn');
    } else if (!isDay && pair.is_stack) {
      setStatus('Time-lapse stack loaded: ' + newSample + ' (' + pair.frames.length + ' frames)', 'ok');
    } else {
      setStatus('Set complete: ' + newSample, 'ok');
    }
  }

  async function loadBiolumStackFrame(idx, stack=null) {
    const activeStack = stack || biolumPair;
    if (!activeStack || !activeStack.is_stack || !activeStack.frames || !activeStack.frames.length) return;
    idx = Math.max(0, Math.min(activeStack.frames.length - 1, idx || 0));
    biolumFrameIndex = idx;
    const frame = activeStack.frames[idx];
    activeStack.jpg = frame.jpg;
    activeStack.nef = frame.nef;

    const controls = document.getElementById('tl-frame-controls');
    const slider = document.getElementById('tl-frame-slider');
    const label = document.getElementById('tl-frame-label');
    controls.style.display = 'flex';
    slider.max = String(activeStack.frames.length - 1);
    slider.value = String(idx);
    label.textContent = `Frame ${idx + 1}/${activeStack.frames.length}`;

    if (!frame.jpg) {
      setStatus('Selected stack frame has no JPEG preview: ' + frame.stem, 'err');
      return;
    }

    await loadImageToCanvas(frame.jpg, biolumCanvas, biolumCtx, 'biolum-placeholder');
    if (!dayPair) {
      const sr = await fetch('/image_size?path=' + encodeURIComponent(frame.jpg));
      const sd = await sr.json();
      jpegSize = {w: sd.width, h: sd.height};
    }
    biolumImg = new Image();
    biolumImg.src = '/image?path=' + encodeURIComponent(frame.jpg);
    await new Promise(r => { biolumImg.onload = r; });
    drawAll();
  }

  function stepBiolumStackFrame(delta) {
    if (!biolumPair || !biolumPair.is_stack) return;
    loadBiolumStackFrame(biolumFrameIndex + delta);
  }

  async function loadImageToCanvas(path, canvas, ctx, placeholderId) {
    const img = new Image();
    img.src = '/image?path=' + encodeURIComponent(path);
    await new Promise(r => { img.onload = r; img.onerror = r; });
    const wrap = canvas.parentElement;
    const maxW = wrap.clientWidth;
    const maxH = wrap.clientHeight;
    const scale = Math.min(maxW / img.naturalWidth, maxH / img.naturalHeight, 1);
    canvas.width  = Math.round(img.naturalWidth  * scale);
    canvas.height = Math.round(img.naturalHeight * scale);
    ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
    canvas.style.display = 'block';
    document.getElementById(placeholderId).style.display = 'none';
  }

  // ── ROI drawing ──
  function setMode(m) {
    mode = m;
    ['draw','move','resize','pan'].forEach(mm => {
      document.getElementById('mode-' + mm).classList.toggle('active', mm === m);
    });
    const cur = m === 'draw' ? 'crosshair' : m === 'move' ? 'move' : m === 'pan' ? 'grab' : 'nw-resize';
    dayCanvas.style.cursor = cur;
    biolumCanvas.style.cursor = cur;
  }

  function resetView() {
    view = {scale: 1, dx: 0, dy: 0};
    drawAll();
  }

  function applyViewTransform(ctx, canvas) {
    const cx = canvas.width / 2, cy = canvas.height / 2;
    ctx.translate(cx + view.dx, cy + view.dy);
    ctx.scale(view.scale, view.scale);
    ctx.translate(-cx, -cy);
  }

  function dayToBiolum(pos) {
    if (!biolumCanvas.width || !dayCanvas.width) return pos;
    return {x: pos.x * biolumCanvas.width / dayCanvas.width,
            y: pos.y * biolumCanvas.height / dayCanvas.height};
  }

  function biolumToDay(pos) {
    if (!dayCanvas.width || !biolumCanvas.width) return pos;
    return {x: pos.x * dayCanvas.width / biolumCanvas.width,
            y: pos.y * dayCanvas.height / biolumCanvas.height};
  }

  function drawPreview(startDay, endDay) {
    [{ctx: dayCtx, canvas: dayCanvas, s: startDay, e: endDay},
     {ctx: biolumCtx, canvas: biolumCanvas,
      s: dayToBiolum(startDay), e: dayToBiolum(endDay)}
    ].forEach(({ctx, canvas, s, e}) => {
      if (!canvas.width) return;
      ctx.save();
      applyViewTransform(ctx, canvas);
      ctx.strokeStyle = 'rgba(74,240,196,0.8)';
      ctx.lineWidth = 2;
      ctx.setLineDash([4,4]);
      const px = Math.min(s.x,e.x), py = Math.min(s.y,e.y);
      const pw = Math.abs(e.x-s.x), ph = Math.abs(e.y-s.y);
      if (roiShape === 'circle') {
        ctx.beginPath();
        ctx.ellipse(px+pw/2, py+ph/2, pw/2, ph/2, 0, 0, Math.PI*2);
        ctx.stroke();
      } else {
        ctx.strokeRect(s.x, s.y, e.x-s.x, e.y-s.y);
      }
      ctx.setLineDash([]);
      ctx.restore();
    });
  }

  function constrainSquare(start, end) {
    const dx = end.x - start.x, dy = end.y - start.y;
    const size = Math.min(Math.abs(dx), Math.abs(dy));
    return {x: start.x + Math.sign(dx) * size, y: start.y + Math.sign(dy) * size};
  }

  function ensureRoiNumbers() {
    let roiMax = 0, bckgMax = 0;
    rois.forEach(r => {
      const n = Number(r.roi_number);
      if (!n || n < 1) return;
      if (r.type === 'bckg') bckgMax = Math.max(bckgMax, n);
      else roiMax = Math.max(roiMax, n);
    });
    rois.forEach(r => {
      if (Number(r.roi_number) > 0) return;
      if (r.type === 'bckg') r.roi_number = ++bckgMax;
      else r.roi_number = ++roiMax;
    });
  }

  function nextRoiNumber(type) {
    ensureRoiNumbers();
    const sameType = rois.filter(r => (type === 'bckg') ? r.type === 'bckg' : r.type !== 'bckg');
    return sameType.reduce((max, r) => Math.max(max, Number(r.roi_number) || 0), 0) + 1;
  }

  function commitRoi(endPosDay) {
    let x = Math.min(drawStart.x, endPosDay.x);
    let y = Math.min(drawStart.y, endPosDay.y);
    let w = Math.abs(endPosDay.x - drawStart.x);
    let h = Math.abs(endPosDay.y - drawStart.y);
    if (fixedSize) { w = fixedSize.w; h = fixedSize.h; }
    if (w > 5 && h > 5) {
      const roi = {x, y, w, h, shape: roiShape, type: roiType, roi_number: nextRoiNumber(roiType)};
      rois.push(roi);
      selectedRoi = rois.indexOf(roi);
      renderRoiSetupTable();
      scheduleRoiAutosave();
    }
    isDrawing = false;
    drawAll();
  }

  function sampleName(stem) {
    let m = stem.match(/^(.+?)_tl\\d+(?:_|$)/i);
    if (m) return m[1].trim();
    m = stem.match(/^(.+?)_biolum/i);
    if (m) return m[1].trim();
    m = stem.match(/^(.+?)_day/i);
    if (m) return m[1].trim();
    m = stem.match(/^(.+?)_light/i);
    if (m) return m[1].trim();
    return stem.trim();
  }

  function toggleFixSize() {
    const checked = document.getElementById('fix-size').checked;
    if (checked && rois.length > 0) {
      const last = rois[rois.length - 1];
      fixedSize = {w: last.w, h: last.h};
    } else {
      fixedSize = null;
    }
  }

  function canvasPos(canvas, e) {
    const rect = canvas.getBoundingClientRect();
    const sx = (e.clientX - rect.left) * (canvas.width / rect.width);
    const sy = (e.clientY - rect.top)  * (canvas.height / rect.height);
    const cx = canvas.width / 2, cy = canvas.height / 2;
    return {
      x: (sx - cx - view.dx) / view.scale + cx,
      y: (sy - cy - view.dy) / view.scale + cy,
      sx, sy
    };
  }

  function onWheel(canvas, e) {
    e.preventDefault();
    const rect = canvas.getBoundingClientRect();
    const sx = (e.clientX - rect.left) * (canvas.width / rect.width);
    const sy = (e.clientY - rect.top)  * (canvas.height / rect.height);
    const cx = canvas.width / 2, cy = canvas.height / 2;
    const factor = e.deltaY < 0 ? 1.12 : 1 / 1.12;
    const ns = Math.max(0.1, Math.min(20, view.scale * factor));
    const ratio = ns / view.scale;
    view.dx = (1 - ratio) * (sx - cx) + ratio * view.dx;
    view.dy = (1 - ratio) * (sy - cy) + ratio * view.dy;
    view.scale = ns;
    drawAll();
  }
  dayCanvas.addEventListener('wheel', e => onWheel(dayCanvas, e), {passive: false});
  biolumCanvas.addEventListener('wheel', e => onWheel(biolumCanvas, e), {passive: false});

  function hitTest(x, y) {
    // returns index of ROI hit, -1 if none
    // check in reverse so top-most ROI is selected first
    for (let i = rois.length - 1; i >= 0; i--) {
      const r = rois[i];
      if (x >= r.x && x <= r.x + r.w && y >= r.y && y <= r.y + r.h) return i;
    }
    return -1;
  }

  function resizeHandle(roi, x, y) {
    // returns true if near bottom-right corner
    const dx = x - (roi.x + roi.w);
    const dy = y - (roi.y + roi.h);
    return Math.abs(dx) < 10 && Math.abs(dy) < 10;
  }

  function roiMouseDown(pos, rawPos) {
    if (mode === 'pan') {
      panStart = {sx: rawPos.sx, sy: rawPos.sy, dx: view.dx, dy: view.dy};
      dayCanvas.style.cursor = 'grabbing';
      biolumCanvas.style.cursor = 'grabbing';
      return;
    }
    if (!dayPair && !biolumPair) return;
    if (mode === 'draw') {
      const preHit = hitTest(pos.x, pos.y);
      if (preHit >= 0) { selectedRoi = preHit; drawAll(); return; }
      isDrawing = true; drawStart = pos; selectedRoi = -1;
    } else if (mode === 'move') {
      const hit = hitTest(pos.x, pos.y);
      if (hit >= 0) {
        selectedRoi = hit; dragStart = pos;
        dragRoiRef = rois[hit]; dragRoiStart = {...rois[hit]};
      }
    } else if (mode === 'resize') {
      const hit = hitTest(pos.x, pos.y);
      if (hit >= 0) {
        selectedRoi = hit; dragStart = pos;
        dragRoiRef = rois[hit]; dragRoiStart = {...rois[hit]};
      }
    }
    drawAll();
  }

  function roiMouseMove(pos, rawPos, buttons, shiftKey) {
    if (mode === 'pan' && panStart && buttons === 1) {
      view.dx = panStart.dx + rawPos.sx - panStart.sx;
      view.dy = panStart.dy + rawPos.sy - panStart.sy;
      drawAll(); return;
    }
    if (!isDrawing && mode !== 'move' && mode !== 'resize') return;
    if (isDrawing && mode === 'draw') {
      const ep = shiftKey ? constrainSquare(drawStart, pos) : pos;
      drawAll(); drawPreview(drawStart, ep);
    } else if (mode === 'move' && dragRoiRef && buttons === 1) {
      dragRoiRef.x = dragRoiStart.x + pos.x - dragStart.x;
      dragRoiRef.y = dragRoiStart.y + pos.y - dragStart.y;
      drawAll();
    } else if (mode === 'resize' && dragRoiRef && buttons === 1) {
      dragRoiRef.w = Math.max(10, dragRoiStart.w + pos.x - dragStart.x);
      dragRoiRef.h = Math.max(10, dragRoiStart.h + pos.y - dragStart.y);
      drawAll();
    }
  }

  function roiMouseUp(pos, shiftKey) {
    if (mode === 'pan') {
      panStart = null;
      const cur = 'grab';
      dayCanvas.style.cursor = cur; biolumCanvas.style.cursor = cur; return;
    }
    if (isDrawing && mode === 'draw') {
      commitRoi(shiftKey ? constrainSquare(drawStart, pos) : pos);
    } else if ((mode === 'move' || mode === 'resize') && dragRoiRef) {
      selectedRoi = rois.indexOf(dragRoiRef);
      dragRoiRef = null;
      renderRoiSetupTable();
      scheduleRoiAutosave();
      drawAll();
    }
  }

  // day canvas events
  dayCanvas.addEventListener('mousedown', e => {
    const p = canvasPos(dayCanvas, e);
    roiMouseDown(p, p);
  });
  dayCanvas.addEventListener('mousemove', e => {
    const p = canvasPos(dayCanvas, e);
    roiMouseMove(p, p, e.buttons, e.shiftKey);
  });
  dayCanvas.addEventListener('mouseup', e => {
    roiMouseUp(canvasPos(dayCanvas, e), e.shiftKey);
  });
  dayCanvas.addEventListener('dblclick', e => {
    const hit = hitTest(canvasPos(dayCanvas, e).x, canvasPos(dayCanvas, e).y);
    if (hit >= 0) { selectedRoi = hit; drawAll(); }
  });

  // biolum canvas events — coords converted to day space
  biolumCanvas.addEventListener('mousedown', e => {
    const raw = canvasPos(biolumCanvas, e);
    roiMouseDown(biolumToDay(raw), raw);
  });
  biolumCanvas.addEventListener('mousemove', e => {
    const raw = canvasPos(biolumCanvas, e);
    roiMouseMove(biolumToDay(raw), raw, e.buttons, e.shiftKey);
  });
  biolumCanvas.addEventListener('mouseup', e => {
    roiMouseUp(biolumToDay(canvasPos(biolumCanvas, e)), e.shiftKey);
  });
  biolumCanvas.addEventListener('dblclick', e => {
    const pos = biolumToDay(canvasPos(biolumCanvas, e));
    const hit = hitTest(pos.x, pos.y);
    if (hit >= 0) { selectedRoi = hit; drawAll(); }
  });

  function sortRois() {
    // Keep ROI order stable. Numbering follows creation/load order so names do not drift.
  }

  function drawOneCanvas(ctx, canvas, img, isBiolum) {
    ctx.setTransform(1, 0, 0, 1, 0, 0);
    ctx.clearRect(0, 0, canvas.width, canvas.height);
    if (!img) return;
    ctx.save();
    applyViewTransform(ctx, canvas);
    ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
    if (isBiolum && dayCanvas.width > 0) {
      drawRoisScaled(ctx, canvas.width / dayCanvas.width, canvas.height / dayCanvas.height);
    } else {
      drawRois(ctx, canvas);
    }
    ctx.restore();
  }

  function drawAll() {
    drawOneCanvas(dayCtx, dayCanvas, dayImg, false);
    drawOneCanvas(biolumCtx, biolumCanvas, biolumImg, true);
  }

  function getRoiLabel(roi, idx) {
    ensureRoiNumbers();
    if (roi.type === 'bckg') {
      return 'B' + roi.roi_number;
    }
    return String(roi.roi_number);
  }

  function getRoiTableLabel(roi, idx) {
    ensureRoiNumbers();
    if (roi.label) return roi.label;
    if (roi.type === 'bckg') {
      return 'Bckg ' + roi.roi_number;
    }
    return String(roi.roi_number);
  }

  function roiAreaPx(roi) {
    if (!roi) return 0;
    if ((roi.shape || 'rect') === 'circle') return Math.round(Math.PI * (roi.w / 2) * (roi.h / 2));
    return Math.round(roi.w * roi.h);
  }

  function roiReferenceCanvas() {
    if (dayPair && dayCanvas.width > 0 && dayCanvas.height > 0) return dayCanvas;
    if (biolumPair && biolumCanvas.width > 0 && biolumCanvas.height > 0) return biolumCanvas;
    return dayCanvas.width > 0 ? dayCanvas : biolumCanvas;
  }

  function scaledRoisToJpeg() {
    const refCanvas = roiReferenceCanvas();
    if (!refCanvas || refCanvas.width < 1 || refCanvas.height < 1) return [];
    const scaleX = jpegSize.w / refCanvas.width;
    const scaleY = jpegSize.h / refCanvas.height;
    return rois.map(r => ({
      ...r,
      x: Math.round(r.x * scaleX),
      y: Math.round(r.y * scaleY),
      w: Math.round(r.w * scaleX),
      h: Math.round(r.h * scaleY),
      shape: r.shape || 'rect',
      type: r.type || 'roi',
      label: r.label || '',
      roi_number: r.roi_number,
    }));
  }

  function drawSingleRoi(ctx, x, y, w, h, shape, color, isSelected, label) {
    ctx.strokeStyle = color;
    ctx.lineWidth = isSelected ? 1.8 : 1.1;
    ctx.fillStyle = color + '0a';
    if (shape === 'circle') {
      ctx.beginPath();
      ctx.ellipse(x+w/2, y+h/2, w/2, h/2, 0, 0, Math.PI*2);
      ctx.stroke(); ctx.fill();
    } else {
      ctx.strokeRect(x, y, w, h);
      ctx.fillRect(x, y, w, h);
    }
    ctx.fillStyle = color;
    ctx.font = '500 11px IBM Plex Mono';
    ctx.textAlign = 'center';
    ctx.textBaseline = 'middle';
    ctx.lineWidth = 1.4;
    ctx.strokeStyle = 'rgba(0,0,0,0.65)';
    ctx.strokeText(label, x + w / 2, y + h / 2);
    ctx.fillStyle = color + 'dd';
    ctx.fillText(label, x + w / 2, y + h / 2);
    ctx.textAlign = 'start';
    ctx.textBaseline = 'alphabetic';
  }

  function drawRois(ctx, canvas) {
    rois.forEach((roi, i) => {
      drawSingleRoi(ctx, roi.x, roi.y, roi.w, roi.h,
        roi.shape||'rect', ROI_COLORS[i % ROI_COLORS.length], roi === dragRoiRef || i === selectedRoi, getRoiLabel(roi,i));
    });
  }

  function drawRoisScaled(ctx, sx, sy) {
    rois.forEach((roi, i) => {
      drawSingleRoi(ctx, roi.x*sx, roi.y*sy, roi.w*sx, roi.h*sy,
        roi.shape||'rect', ROI_COLORS[i % ROI_COLORS.length], roi === dragRoiRef || i === selectedRoi, getRoiLabel(roi,i));
    });
  }

  function deleteSelected() {
    if (selectedRoi >= 0 && selectedRoi < rois.length) {
      rois.splice(selectedRoi, 1);
      selectedRoi = -1;
      dragRoiRef = null;
      measurements = [];
      renderRoiSetupTable();
      scheduleRoiAutosave();
      drawAll();
    }
  }

  function clearAllRois() {
    rois = [];
    selectedRoi = -1;
    measurements = [];
    sessionMeasurements = [];
    analysisData = [];
    analysisDataUid = 0;
    summarySampleOrder = [];
    unsavedResults = false;
    savePromptShown = false;
    document.getElementById('results-table-wrap').innerHTML =
      '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click ▶ Measure ROIs</div>';
    document.getElementById('download-links').innerHTML = '';
    renderSummary();
    scheduleRoiAutosave(true);
    setStatus('ROIs and measurements cleared', '');
    drawAll();
  }

  // ── measurement ──
  async function measureAll() {
    if (!dayPair && !biolumPair) { setStatus('Load images first', 'err'); return; }
    if (!rois.length) { setStatus('Draw at least one ROI first', 'wrn'); return; }
    if (!biolumPair || !biolumPair.nef) { setStatus('No biolum NEF loaded — select a biolum image with a NEF file', 'err'); return; }

    syncRoiLabelsFromInputs();
    ensureRoiNumbers();
    const isStackMeasure = biolumPair && biolumPair.is_stack;
    setStatus(isStackMeasure ? 'Reading all stack NEFs and measuring...' : 'Reading NEF and measuring... (this may take a few seconds)', '');
    document.getElementById('btn-measure').disabled = true;

    // convert canvas ROIs to JPEG pixel coords using day image when present,
    // otherwise the biolum preview itself for biolum-only workflows.
    const scaledRois = scaledRoisToJpeg();
    if (!scaledRois.length) { setStatus('Could not scale ROIs - reload the image and try again', 'err'); return; }

    const sample = biolumPair
      ? (biolumPair.sample || sampleName(biolumPair.stem))
      : (dayPair ? (dayPair.sample || sampleName(dayPair.stem)) : '');

    let rawMeasurements = [];
    try {
      if (isStackMeasure) {
        rawMeasurements = await measureStackFramesWithProgress(scaledRois);
      } else {
        rawMeasurements = await measureSingleImageRoisWithProgress(scaledRois);
      }
    } catch (err) {
      setStatus('Error: ' + err.message, 'err');
      return;
    } finally {
      document.getElementById('btn-measure').disabled = false;
    }

    measurements = buildMeasurements(rawMeasurements, scaledRois, sample);

    // accumulate session measurements (all ROIs including bckg)
    sessionMeasurements.push(...measurements);

    // accumulate for Analysis Summary (non-background ROIs only)
    measurements.filter(m => m.roi_type !== 'bckg').forEach(m => {
      const uid = analysisDataUid++;
      m._uid = uid;
      analysisData.push({
        uid,
        sample: m.sample_name,
        roiLabel: m.roi_label,
        mean_R: m.mean_R,
        mean_G: m.mean_G,
        mean_B: m.mean_B,
        mean_R_nb: m.mean_R_nb,
        mean_G_nb: m.mean_G_nb,
        mean_B_nb: m.mean_B_nb,
        has_bckg: m.mean_R_nb !== undefined && m.mean_R_nb !== null,
        frame_number: m.frame_number,
        elapsed_min: m.elapsed_min,
      });
    });

    renderTable(measurements);
    unsavedResults = true;
    savePromptShown = false;
    setStatus(isStackMeasure
      ? `Measured ${measurements.length} ROI rows across ${biolumPair.frames.length} stack frames`
      : `Measured ${measurements.length} ROIs from NEF (16-bit)`, 'ok');
    promptSaveResultsSoon();
  }

  function renderMeasureProgress(done, total, message, unit = 'ROIs') {
    const wrap = document.getElementById('results-table-wrap');
    const pct = total ? Math.round((done / total) * 100) : 0;
    wrap.innerHTML = `
      <div style="font-family:var(--mono);font-size:12px;color:var(--text);padding:14px 12px;">
        <div style="display:flex;justify-content:space-between;gap:12px;margin-bottom:8px;">
          <span>${message}</span>
          <span style="color:var(--accent);">${done}/${total} ${unit}</span>
        </div>
        <div style="height:12px;background:#0b1018;border:1px solid var(--border2);border-radius:3px;overflow:hidden;">
          <div style="width:${pct}%;height:100%;background:linear-gradient(90deg,var(--accent),var(--amber));transition:width 0.2s ease;"></div>
        </div>
      </div>`;
  }

  function frameTimestampSeconds(stem) {
    const m = String(stem || '').match(/_(\\d{6})$/);
    if (!m) return null;
    const hh = parseInt(m[1].slice(0, 2), 10);
    const mm = parseInt(m[1].slice(2, 4), 10);
    const ss = parseInt(m[1].slice(4, 6), 10);
    if (hh > 23 || mm > 59 || ss > 59) return null;
    return hh * 3600 + mm * 60 + ss;
  }

  async function measureStackFramesWithProgress(scaledRois) {
    const frames = (biolumPair.frames || [])
      .filter(f => f.nef)
      .slice()
      .sort((a, b) => (a.frame || 0) - (b.frame || 0));
    if (!frames.length) throw new Error('No NEF files found in stack');
    return measureJobWithProgress({
      mode: 'stack',
      frames,
      rois: scaledRois,
      jpeg_size: [jpegSize.w, jpegSize.h]
    }, 'Preparing stack measurement...');
  }

  async function measureSingleImageRoisWithProgress(scaledRois) {
    return measureJobWithProgress({
      mode: 'single',
      nef_path: biolumPair.nef,
      rois: scaledRois,
      jpeg_size: [jpegSize.w, jpegSize.h]
    }, 'Preparing single-image ROI measurement...');
  }

  async function measureJobWithProgress(payload, initialMessage) {
    const totalHint = payload.mode === 'stack'
      ? Math.max(1, (payload.frames || []).length * (payload.rois || []).length)
      : Math.max(1, (payload.rois || []).length);
    const progressUnit = payload.mode === 'stack' ? 'ROI steps' : 'ROIs';
    renderMeasureProgress(0, totalHint, initialMessage, progressUnit);
    await new Promise(requestAnimationFrame);

    const start = await fetch('/measure_job', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify(payload)
    });
    const started = await start.json();
    if (started.error) throw new Error(started.error);
    if (!started.job_id) throw new Error('Measurement job did not start');

    while (true) {
      await new Promise(resolve => setTimeout(resolve, 180));
      const r = await fetch('/measure_job/' + encodeURIComponent(started.job_id));
      const data = await r.json();
      if (data.error) throw new Error(data.error);
      renderMeasureProgress(data.done || 0, data.total || totalHint, data.message || 'Measuring...', progressUnit);
      if (data.status === 'error') throw new Error(data.error || 'Measurement failed');
      if (data.status === 'done') return data.measurements || [];
    }
  }

  function buildMeasurements(rawMeasurements, scaledRois, sample) {
    const groups = new Map();
    rawMeasurements.forEach(m => {
      const key = m.frame_index !== undefined ? m.frame_index : 0;
      if (!groups.has(key)) groups.set(key, []);
      groups.get(key).push(m);
    });

    const out = [];
    [...groups.keys()].sort((a,b) => a-b).forEach(key => {
      const rows = groups.get(key);
      const bckgMeas = rows.filter((m, i) => scaledRois[i] && scaledRois[i].type === 'bckg');
      let bckgMean = null;
      if (bckgMeas.length > 0) {
        const avg = k => bckgMeas.reduce((s, m) => s + m[k], 0) / bckgMeas.length;
        bckgMean = {
          mean_R: avg('mean_R'), mean_G: avg('mean_G'), mean_B: avg('mean_B'),
          intden_R: avg('intden_R'), intden_G: avg('intden_G'), intden_B: avg('intden_B'),
        };
      }

      let roiCount = 0, bckgCount = 0;
      rows.forEach((m, i) => {
        const roiObj = scaledRois[i] || {};
        const isB = roiObj.type === 'bckg';
        if (isB) bckgCount++; else roiCount++;
        const stableNumber = roiObj.roi_number || (isB ? bckgCount : roiCount);
        const defaultLabel = isB ? ('Bckg ' + stableNumber) : String(stableNumber);
        const roi_label = (roiObj.label || '').trim() || defaultLabel;
        const nb = bckgMean ? {
          mean_R_nb:   isB ? null : m.mean_R   - bckgMean.mean_R,
          mean_G_nb:   isB ? null : m.mean_G   - bckgMean.mean_G,
          mean_B_nb:   isB ? null : m.mean_B   - bckgMean.mean_B,
          intden_R_nb: isB ? null : m.intden_R - bckgMean.intden_R,
          intden_G_nb: isB ? null : m.intden_G - bckgMean.intden_G,
          intden_B_nb: isB ? null : m.intden_B - bckgMean.intden_B,
        } : {};
        out.push({...m, sample_name: sample, roi_type: roiObj.type||'roi', roi_label, roi_index: i, ...nb});
      });
    });
    return out;
  }

  function currentSampleName() {
    if (biolumPair) return biolumPair.sample || sampleName(biolumPair.stem);
    if (dayPair) return dayPair.sample || sampleName(dayPair.stem);
    return '';
  }

  function renderRoiSetupTable() {
    if (!rois.length) {
      measurements = [];
      document.getElementById('results-table-wrap').innerHTML =
        '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click ▶ Measure ROIs</div>';
      return;
    }
    ensureRoiNumbers();
    measurements = rois.map((roi, i) => ({
      sample_name: currentSampleName(),
      roi_type: roi.type || 'roi',
      roi_label: getRoiTableLabel(roi, i),
      roi_index: i,
      area_px: roiAreaPx(roi),
      mean_B: null,
      mean_G: null,
      mean_R: null,
    }));
    renderTable(measurements);
  }

  function renderTable(data) {
    if (!data.length) return;
    const wrap = document.getElementById('results-table-wrap');
    const esc = v => String(v ?? '').replace(/[&<>"']/g, ch => ({
      '&':'&amp;', '<':'&lt;', '>':'&gt;', '"':'&quot;', "'":'&#39;'
    }[ch]));
    const hasBckg = data.some(m => m.mean_R_nb !== undefined && m.mean_R_nb !== null);
    const hasFrame = data.some(m => m.frame_number !== undefined && m.frame_number !== null);
    const hasElapsed = data.some(m => m.elapsed_min !== undefined && m.elapsed_min !== null);
    const valueCols = [
      ...(hasFrame ? [{label:'Frame #', key:'frame_number'}] : []),
      ...(hasElapsed ? [{label:'Elapsed (min)', key:'elapsed_min'}] : []),
      {label:'Area (px)', key:'area_px'},
      {label:'Raw Mean B', key:'mean_B', cls:'ch-b'},
      {label:'Raw Mean G', key:'mean_G', cls:'ch-g'},
      {label:'Raw Mean R', key:'mean_R', cls:'ch-r'},
      ...(hasBckg ? [
        {label:'Mean B-Bckg', key:'mean_B_nb', cls:'ch-b-nb'},
        {label:'Mean G-Bckg', key:'mean_G_nb', cls:'ch-g-nb'},
        {label:'Mean R-Bckg', key:'mean_R_nb', cls:'ch-r-nb'},
      ] : []),
    ];
    const fmtVal = v => {
      if (v === null || v === undefined) return '—';
      if (typeof v === 'number') return v.toLocaleString(undefined, {maximumFractionDigits: 1});
      return v;
    };
    const roiHeader = renameMode ? 'ROI Name' : 'ROI';
    const sampleHeader = renameMode ? '' : '<th>Sample</th>';
    wrap.innerHTML = `
      <table>
        <thead><tr>${sampleHeader}<th>${roiHeader}</th>${valueCols.map(c => `<th>${c.label}</th>`).join('')}</tr></thead>
        <tbody>
          ${data.map((row, ri) => {
            const isB = row.roi_type === 'bckg';
            const rowCls = isB ? ' class="bckg-row"' : '';
            const sampleCell = renameMode ? '' : `<td>${esc(fmtVal(row.sample_name))}</td>`;
            const roiCell = (renameMode && !isB)
              ? `<td><input type="text" value="${esc(row.roi_label)}" data-roi-index="${row.roi_index}" style="font-family:var(--mono);font-size:11px;background:transparent;border:1px solid var(--border2);color:var(--text);border-radius:2px;padding:1px 4px;width:80px;" oninput="renameMeasurement(${ri}, this.value)"></td>`
              : `<td>${esc(fmtVal(row.roi_label))}</td>`;
            return `<tr${rowCls}>${sampleCell}${roiCell}${valueCols.map(c => {
              const cls = isB && (c.cls||'').includes('-nb') ? '' : (c.cls||'');
              const val = c.key === 'elapsed_min' && typeof row[c.key] === 'number'
                ? row[c.key].toLocaleString(undefined, {maximumFractionDigits: 2})
                : fmtVal(row[c.key]);
              return `<td class="${cls}">${esc(val)}</td>`;
            }).join('')}</tr>`;
          }).join('')}
        </tbody>
      </table>`;
  }

  // ── rename ROIs ──
  function toggleRenameMode() {
    renameMode = document.getElementById('rename-mode').checked;
    if (measurements.length) renderTable(measurements);
    else renderRoiSetupTable();
    renderSummary();
  }

  function syncRoiLabelsFromInputs() {
    document.querySelectorAll('#results-table-wrap input[data-roi-index]').forEach(inp => {
      const idx = Number(inp.dataset.roiIndex);
      if (Number.isInteger(idx) && rois[idx]) {
        rois[idx].label = inp.value.trim();
      }
    });
  }

  function renameMeasurement(idx, newLabel) {
    const m = measurements[idx];
    if (!m || m.roi_type === 'bckg') return;
    const roiIndex = m.roi_index;
    const cleanLabel = newLabel.trim();
    if (roiIndex !== undefined && rois[roiIndex]) {
      rois[roiIndex].label = cleanLabel;
    }
    measurements.forEach(row => {
      if (row.roi_index === roiIndex && row.roi_type !== 'bckg') {
        row.roi_label = cleanLabel;
        if (row._uid !== undefined) {
          const ad = analysisData.find(d => d.uid === row._uid);
          if (ad) ad.roiLabel = cleanLabel;
        }
      }
    });
    if (roiIndex === undefined) {
      m.roi_label = cleanLabel;
    }
    drawAll();
    renderSummary();
    scheduleRoiAutosave();
  }

  // ── save ──
  function analysisStem() {
    return biolumPair ? biolumPair.stem : (dayPair ? dayPair.stem : 'analysis');
  }

  function scaledRoisForSave() {
    syncRoiLabelsFromInputs();
    ensureRoiNumbers();
    return scaledRoisToJpeg();
  }

  function scheduleRoiAutosave(force=false) {
    if (roiAutosaveTimer) clearTimeout(roiAutosaveTimer);
    roiAutosaveTimer = setTimeout(() => autosaveRoisNow(force), 700);
  }

  async function autosaveRoisNow(force=false) {
    if ((!dayPair && !biolumPair) || (!force && !rois.length) || !currentFolder || dayCanvas.width < 1) return;
    try {
      const r = await fetch('/autosave_rois', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({
          folder: currentFolder,
          save_dir: lastSaveDir,
          stem: analysisStem(),
          rois: scaledRoisForSave(),
          jpeg_size: [jpegSize.w, jpegSize.h],
        })
      });
      const data = await r.json();
      if (data.ok && data.dir) lastSaveDir = data.dir;
    } catch(e) {}
  }

  function promptSaveResultsSoon() {
    if (savePromptShown) return;
    savePromptShown = true;
    setTimeout(() => {
      if (unsavedResults && confirm('Measurements are complete. Save results now?')) {
        saveResults();
      }
    }, 200);
  }

  async function saveResults() {
    if ((!dayPair && !biolumPair) || !rois.length) { setStatus('Nothing to save', 'wrn'); return; }

    const scaledRois = scaledRoisForSave();

    const saveStem = analysisStem();
    const snapshot_png = captureSnapshotDataURL() || '';
    setStatus('Saving...', '');
    const r = await fetch('/save', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({
        folder: currentFolder,
        save_dir: lastSaveDir,
        stem: saveStem,
        rois: scaledRois,
        measurements,
        session_measurements: sessionMeasurements,
        jpeg_size: [jpegSize.w, jpegSize.h],
        snapshot_png,
      })
    });
    const data = await r.json();
    if (data.ok) {
      lastSaveDir = data.dir;
      unsavedResults = false;
      savePromptShown = false;
      setStatus('Saved to: ' + data.dir, 'ok');
      const links = document.getElementById('download-links');
      links.innerHTML = `
        <a href="/download?path=${encodeURIComponent(fixPath(data.dir) + '/session_measurements.xlsx')}"
           style="font-family:var(--mono);font-size:10px;color:var(--amber);text-decoration:none;border:1px solid var(--border2);padding:2px 8px;border-radius:2px;">
           ↓ Excel</a>`;
    } else {
      setStatus('Save failed: ' + data.error, 'err');
    }
  }

  // ── load ROIs ──
  async function showLoadRoiDialog() {
    if (!currentFolder) { setStatus('Load a folder first', 'wrn'); return; }
    document.getElementById('roi-search-folder').value = currentFolder;
    await refreshRoiFileList();
    document.getElementById('roi-dialog').style.display = 'flex';
  }

  async function pickRoiSearchFolder() {
    const r = await fetch('/pick_folder');
    const data = await r.json();
    if (!data.path) return;
    document.getElementById('roi-search-folder').value = data.path;
    await refreshRoiFileList();
  }

  async function refreshRoiFileList() {
    const folder = document.getElementById('roi-search-folder').value || currentFolder;
    const r = await fetch('/list_roi_files', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({folder})
    });
    const data = await r.json();
    const list = document.getElementById('roi-file-list');
    if (data.error) {
      list.innerHTML = `<div style="font-family:var(--mono);font-size:11px;color:var(--warn);">${escHtml(data.error)}</div>`;
      return;
    }
    if (!data.files.length) {
      list.innerHTML = '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">No saved ROI files found in this folder.</div>';
    } else {
      list.innerHTML = data.files.map(f => `
        <div data-path="${escHtml(f.path)}" onclick="loadRoiFile(this.dataset.path)"
             style="padding:8px 10px;border:1px solid var(--border2);border-radius:3px;cursor:pointer;font-family:var(--mono);font-size:11px;color:var(--text);">
          <div>${escHtml(f.name)}</div>
          <div style="font-size:9px;color:var(--muted);margin-top:3px;">${escHtml(f.folder)}</div>
        </div>`).join('');
    }
  }

  function clearPanel(side) {
    if (side === 'day') {
      dayPair = null; dayImg = null;
      dayCanvas.style.display = 'none';
      document.getElementById('day-placeholder').style.display = '';
      document.getElementById('day-clear-btn').style.display = 'none';
      setSampleLabel('day', null);
      document.querySelectorAll('.pair-item.active-day').forEach(e => e.classList.remove('active', 'active-day'));
      rois = []; measurements = []; selectedRoi = -1;
      document.getElementById('results-table-wrap').innerHTML =
        '<div style="font-family:var(--mono);font-size:11px;color:var(--muted);">Draw ROIs then click ▶ Measure ROIs</div>';
      document.getElementById('download-links').innerHTML = '';
    } else {
      biolumPair = null; biolumImg = null;
      biolumCanvas.style.display = 'none';
      document.getElementById('biolum-placeholder').style.display = '';
      document.getElementById('biolum-clear-btn').style.display = 'none';
      document.getElementById('tl-frame-controls').style.display = 'none';
      setSampleLabel('biolum', null);
      document.querySelectorAll('.pair-item.active-biolum').forEach(e => e.classList.remove('active', 'active-biolum'));
      clearCurrentMeasurementView();
    }
    drawAll();
    setStatus('Panel cleared. Select a new image.', '');
  }

  // ── background type toggle ──
  function toggleBckgType() {
    roiType = roiType === 'bckg' ? 'roi' : 'bckg';
    document.getElementById('type-bckg').classList.toggle('active', roiType === 'bckg');
  }

  // ── top-level tab switching ──
  function switchTopTab(tab) {
    ['measure','summary'].forEach(t => {
      document.getElementById('top-tab-' + t).classList.toggle('active', t === tab);
      document.getElementById('top-tab-btn-' + t).classList.toggle('active', t === tab);
    });
    if (tab === 'summary') renderSummary();
  }

  // ── analysis summary / box plots ──
  function clearAnalysisData() {
    analysisData = [];
    summarySampleOrder = [];
    renderSummary();
  }

  // ── Tukey HSD + CLD ──

  // Studentized range critical values q(α=0.05, k, df) — Tukey-Kramer table
  const _TDF = [4,5,6,7,8,9,10,12,15,20,24,30,40,60,120,Infinity];
  const _TQCRIT = {
    2:[3.93,3.64,3.46,3.34,3.26,3.20,3.15,3.08,3.01,2.95,2.92,2.89,2.86,2.83,2.80,2.77],
    3:[5.04,4.60,4.34,4.16,4.04,3.95,3.88,3.77,3.67,3.58,3.53,3.49,3.44,3.40,3.36,3.31],
    4:[5.76,5.22,4.90,4.68,4.53,4.41,4.33,4.20,4.08,3.96,3.90,3.85,3.79,3.74,3.68,3.63],
    5:[6.29,5.67,5.30,5.06,4.89,4.76,4.65,4.51,4.37,4.23,4.17,4.10,4.04,3.98,3.92,3.86],
    6:[6.71,6.03,5.63,5.36,5.17,5.02,4.91,4.75,4.59,4.45,4.37,4.30,4.23,4.16,4.10,4.03],
    7:[7.05,6.33,5.90,5.61,5.40,5.24,5.12,4.95,4.78,4.62,4.54,4.46,4.39,4.31,4.24,4.17],
    8:[7.35,6.58,6.12,5.82,5.60,5.43,5.30,5.12,4.94,4.77,4.68,4.60,4.52,4.44,4.36,4.29],
  };

  function tukeyQCrit(k, df) {
    const row = _TQCRIT[Math.min(Math.max(k, 2), 8)];
    if (df <= _TDF[0]) return row[0];
    if (!isFinite(df)) return row[row.length - 1];
    for (let i = 0; i < _TDF.length - 1; i++) {
      if (df >= _TDF[i] && df <= _TDF[i+1]) {
        const t = (df - _TDF[i]) / (_TDF[i+1] - _TDF[i]);
        return row[i] + t * (row[i+1] - row[i]);
      }
    }
    return row[row.length - 1];
  }

  function tukeyHSD(groupData) {
    // groupData: [{name, vals}]
    const k = groupData.length;
    if (k < 2) return null;
    const grps = groupData.map(g => {
      const n = g.vals.length, mean = g.vals.reduce((a,b)=>a+b,0)/n;
      const ss = g.vals.reduce((a,b)=>a+(b-mean)**2, 0);
      return {name:g.name, n, mean, ss};
    });
    const N = grps.reduce((a,g)=>a+g.n, 0);
    const df = N - k;
    if (df < 1) return null;
    const MSE = grps.reduce((a,g)=>a+g.ss,0) / df;
    if (MSE <= 0) return null;
    const qCrit = tukeyQCrit(k, df);
    const sig = Array.from({length:k}, ()=>Array(k).fill(false));
    for (let i = 0; i < k; i++)
      for (let j = i+1; j < k; j++) {
        const q = Math.abs(grps[i].mean - grps[j].mean) /
                  (Math.sqrt(MSE) * Math.sqrt(0.5*(1/grps[i].n + 1/grps[j].n)));
        sig[i][j] = sig[j][i] = q > qCrit;
      }
    return {sig};
  }

  function computeCLD(k, sig, sortedIdx) {
    // Piepho (2004) absorption algorithm
    if (k === 0) return [];
    if (k === 1) return ['a'];
    const sets = Array.from({length:k}, ()=>new Set([0]));
    let next = 1, changed = true;
    while (changed) {
      changed = false;
      for (let pi = 0; pi < k; pi++) {
        const i = sortedIdx[pi];
        for (let pj = pi+1; pj < k; pj++) {
          const j = sortedIdx[pj];
          if (!sig[i][j]) continue;
          const common = [...sets[i]].filter(l=>sets[j].has(l));
          if (!common.length) continue;
          changed = true;
          for (const l of common) {
            const lp = next++;
            const Li = new Set(), Lj = new Set();
            for (let m = 0; m < k; m++) {
              if (!sets[m].has(l)) continue;
              if (!sig[i][m]) Li.add(m);
              if (!sig[j][m]) Lj.add(m);
            }
            for (let m = 0; m < k; m++) {
              if (!sets[m].has(l)) continue;
              if (Li.has(m) && !Lj.has(m)) { sets[m].delete(l); sets[m].add(lp); }
              else if (Li.has(m) &&  Lj.has(m)) { sets[m].add(lp); }
              // Lj only or neither: keep l unchanged
            }
          }
        }
      }
    }
    // Build ordered letter map: first letter of highest-mean group → 'a'
    const order = [];
    for (const si of sortedIdx)
      for (const l of [...sets[si]].sort((a,b)=>a-b))
        if (!order.includes(l)) order.push(l);
    const alpha = 'abcdefghijklmnop';
    const lmap = Object.fromEntries(order.map((l,i)=>[l, alpha[i]]));
    return Array.from({length:k}, (_,i) =>
      [...sets[i]].sort((a,b)=>order.indexOf(a)-order.indexOf(b)).map(l=>lmap[l]).join('')
    );
  }

  function computeStats(values) {
    if (!values.length) return null;
    const s = [...values].sort((a,b)=>a-b);
    const n = s.length;
    const q = p => { const i=p*(n-1); const lo=Math.floor(i),hi=Math.ceil(i); return s[lo]+(s[hi]-s[lo])*(i-lo); };
    return {min:s[0], q1:q(0.25), median:q(0.5), q3:q(0.75), max:s[n-1], values:s};
  }

  function summaryGroupKey(d) { return renameMode ? (d.roiLabel || d.sample) : d.sample; }
  function hasTimeSeriesData() {
    return analysisData.some(d => d.elapsed_min !== undefined && d.elapsed_min !== null && isFinite(d.elapsed_min));
  }
  function timeSeriesKey(d) {
    return summaryGroupKey(d) || 'sample';
  }
  function summaryGroups(useTimeSeries) {
    const rawGroups = [...new Set(analysisData.map(useTimeSeries ? timeSeriesKey : summaryGroupKey))];
    summarySampleOrder = summarySampleOrder.filter(name => rawGroups.includes(name));
    rawGroups.forEach(name => {
      if (!summarySampleOrder.includes(name)) summarySampleOrder.push(name);
    });
    return summarySampleOrder.length ? summarySampleOrder.filter(name => rawGroups.includes(name)) : rawGroups;
  }
  function updateSummaryOrderControls(groups, useTimeSeries) {
    const wrap = document.getElementById('summary-order-controls');
    const sel = document.getElementById('summary-order-sample');
    if (!wrap || !sel) return;
    if (useTimeSeries || groups.length < 2) {
      wrap.style.display = 'none';
      return;
    }
    const prev = sel.value;
    wrap.style.display = 'flex';
    sel.innerHTML = groups.map(name => `<option value="${escHtml(name)}">${escHtml(name)}</option>`).join('');
    sel.value = groups.includes(prev) ? prev : groups[0];
  }
  function moveSummarySample(dir) {
    const sel = document.getElementById('summary-order-sample');
    if (!sel || !sel.value) return;
    const idx = summarySampleOrder.indexOf(sel.value);
    const nextIdx = idx + dir;
    if (idx < 0 || nextIdx < 0 || nextIdx >= summarySampleOrder.length) return;
    [summarySampleOrder[idx], summarySampleOrder[nextIdx]] = [summarySampleOrder[nextIdx], summarySampleOrder[idx]];
    renderSummary();
    const updated = document.getElementById('summary-order-sample');
    if (updated) updated.value = summarySampleOrder[nextIdx];
  }
  function timeSeriesColor(i) {
    const palette = ['#e7c84b', '#5bd6c6', '#ff8b72', '#8fb8ff', '#d59bff', '#74d36f', '#ffb45c', '#f27bb2'];
    return palette[i % palette.length];
  }

  function useBackgroundSubtracted() {
    return document.getElementById('summary-use-bckg')?.checked ?? true;
  }

  function valueKey(rawKey, nbKey, d = null) {
    if (!useBackgroundSubtracted()) return rawKey;
    return !d || d.has_bckg ? nbKey : rawKey;
  }

  function channelPlotValue(d, ch) {
    return d[valueKey(ch.rawKey, ch.nbKey, d)];
  }

  function channelPlotLabel(ch) {
    return useBackgroundSubtracted() ? ch.nbLabel : ch.rawLabel;
  }

  function ratioSpec() {
    const num = document.getElementById('ratio-num')?.value || 'G';
    const den = document.getElementById('ratio-den')?.value || 'R';
    const numCh = RATIO_CHANNELS[num];
    const denCh = RATIO_CHANNELS[den];
    return {
      num, den,
      numKey: valueKey(numCh.rawKey, numCh.nbKey),
      denKey: valueKey(denCh.rawKey, denCh.nbKey),
      useBckg: useBackgroundSubtracted(),
      label: `${numCh.label}/${denCh.label}`,
      shortLabel: `${numCh.short}/${denCh.short}`,
    };
  }

  function ratioValue(d) {
    const spec = ratioSpec();
    if (spec.useBckg && !d.has_bckg) return null;
    const numCh = RATIO_CHANNELS[spec.num];
    const denCh = RATIO_CHANNELS[spec.den];
    const num = d[valueKey(numCh.rawKey, numCh.nbKey, d)];
    const den = d[valueKey(denCh.rawKey, denCh.nbKey, d)];
    if (!validRatioPair(num, den, spec, ratioDenominatorFloor(spec))) return null;
    return num / den;
  }

  function validRatioPair(num, den, spec, denFloor) {
    if (num === null || num === undefined || den === null || den === undefined) return false;
    if (!isFinite(num) || !isFinite(den)) return false;
    if (spec.useBckg) return num > 0 && den > denFloor;
    return num >= 0 && den > 0;
  }

  function meanFinite(vals) {
    const finite = vals.filter(v => v !== null && v !== undefined && isFinite(v));
    return finite.length ? finite.reduce((sum, v) => sum + v, 0) / finite.length : null;
  }

  function medianValue(vals) {
    if (!vals.length) return null;
    const sorted = [...vals].sort((a, b) => a - b);
    const mid = Math.floor(sorted.length / 2);
    return sorted.length % 2 ? sorted[mid] : (sorted[mid - 1] + sorted[mid]) / 2;
  }

  function ratioDenominatorFloor(spec = ratioSpec()) {
    const positiveDens = analysisData
      .map(d => spec.useBckg && !d.has_bckg ? null : d[valueKey(RATIO_CHANNELS[spec.den].rawKey, RATIO_CHANNELS[spec.den].nbKey, d)])
      .filter(v => v !== null && v !== undefined && isFinite(v) && v > 0);
    const med = medianValue(positiveDens);
    return med !== null ? Math.max(med * 0.01, Number.EPSILON) : Number.EPSILON;
  }

  function ratioFromRows(rows, spec = ratioSpec(), denFloor = ratioDenominatorFloor(spec)) {
    const numCh = RATIO_CHANNELS[spec.num];
    const denCh = RATIO_CHANNELS[spec.den];
    const paired = rows
      .filter(d => !spec.useBckg || d.has_bckg)
      .map(d => ({
        num: d[valueKey(numCh.rawKey, numCh.nbKey, d)],
        den: d[valueKey(denCh.rawKey, denCh.nbKey, d)],
      }))
      .filter(p => validRatioPair(p.num, p.den, spec, denFloor));
    if (!paired.length) return null;
    const numMean = meanFinite(paired.map(p => p.num));
    const denMean = meanFinite(paired.map(p => p.den));
    if (!validRatioPair(numMean, denMean, spec, denFloor)) return null;
    return numMean / denMean;
  }

  function timeSeriesRatioPoints(seriesName, spec = ratioSpec()) {
    const denFloor = ratioDenominatorFloor(spec);
    const byTime = new Map();
    analysisData
      .filter(d =>
        timeSeriesKey(d) === seriesName &&
        d.elapsed_min !== undefined && d.elapsed_min !== null && isFinite(d.elapsed_min)
      )
      .forEach(d => {
        const t = d.elapsed_min;
        if (!byTime.has(t)) byTime.set(t, []);
        byTime.get(t).push(d);
      });
    return [...byTime.entries()]
      .map(([elapsed_min, rows]) => ({
        elapsed_min: Number(elapsed_min),
        value: ratioFromRows(rows, spec, denFloor),
      }))
      .filter(p => p.value !== null && p.value !== undefined && isFinite(p.value))
      .sort((a, b) => a.elapsed_min - b.elapsed_min);
  }

  function timeSeriesDenominatorPoints(seriesName, spec = ratioSpec()) {
    const denCh = RATIO_CHANNELS[spec.den];
    const byTime = new Map();
    analysisData
      .filter(d =>
        timeSeriesKey(d) === seriesName &&
        (!spec.useBckg || d.has_bckg) &&
        d.elapsed_min !== undefined && d.elapsed_min !== null && isFinite(d.elapsed_min)
      )
      .forEach(d => {
        const t = d.elapsed_min;
        if (!byTime.has(t)) byTime.set(t, []);
        byTime.get(t).push(d[valueKey(denCh.rawKey, denCh.nbKey, d)]);
      });
    return [...byTime.entries()]
      .map(([elapsed_min, vals]) => ({
        elapsed_min: Number(elapsed_min),
        value: meanFinite(vals),
      }))
      .filter(p => p.value !== null && p.value !== undefined && isFinite(p.value))
      .sort((a, b) => a.elapsed_min - b.elapsed_min);
  }

  function formatRatioTitle(spec, groups = []) {
    const mode = spec.useBckg ? 'Bckg-subtracted' : 'Raw';
    const denVals = groups.flatMap(name => timeSeriesDenominatorPoints(name, spec).map(p => p.value));
    if (!denVals.length) return `${spec.label} ratio - ${mode}`;
    return `${spec.label} ratio - ${mode} (${spec.den} denom ${fmtVal(Math.min(...denVals))}-${fmtVal(Math.max(...denVals))})`;
  }

  function timeAxisRange() {
    const times = analysisData
      .map(d => d.elapsed_min)
      .filter(v => v !== null && v !== undefined && isFinite(v));
    if (!times.length) return null;
    return {xMin: Math.min(...times), xMax: Math.max(...times)};
  }

  function plotValue(d, key) {
    if (key === 'ratio') return ratioValue(d);
    const ch = SUMMARY_CHANNELS.find(c => c.rawKey === key || c.nbKey === key || c.channel === key);
    return ch ? channelPlotValue(d, ch) : d[key];
  }

  function summaryYRange() {
    const allVals = SUMMARY_CHANNELS.flatMap(ch =>
      analysisData.map(d => channelPlotValue(d, ch)).filter(v => v != null && isFinite(v))
    );
    if (!allVals.length) return {yMin: 0, yMax: 1};
    const gMin = Math.min(0, ...allVals), gMax = Math.max(0, ...allVals);
    const pad = (gMax - gMin) * 0.12 || Math.abs(gMax) * 0.12 || 1;
    return {yMin: gMin - pad, yMax: gMax + pad};
  }

  function ratioYRange(useTimeSeries = false, seriesNames = []) {
    const spec = ratioSpec();
    const allVals = useTimeSeries
      ? seriesNames.flatMap(name => timeSeriesRatioPoints(name, spec).map(p => p.value))
      : analysisData.map(d => ratioValue(d)).filter(v => v != null && isFinite(v));
    if (!allVals.length) return null;
    const gMin = Math.min(...allVals), gMax = Math.max(...allVals);
    const pad = (gMax - gMin) * 0.12 || Math.abs(gMax) * 0.12 || 0.1;
    return {yMin: gMin - pad, yMax: gMax + pad};
  }

  function renderSummary() {
    const empty = document.getElementById('summary-empty');
    const plots = document.getElementById('summary-plots');
    const ratioPlot = document.getElementById('summary-ratio-plot');
    if (!analysisData.length) {
      empty.style.display = '';
      plots.style.display = 'none';
      if (ratioPlot) ratioPlot.style.display = 'none';
      const orderControls = document.getElementById('summary-order-controls');
      if (orderControls) orderControls.style.display = 'none';
      return;
    }
    empty.style.display = 'none';
    plots.style.display = 'flex';
    const useTimeSeries = hasTimeSeriesData();
    const groups = summaryGroups(useTimeSeries);
    updateSummaryOrderControls(groups, useTimeSeries);
    const showRatio = document.getElementById('summary-show-ratio')?.checked ?? false;
    if (ratioPlot) ratioPlot.style.display = showRatio ? 'grid' : 'none';
    const {yMin, yMax} = summaryYRange();
    const plotsDiv = document.getElementById('summary-plots');
    const gap = 16;
    const n = SUMMARY_CHANNELS.length;
    const canvasW = Math.floor((plotsDiv.clientWidth  - gap * (n - 1)) / n) || 400;
    const canvasH = plotsDiv.clientHeight || 500;
    SUMMARY_CHANNELS.forEach(ch => {
      const canvas = document.getElementById(ch.id);
      canvas.width  = canvasW;
      canvas.height = canvasH;
      canvas.style.width  = canvasW + 'px';
      canvas.style.height = canvasH + 'px';
      const title = channelPlotLabel(ch);
      if (useTimeSeries) {
        drawTimeSeriesPlot(canvas, groups, ch.channel, ch.color, title, yMin, yMax, 'dark');
      } else {
        drawBoxPlot(canvas, groups, ch.channel, ch.color, title, yMin, yMax, 'dark');
      }
    });

    const ratioRange = showRatio ? ratioYRange(useTimeSeries, groups) : null;
    const ratioCanvas = document.getElementById('plot-ratio');
    const ratioWrap = document.getElementById('summary-ratio-plot');
    if (showRatio && ratioCanvas && ratioWrap) {
      const ratioW = canvasW;
      const ratioH = ratioWrap.clientHeight || 260;
      ratioCanvas.width = ratioW;
      ratioCanvas.height = ratioH;
      ratioCanvas.style.width = ratioW + 'px';
      ratioCanvas.style.height = ratioH + 'px';
      const spec = ratioSpec();
      const title = formatRatioTitle(spec, groups);
      if (ratioRange) {
        if (useTimeSeries) {
          drawTimeSeriesPlot(ratioCanvas, groups, 'ratio', '#e7c84b', title, ratioRange.yMin, ratioRange.yMax, 'dark');
        } else {
          drawBoxPlot(ratioCanvas, groups, 'ratio', '#e7c84b', title, ratioRange.yMin, ratioRange.yMax, 'dark');
        }
      } else {
        const ctx = ratioCanvas.getContext('2d');
        ctx.clearRect(0, 0, ratioW, ratioH);
        ctx.fillStyle = '#0b0f15';
        ctx.fillRect(0, 0, ratioW, ratioH);
        ctx.fillStyle = '#7090a8';
        ctx.font = '12px IBM Plex Mono';
        ctx.textAlign = 'center';
        ctx.fillText('No finite normalized values for ' + spec.shortLabel, ratioW / 2, ratioH / 2);
      }
    }
  }

  function fmtVal(v) {
    const a = Math.abs(v);
    if (a === 0) return '0';
    if (a >= 1e9) return (v/1e9).toFixed(2) + ' G';
    if (a >= 1e6) return (v/1e6).toFixed(2) + ' M';
    if (a >= 1e3) return (v/1e3).toFixed(1) + ' k';
    if (a >= 1)   return v.toFixed(1);
    return v.toPrecision(3);
  }

  function drawBoxPlot(canvas, samples, key, color, title, yMinOverride = null, yMaxOverride = null, theme = 'dark') {
    const ctx = canvas.getContext('2d');
    const W = canvas.width, H = canvas.height;
    ctx.clearRect(0, 0, W, H);

    const isLight = theme === 'light';
    const bgCol    = isLight ? '#ffffff' : '#0b0f15';
    const gridCol  = isLight ? '#cccccc' : '#1e2d42';
    const axisCol  = isLight ? '#666666' : '#2e4060';
    const tickCol  = isLight ? '#333333' : '#7090a8';
    const xLblCol  = isLight ? '#111111' : '#c8dff0';
    const medCol   = isLight ? '#000000' : '#ffffff';

    ctx.fillStyle = bgCol; ctx.fillRect(0, 0, W, H);

    // determine Y range
    const allVals = analysisData.map(d => plotValue(d, key)).filter(v => v != null && isFinite(v));
    if (!allVals.length) return;
    let yMin, yMax;
    if (yMinOverride !== null && yMaxOverride !== null) {
      yMin = yMinOverride; yMax = yMaxOverride;
    } else {
      yMin = Math.min(...allVals); yMax = Math.max(...allVals);
      const pad = (yMax - yMin) * 0.12 || Math.abs(yMax) * 0.12 || 1;
      yMin -= pad; yMax += pad;
    }

    ctx.font = '13px IBM Plex Mono';
    let maxLabelW = 0;
    for (let t = 0; t <= 5; t++) {
      const v = yMin + (yMax - yMin) * t / 5;
      maxLabelW = Math.max(maxLabelW, ctx.measureText(fmtVal(v)).width);
    }
    const mg = {top: 44, right: 20, bottom: 80, left: Math.ceil(maxLabelW) + 20};
    const pw = W - mg.left - mg.right;
    const ph = H - mg.top  - mg.bottom;
    const toY = v => mg.top + ph - (v - yMin) / (yMax - yMin) * ph;

    // title
    ctx.fillStyle = color;
    ctx.font = 'bold 15px IBM Plex Mono';
    ctx.textAlign = 'center';
    ctx.fillText(title, W / 2, 26);

    // grid lines + y-axis ticks
    const ticks = 5;
    for (let t = 0; t <= ticks; t++) {
      const v = yMin + (yMax - yMin) * t / ticks;
      const y = toY(v);
      ctx.strokeStyle = gridCol; ctx.lineWidth = 0.5;
      ctx.beginPath(); ctx.moveTo(mg.left, y); ctx.lineTo(mg.left + pw, y); ctx.stroke();
      ctx.fillStyle = tickCol;
      ctx.font = '13px IBM Plex Mono';
      ctx.textAlign = 'right';
      ctx.fillText(fmtVal(v), mg.left - 8, y + 5);
    }

    // axes
    ctx.strokeStyle = axisCol; ctx.lineWidth = 1.5;
    ctx.beginPath();
    ctx.moveTo(mg.left, mg.top); ctx.lineTo(mg.left, mg.top + ph);
    ctx.lineTo(mg.left + pw, mg.top + ph); ctx.stroke();

    const slotW = pw / samples.length;
    const boxW = Math.min(slotW * 0.5, 70);

    const sampleData = samples.map(s => ({
      sample: s,
      vals: analysisData.filter(d => summaryGroupKey(d) === s).map(d => plotValue(d, key)).filter(v => v != null && isFinite(v)),
    }));
    const stats = sampleData.map(sd => ({...sd, stats: computeStats(sd.vals)}));

    stats.forEach(({sample, vals, stats: st}, si) => {
      const cx = mg.left + (si + 0.5) * slotW;

      // x-axis label
      ctx.fillStyle = xLblCol;
      ctx.font = '13px IBM Plex Mono';
      ctx.textAlign = 'center';
      const maxChars = Math.floor(slotW / 8);
      const lbl = sample.length > maxChars ? sample.slice(0, maxChars - 1) + '…' : sample;
      ctx.fillText(lbl, cx, mg.top + ph + 22);

      if (!vals.length) return;

      // jittered data points
      const ptAlpha = isLight ? 'aa' : 'bb';
      vals.forEach(v => {
        const jitter = (Math.random() - 0.5) * boxW * 0.55;
        ctx.beginPath();
        ctx.arc(cx + jitter, toY(v), 4, 0, Math.PI * 2);
        ctx.fillStyle = color + ptAlpha; ctx.fill();
      });

      if (!st || vals.length < 2) return;

      const capW = boxW * 0.35;

      // whiskers + caps
      const wAlpha = isLight ? 'dd' : 'cc';
      ctx.strokeStyle = color + wAlpha; ctx.lineWidth = 1.5;
      ctx.beginPath();
      ctx.moveTo(cx, toY(st.min)); ctx.lineTo(cx, toY(st.q1));
      ctx.moveTo(cx, toY(st.q3)); ctx.lineTo(cx, toY(st.max));
      ctx.moveTo(cx - capW, toY(st.min)); ctx.lineTo(cx + capW, toY(st.min));
      ctx.moveTo(cx - capW, toY(st.max)); ctx.lineTo(cx + capW, toY(st.max));
      ctx.stroke();

      // IQR box
      const qy = toY(st.q3), qh = toY(st.q1) - toY(st.q3);
      const boxAlpha = isLight ? '28' : '30';
      ctx.fillStyle = color + boxAlpha; ctx.strokeStyle = color; ctx.lineWidth = 2;
      ctx.fillRect(cx - boxW/2, qy, boxW, qh);
      ctx.strokeRect(cx - boxW/2, qy, boxW, qh);

      // median line
      const my = toY(st.median);
      ctx.strokeStyle = medCol; ctx.lineWidth = 2.5;
      ctx.beginPath(); ctx.moveTo(cx - boxW/2, my); ctx.lineTo(cx + boxW/2, my); ctx.stroke();
    });

    // ── Tukey HSD CLD letters ──
    if (stats.length >= 2) {
      const tukey = tukeyHSD(stats.map(s => ({name: s.sample, vals: s.vals})));
      if (tukey) {
        const sortedIdx = [...Array(stats.length).keys()]
          .sort((a, b) => (stats[b].stats ? stats[b].stats.median : 0) -
                          (stats[a].stats ? stats[a].stats.median : 0));
        const cld = computeCLD(stats.length, tukey.sig, sortedIdx);
        ctx.font = 'bold 13px IBM Plex Mono';
        ctx.textAlign = 'center';
        ctx.fillStyle = isLight ? '#222222' : '#dddddd';
        stats.forEach(({vals, stats: st}, si) => {
          if (!vals.length || !cld[si]) return;
          const cx = mg.left + (si + 0.5) * slotW;
          const topY = st ? toY(st.max) : Math.min(...vals.map(v => toY(v)));
          ctx.fillText(cld[si], cx, topY - 6);
        });
      }
    }

    // ── legend ──
    ctx.font = '9px IBM Plex Mono';
    ctx.textAlign = 'left';
    ctx.fillStyle = isLight ? '#888888' : '#506070';
    ctx.fillText('Tukey HSD α=0.05 — groups sharing a letter are not significantly different', mg.left, H - 6);
  }

  // ── PDF export ──
  function drawTimeSeriesPlot(canvas, seriesNames, key, color, title, yMinOverride = null, yMaxOverride = null, theme = 'dark') {
    const ctx = canvas.getContext('2d');
    const W = canvas.width, H = canvas.height;
    ctx.clearRect(0, 0, W, H);

    const isLight = theme === 'light';
    const bgCol    = isLight ? '#ffffff' : '#0b0f15';
    const gridCol  = isLight ? '#cccccc' : '#1e2d42';
    const axisCol  = isLight ? '#666666' : '#2e4060';
    const tickCol  = isLight ? '#333333' : '#7090a8';
    const labelCol = isLight ? '#111111' : '#c8dff0';

    ctx.fillStyle = bgCol;
    ctx.fillRect(0, 0, W, H);

    const values = key === 'ratio'
      ? seriesNames.flatMap(name =>
          timeSeriesRatioPoints(name).map(p => ({
            elapsed_min: p.elapsed_min,
            _plotValue: p.value,
          }))
        )
      : analysisData
          .map(d => ({...d, _plotValue: plotValue(d, key)}))
          .filter(d =>
            d.elapsed_min !== undefined && d.elapsed_min !== null && isFinite(d.elapsed_min) &&
            d._plotValue !== null && d._plotValue !== undefined && isFinite(d._plotValue)
          );
    if (!values.length) return;

    const sharedTimeRange = key === 'ratio' ? timeAxisRange() : null;
    let xMin = sharedTimeRange ? sharedTimeRange.xMin : Math.min(...values.map(d => d.elapsed_min));
    let xMax = sharedTimeRange ? sharedTimeRange.xMax : Math.max(...values.map(d => d.elapsed_min));
    if (xMax === xMin) xMax = xMin + 1;

    let yMin, yMax;
    if (yMinOverride !== null && yMaxOverride !== null) {
      yMin = yMinOverride;
      yMax = yMaxOverride;
    } else {
      yMin = Math.min(...values.map(d => d._plotValue));
      yMax = Math.max(...values.map(d => d._plotValue));
      const pad = (yMax - yMin) * 0.12 || Math.abs(yMax) * 0.12 || 1;
      yMin -= pad;
      yMax += pad;
    }

    ctx.font = '13px IBM Plex Mono';
    let maxLabelW = 0;
    for (let t = 0; t <= 5; t++) {
      const v = yMin + (yMax - yMin) * t / 5;
      maxLabelW = Math.max(maxLabelW, ctx.measureText(fmtVal(v)).width);
    }

    const mg = {top: 86, right: 20, bottom: 56, left: Math.ceil(maxLabelW) + 20};
    const pw = W - mg.left - mg.right;
    const ph = H - mg.top - mg.bottom;
    const toX = v => mg.left + (v - xMin) / (xMax - xMin) * pw;
    const toY = v => mg.top + ph - (v - yMin) / (yMax - yMin) * ph;

    ctx.fillStyle = color;
    ctx.font = 'bold 15px IBM Plex Mono';
    ctx.textAlign = 'center';
    ctx.fillText(title + ' over time', W / 2, 26);

    const legend = seriesNames.slice(0, 6);
    ctx.font = '9px IBM Plex Mono';
    ctx.textAlign = 'left';
    let legendX = mg.left;
    let legendY = 44;
    legend.forEach((name, i) => {
      const shortName = name.length > 18 ? name.slice(0, 17) + '...' : name;
      const itemW = Math.min(150, ctx.measureText(shortName).width + 30);
      if (legendX + itemW > W - mg.right) {
        legendX = mg.left;
        legendY += 12;
      }
      const lineCol = timeSeriesColor(i);
      ctx.strokeStyle = lineCol;
      ctx.lineWidth = 1.6;
      ctx.beginPath();
      ctx.moveTo(legendX, legendY - 3);
      ctx.lineTo(legendX + 14, legendY - 3);
      ctx.stroke();
      ctx.fillStyle = lineCol;
      ctx.fillText(shortName, legendX + 18, legendY);
      legendX += itemW;
    });

    for (let t = 0; t <= 5; t++) {
      const v = yMin + (yMax - yMin) * t / 5;
      const y = toY(v);
      ctx.strokeStyle = gridCol;
      ctx.lineWidth = 0.5;
      ctx.beginPath();
      ctx.moveTo(mg.left, y);
      ctx.lineTo(mg.left + pw, y);
      ctx.stroke();
      ctx.fillStyle = tickCol;
      ctx.font = '13px IBM Plex Mono';
      ctx.textAlign = 'right';
      ctx.fillText(fmtVal(v), mg.left - 8, y + 5);
    }

    for (let t = 0; t <= 4; t++) {
      const v = xMin + (xMax - xMin) * t / 4;
      const x = toX(v);
      ctx.strokeStyle = gridCol;
      ctx.lineWidth = 0.5;
      ctx.beginPath();
      ctx.moveTo(x, mg.top);
      ctx.lineTo(x, mg.top + ph);
      ctx.stroke();
      ctx.fillStyle = tickCol;
      ctx.font = '12px IBM Plex Mono';
      ctx.textAlign = 'center';
      ctx.fillText(v.toFixed(v >= 10 ? 0 : 1), x, mg.top + ph + 20);
    }

    ctx.strokeStyle = axisCol;
    ctx.lineWidth = 1.5;
    ctx.beginPath();
    ctx.moveTo(mg.left, mg.top);
    ctx.lineTo(mg.left, mg.top + ph);
    ctx.lineTo(mg.left + pw, mg.top + ph);
    ctx.stroke();

    ctx.fillStyle = labelCol;
    ctx.font = '12px IBM Plex Mono';
    ctx.textAlign = 'center';
    ctx.fillText('Elapsed time (min)', mg.left + pw / 2, H - 12);

    seriesNames.forEach((name, si) => {
      let pts;
      if (key === 'ratio') {
        pts = timeSeriesRatioPoints(name);
      } else {
        const byTime = new Map();
        values.filter(d => timeSeriesKey(d) === name).forEach(d => {
          const t = d.elapsed_min;
          if (!byTime.has(t)) byTime.set(t, []);
          byTime.get(t).push(d._plotValue);
        });
        pts = [...byTime.entries()]
          .map(([elapsed_min, vals]) => ({
            elapsed_min: Number(elapsed_min),
            value: vals.reduce((sum, v) => sum + v, 0) / vals.length,
          }))
          .sort((a, b) => a.elapsed_min - b.elapsed_min);
      }
      if (!pts.length) return;
      const lineCol = timeSeriesColor(si);
      ctx.strokeStyle = lineCol;
      ctx.lineWidth = 1.8;
      ctx.beginPath();
      pts.forEach((p, pi) => {
        const x = toX(p.elapsed_min);
        const y = toY(p.value);
        if (pi === 0) ctx.moveTo(x, y);
        else ctx.lineTo(x, y);
      });
      ctx.stroke();
      ctx.save();
      ctx.globalAlpha = isLight ? 0.42 : 0.55;
      pts.forEach(p => {
        ctx.beginPath();
        ctx.arc(toX(p.elapsed_min), toY(p.value), 4.2, 0, Math.PI * 2);
        ctx.fillStyle = lineCol;
        ctx.fill();
      });
      ctx.restore();
    });

  }

  async function exportSummaryPDF() {
    if (!analysisData.length) { setStatus('No data to export', 'wrn'); return; }
    if (!window.jspdf) { setStatus('PDF library not loaded', 'wrn'); return; }
    const useTimeSeries = hasTimeSeriesData();
    const groups = summaryGroups(useTimeSeries);
    const showRatio = document.getElementById('summary-show-ratio')?.checked ?? false;
    const {yMin, yMax} = summaryYRange();

    // Canvas sized to match the per-plot slot on landscape A4 (3 plots, 8mm margins)
    // A4 landscape: 297x210mm → slot ≈ 88mm wide × 194mm tall → ratio 0.454
    const CW = 720, CH = 1590;
    const dataURLs = SUMMARY_CHANNELS.map(ch => {
      const c = document.createElement('canvas');
      c.width = CW; c.height = CH;
      const title = channelPlotLabel(ch);
      if (useTimeSeries) {
        drawTimeSeriesPlot(c, groups, ch.channel, ch.colorExport, title, yMin, yMax, 'light');
      } else {
        drawBoxPlot(c, groups, ch.channel, ch.colorExport, title, yMin, yMax, 'light');
      }
      return c.toDataURL('image/png');
    });

    const {jsPDF} = window.jspdf;
    const pdf = new jsPDF({orientation: 'landscape', unit: 'mm', format: 'a4'});
    const PW = pdf.internal.pageSize.getWidth();
    const PH = pdf.internal.pageSize.getHeight();
    const margin = 8;
    const n = SUMMARY_CHANNELS.length;
    const plotW = (PW - margin * (n + 1)) / n;
    const plotH = plotW * CH / CW;           // maintain canvas aspect ratio — no distortion
    const yOff = (PH - plotH) / 2;          // centre vertically on page
    dataURLs.forEach((url, i) => {
      pdf.addImage(url, 'PNG', margin + i * (plotW + margin), yOff, plotW, plotH);
    });

    const ratioRange = showRatio ? ratioYRange(useTimeSeries, groups) : null;
    if (ratioRange) {
      const spec = ratioSpec();
      const ratioCanvas = document.createElement('canvas');
      ratioCanvas.width = 1440;
      ratioCanvas.height = 810;
      const ratioTitle = formatRatioTitle(spec, groups);
      if (useTimeSeries) {
        drawTimeSeriesPlot(ratioCanvas, groups, 'ratio', '#9a7200', ratioTitle, ratioRange.yMin, ratioRange.yMax, 'light');
      } else {
        drawBoxPlot(ratioCanvas, groups, 'ratio', '#9a7200', ratioTitle, ratioRange.yMin, ratioRange.yMax, 'light');
      }
      const ratioUrl = ratioCanvas.toDataURL('image/png');
      pdf.addPage('a4', 'landscape');
      const ratioW = PW - margin * 2;
      const ratioH = ratioW * ratioCanvas.height / ratioCanvas.width;
      pdf.addImage(ratioUrl, 'PNG', margin, (PH - ratioH) / 2, ratioW, ratioH);
    }

    const pdfDataUrl = pdf.output('datauristring');
    if (lastSaveDir) {
      const r = await fetch('/save_pdf', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({dir: lastSaveDir, pdf_data: pdfDataUrl, filename: 'biolum_summary.pdf'})
      });
      const d = await r.json();
      if (d.ok) setStatus('PDF saved: ' + d.path, 'ok');
      else setStatus('PDF save failed: ' + d.error, 'err');
    } else {
      pdf.save('biolum_summary.pdf');
      setStatus('PDF downloaded — save measurements first to write to folder', 'wrn');
    }
  }

  // ── snapshot ──
  function captureSnapshotDataURL() {
    if (!dayImg && !biolumImg) return null;
    const dW = dayCanvas.width, dH = dayCanvas.height;
    const bW = biolumCanvas.width, bH = biolumCanvas.height;
    const gap = 4;
    const W = (dW||0) + (bW||0) + (dW && bW ? gap : 0);
    const H = Math.max(dH||0, bH||0);
    const off = document.createElement('canvas');
    off.width = W||800; off.height = H||600;
    const ctx = off.getContext('2d');
    ctx.fillStyle = '#06090d'; ctx.fillRect(0, 0, off.width, off.height);

    if (dayImg && dW) {
      ctx.drawImage(dayImg, 0, 0, dW, dH);
      ctx.save(); drawRoisOnCtx(ctx, 1, 1); ctx.restore();
    }
    const bx = (dW||0) + (dW && bW ? gap : 0);
    if (biolumImg && bW) {
      ctx.drawImage(biolumImg, bx, 0, bW, bH);
      ctx.save(); ctx.translate(bx, 0);
      drawRoisOnCtx(ctx, dW ? bW/dW : 1, dH ? bH/dH : 1);
      ctx.restore();
    }

    const addLabel = (text, x, y) => {
      ctx.font = '10px IBM Plex Mono';
      const tw = ctx.measureText(text).width;
      ctx.fillStyle = 'rgba(6,9,13,0.75)'; ctx.fillRect(x, y, tw+12, 18);
      ctx.fillStyle = '#4af0c4'; ctx.fillText(text, x+6, y+13);
    };
    if (dayImg && dW)    addLabel('DAY',    6,    6);
    if (biolumImg && bW) addLabel('BIOLUM', bx+6, 6);

    return off.toDataURL('image/png');
  }

  function saveSnapshot() {
    const dataURL = captureSnapshotDataURL();
    if (!dataURL) { setStatus('Load images first', 'wrn'); return; }
    const stem = biolumPair ? biolumPair.stem : (dayPair ? dayPair.stem : 'snapshot');
    const a = document.createElement('a');
    a.href = dataURL; a.download = stem + '_snapshot.png';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
    setStatus('Snapshot downloaded: ' + stem + '_snapshot.png', 'ok');
  }

  function drawRoisOnCtx(ctx, sx, sy) {
    rois.forEach((roi, i) => {
      const color = ROI_COLORS[i % ROI_COLORS.length];
      const label = getRoiLabel(roi, i);
      const x = roi.x*sx, y = roi.y*sy, w = roi.w*sx, h = roi.h*sy;
      drawSingleRoi(ctx, x, y, w, h, roi.shape||'rect', color, false, label);
    });
  }

  async function loadRoiFile(path) {
    document.getElementById('roi-dialog').style.display = 'none';
    const r = await fetch('/load_rois', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({path})
    });
    const data = await r.json();
    if (!data.ok) { setStatus('Failed to load ROIs: ' + data.error, 'err'); return; }

    const saved = data.data;
    const scaleX = dayCanvas.width / saved.jpeg_size[0];
    const scaleY = dayCanvas.height / saved.jpeg_size[1];

    rois = saved.rois.map(r => ({
      ...r,
      x: r.x * scaleX, y: r.y * scaleY,
      w: r.w * scaleX, h: r.h * scaleY,
    }));
    ensureRoiNumbers();
    selectedRoi = -1;
    renderRoiSetupTable();
    drawAll();
    scheduleRoiAutosave();
    setStatus(`Loaded ${rois.length} ROIs from ${saved.stem}`, 'ok');
  }
</script>
</body>
</html>"""

# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print()
    print("  Biolum Analyzer")
    print("  ─────────────────────────────────────────")
    print("  Opening browser at http://localhost:5001")
    print("  Press Ctrl+C to quit")
    print("  ─────────────────────────────────────────")
    print()

    def open_browser():
        import time
        time.sleep(1)
        webbrowser.open("http://localhost:5001")

    threading.Thread(target=open_browser, daemon=True).start()

    try:
        app.run(host="127.0.0.1", port=5001, threaded=True)
    except KeyboardInterrupt:
        print("\nBye!")
        sys.exit(0)
